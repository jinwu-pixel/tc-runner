"""Tests for scripts/settings_anchor_gap.py (read-only static decomposition).

Pure parser / classifier functions are tested with synthetic inputs only.
NO Excel / CSV / baseline file IO, NO device, NO openpyxl import required here.
The real analysis IO (load_export_rows / load_excel_procedures / main) is
covered separately with a small golden snapshot.
"""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parent.parent
_PATH = _ROOT / "scripts" / "settings_anchor_gap.py"
_spec = importlib.util.spec_from_file_location("settings_anchor_gap", _PATH)
sag = importlib.util.module_from_spec(_spec)
sys.modules["settings_anchor_gap"] = sag
_spec.loader.exec_module(sag)


# --------------------------------------------------------------------------
# Cycle 1 — parse_menu_path / compute_depth
# --------------------------------------------------------------------------

def test_parse_menu_path_simple_three_levels():
    assert sag.parse_menu_path("1. 설정 > 앱 > 모두 보기") == ["설정", "앱", "모두 보기"]


def test_parse_menu_path_strips_trailing_action_word_진입():
    assert sag.parse_menu_path(" 1.설정 > 네트워크 및 인터넷 진입") == [
        "설정",
        "네트워크 및 인터넷",
    ]


def test_parse_menu_path_strips_trailing_tap_token():
    assert sag.parse_menu_path("1. 설정 > 검색 TAP 하드키") == ["설정", "검색"]


def test_parse_menu_path_multiline_picks_deepest_and_strips_quotes():
    proc = '1. 설정 > 검색\n 2. 설정 > 검색 > "화면" '
    assert sag.parse_menu_path(proc) == ["설정", "검색", "화면"]


def test_parse_menu_path_no_separator_returns_empty():
    assert sag.parse_menu_path("설정 검색") == []


def test_parse_menu_path_none_or_blank_returns_empty():
    assert sag.parse_menu_path(None) == []
    assert sag.parse_menu_path("   ") == []


def test_parse_menu_path_strips_comma_step_prefix():
    assert sag.parse_menu_path("1, 설정 > 알림 읽어주기 > 정각") == [
        "설정",
        "알림 읽어주기",
        "정각",
    ]


def test_parse_menu_path_strips_doubled_step_prefix():
    assert sag.parse_menu_path("1. 1. 설정 > 배경화면 및 스타일") == [
        "설정",
        "배경화면 및 스타일",
    ]


def test_parse_menu_path_prefers_settings_rooted_line_over_deeper_subaction():
    # Real #276: the sub-action line is DEEPER (3 '>') than the menu-root line
    # (2 '>'), so naive deepest-wins picks the wrong root.
    proc = (
        "1. 설정 > 소리 및 진동 > 미디어\n"
        " 1. 미디어 플레이어 고정 On > 음악 재생 > 왼쪽으로 드레그 > 퀵셋팅 하단 확인"
    )
    assert sag.parse_menu_path(proc) == ["설정", "소리 및 진동", "미디어"]


def test_parse_menu_path_falls_back_to_non_rooted_when_no_settings_line():
    path = sag.parse_menu_path("1. 잠금화면 > 화면 하단 shortcut 확인")
    assert path and path[0] == "잠금화면"


def test_compute_depth_counts_levels_below_root():
    assert sag.compute_depth(["설정", "앱", "모두 보기"]) == 2
    assert sag.compute_depth(["설정", "검색"]) == 1
    assert sag.compute_depth([]) == 0


# --------------------------------------------------------------------------
# Cycle 2 — text_input_required / focus_nav_required / mutation_suspected
# --------------------------------------------------------------------------

def test_text_input_required_true_on_keyboard_cue():
    assert sag.text_input_required("1.쿼티키보드 표시 되는지 확인", "READ_ONLY") is True


def test_text_input_required_true_on_csv_safety_flag():
    assert sag.text_input_required("1. 설정 > 앱 > 모두 보기", "INPUT_REQUIRED") is True


def test_text_input_required_false_for_plain_navigation():
    assert sag.text_input_required("1. 설정 > 앱 > 모두 보기", "NAVIGATION_ONLY") is False


def test_focus_nav_required_true_on_hardkey_cue():
    assert sag.focus_nav_required("1. 설정 > 검색 TAP 하드키") is True


def test_focus_nav_required_true_on_direction_focus_cue():
    assert sag.focus_nav_required("방향키로 포커스 이동") is True


def test_focus_nav_required_false_for_plain_tap_navigation():
    assert sag.focus_nav_required("1. 설정 > 앱 > 모두 보기") is False


def test_mutation_suspected_true_on_toggle():
    assert sag.mutation_suspected("위치 사용 토글 Off") is True


def test_mutation_suspected_true_on_add_data():
    assert sag.mutation_suspected("비상 연락처 추가") is True


def test_mutation_suspected_false_for_readonly_navigation():
    assert sag.mutation_suspected("1. 설정 > 앱 > 모두 보기") is False


def test_mutation_suspected_true_from_expected_result_change_verb():
    # #70: procedure has no mutation cue, but the expected result confirms a
    # state change ("이름이 변경된다").
    proc = "1. 설정 > 연결된 기기 > + 새 기기와 페어링 > 기기 이름 TAP"
    expected = "이름 바꾸기 TAP하면 이름이 변경된다."
    assert sag.mutation_suspected(proc, expected) is True


def test_mutation_suspected_false_for_onoff_list_exposure_in_expected():
    # #162: expected merely exposes an ON/OFF list — observation, not mutation.
    proc = "1. 설정 > 알림 > 방해 금지 모드 > 알람 및 기타 방해 요소"
    expected = "하기 메뉴 노출된다.\n- 알람 (ON/OFF)\n- 미디어 소리 (ON/OFF)"
    assert sag.mutation_suspected(proc, expected) is False


def test_mutation_suspected_false_for_menu_exposed_expected():
    # #149: "메뉴 노출됨" is an observation verb, not a state change.
    proc = "1. 설정 > 알림 > 방해금지 모드"
    expected = "하기 메뉴 노출됨\n- 지금 사용 설정 (버튼)"
    assert sag.mutation_suspected(proc, expected) is False


def test_mutation_suspected_true_for_declarative_set_result():
    # #287: declarative "설정된다" — the action sets a value.
    assert sag.mutation_suspected("1. 설정 > 디스플레이 > 밝기", "디스플레이 밝기가 0% 설정된다") is True


def test_mutation_suspected_false_for_state_is_set_observation():
    # #975: "설정되어 있음" describes a pre-existing state, not a mutation.
    exp = "Google 음성 서비스로 설정되어 있음\n- 언어 추가"
    assert sag.mutation_suspected("1. 설정 > 시스템 > 언어", exp) is False


def test_mutation_suspected_false_for_default_set_and_displayed():
    # #337: "설정되어 노출된다" — a default value is set and shown (observation).
    exp = "화면 자동 잠금 시간 1분(Default)으로 설정되어 노출된다"
    assert sag.mutation_suspected("1. 설정 > 디스플레이 > 화면 자동 잠금", exp) is False


def test_mutation_suspected_false_for_adnominal_set_noun():
    # #405/#623: "설정된 시간/값" adnominal — a pre-set thing being observed.
    exp = "시작시간 또는 종료시간이 설정된 시간으로 노출된다"
    assert sag.mutation_suspected("1. 설정 > 디스플레이 > 화면 보호기", exp) is False


def test_mutation_suspected_false_for_screen_transition_navigation():
    # 전환된다 is overwhelmingly screen navigation ("QnA 화면으로 전환된다"), not a
    # persistent state mutation — it must not be flagged.
    exp = "고객센터 QnA 화면으로 전환된다"
    assert sag.mutation_suspected("1. 설정 > 고객센터", exp) is False


# --------------------------------------------------------------------------
# Cycle 3 — build_baseline_index / classify_entry_method / resolve_anchor
# --------------------------------------------------------------------------

_SYNTH_BASELINE = {
    "screens": [
        {
            "screen_id": "settings_d1_location",
            "label_ko": "위치",
            "nav_path": ["설정", "위치"],
            "reach_status": "REACHED",
            "entry": {
                "method": "deeplink",
                "action": "android.settings.LOCATION_SOURCE_SETTINGS",
                "launched_cmd": "am start -a android.settings.LOCATION_SOURCE_SETTINGS",
            },
            "observed_texts": {
                "ko": ["모두 보기", "앱 위치 정보 액세스 권한"],
                "en": [],
                "other": [],
            },
        },
        {
            "screen_id": "settings_d1_apps",
            "label_ko": "앱",
            "nav_path": ["설정", "앱"],
            "reach_status": "REACHED",
            "entry": {
                "method": "deeplink",
                "action": "android.settings.APPLICATION_SETTINGS",
                "launched_cmd": "am start -a android.settings.APPLICATION_SETTINGS",
            },
            "observed_texts": {"ko": ["기본 앱"], "en": [], "other": []},
        },
        {
            "screen_id": "settings_d1_wellbeing",
            "label_ko": "디지털 웰빙 및 자녀 보호 기능",
            "nav_path": ["설정", "디지털 웰빙 및 자녀 보호 기능"],
            "reach_status": "FOCUS_MISMATCH",
            "entry": {"method": "deeplink", "action": "X"},
            "observed_texts": {"ko": [], "en": [], "other": []},
        },
    ]
}


@pytest.fixture
def idx():
    return sag.build_baseline_index(_SYNTH_BASELINE)


def test_build_baseline_index_maps_d1_label_to_screen(idx):
    assert idx["위치"]["screen_id"] == "settings_d1_location"
    assert idx["앱"]["screen_id"] == "settings_d1_apps"
    assert idx["디지털 웰빙 및 자녀 보호 기능"]["screen_id"] == "settings_d1_wellbeing"


def test_resolve_anchor_leaf_label_observed_when_leaf_text_seen_on_d1(idx):
    # Leaf text seen in the d1 dashboard observation is NOT a reached target —
    # only the label was observed.
    r = sag.resolve_anchor(["설정", "위치", "모두 보기"], idx)
    assert r["anchor_state"] == "LEAF_LABEL_OBSERVED"
    assert r["baseline_screen_id"] == "settings_d1_location"
    assert r["confidence"] >= 0.7


def test_resolve_anchor_leaf_label_observed_is_spacing_insensitive(idx):
    r = sag.resolve_anchor(["설정", "위치", "모두보기"], idx)
    assert r["anchor_state"] == "LEAF_LABEL_OBSERVED"


def test_resolve_anchor_partial_when_d1_present_but_leaf_unobserved(idx):
    r = sag.resolve_anchor(["설정", "앱", "모두 보기"], idx)
    assert r["anchor_state"] == "PARTIAL"
    assert r["baseline_screen_id"] == "settings_d1_apps"
    assert r["confidence"] < 0.7


def test_resolve_anchor_missing_when_no_baseline_d1_screen(idx):
    r = sag.resolve_anchor(["설정", "배터리", "절약 모드"], idx)
    assert r["anchor_state"] == "MISSING"
    assert r["baseline_screen_id"] is None


def test_resolve_anchor_depth1_target_reached_when_screen_reached(idx):
    r = sag.resolve_anchor(["설정", "앱"], idx)
    assert r["anchor_state"] == "TARGET_REACHED"


def test_resolve_anchor_depth1_partial_when_screen_focus_mismatch(idx):
    r = sag.resolve_anchor(["설정", "디지털 웰빙 및 자녀 보호 기능"], idx)
    assert r["anchor_state"] == "PARTIAL"


def test_resolve_anchor_unknown_for_empty_path(idx):
    r = sag.resolve_anchor([], idx)
    assert r["anchor_state"] == "UNKNOWN"
    assert r["baseline_screen_id"] is None


def test_classify_entry_method_deeplink_candidate_from_baseline(idx):
    method, evidence = sag.classify_entry_method(
        ["설정", "위치", "모두 보기"], "1. 설정 > 위치 > 모두 보기", idx
    )
    assert method == "DEEPLINK_CANDIDATE"
    assert evidence


def test_classify_entry_method_search_candidate_when_search_in_path(idx):
    method, _ = sag.classify_entry_method(
        ["설정", "검색", "화면"], "1. 설정 > 검색 > 화면", idx
    )
    assert method == "SEARCH_CANDIDATE"


def test_classify_entry_method_hardkey_takes_precedence(idx):
    method, _ = sag.classify_entry_method(
        ["설정", "앱"], "1. 설정 > 앱 TAP 하드키", idx
    )
    assert method == "HARDKEY_NAVIGATION"


def test_classify_entry_method_menu_navigation_when_area_not_in_baseline(idx):
    method, _ = sag.classify_entry_method(
        ["설정", "배터리", "절약 모드"], "1. 설정 > 배터리 > 절약 모드", idx
    )
    assert method == "MENU_NAVIGATION"


def test_classify_entry_method_unknown_for_empty_path(idx):
    method, _ = sag.classify_entry_method([], "주절주절", idx)
    assert method == "UNKNOWN"


# --------------------------------------------------------------------------
# Cycle 4 — recommend_probe / enrich_row (schema lock)
# --------------------------------------------------------------------------

def test_recommend_probe_no_deepen_when_target_reached():
    assert sag.recommend_probe("TARGET_REACHED", False, False) == "NO_ANCHOR_DEEPEN_NEEDED"


def test_recommend_probe_no_deepen_when_leaf_label_observed():
    assert (
        sag.recommend_probe("LEAF_LABEL_OBSERVED", False, False)
        == "NO_ANCHOR_DEEPEN_NEEDED"
    )


def test_recommend_probe_partial_is_high_priority():
    assert sag.recommend_probe("PARTIAL", False, False) == "PROBE_PRIORITY_HIGH"


def test_recommend_probe_missing_is_medium_priority():
    assert sag.recommend_probe("MISSING", False, False) == "PROBE_PRIORITY_MEDIUM"


def test_recommend_probe_does_not_promote_mutation():
    assert sag.recommend_probe("PARTIAL", True, False) == "PROBE_DEFER_MUTATION"


def test_recommend_probe_does_not_promote_input():
    assert sag.recommend_probe("PARTIAL", False, True) == "PROBE_DEFER_INPUT"


def test_recommend_probe_anchored_excluded_regardless_of_mutation():
    assert (
        sag.recommend_probe("TARGET_REACHED", True, False) == "NO_ANCHOR_DEEPEN_NEEDED"
    )


def test_recommend_probe_unknown_routes_to_review():
    assert sag.recommend_probe("UNKNOWN", False, False) == "REVIEW_SOURCE"


def test_enrich_row_end_to_end_resolved(idx):
    csv_row = {
        "excel_tc_id": "23.Settings#827.0",
        "excel_title": "Settings / 위치",
        "source_sheet": "23.Settings",
        "source_row_range": "827",
        "safety_class": "NAVIGATION_ONLY",
    }
    rec = sag.enrich_row(csv_row, "1. 설정 > 위치 > 모두 보기", idx)
    assert rec["tc_id"] == "23.Settings#827.0"
    assert rec["area"] == "위치"
    assert rec["menu_path"] == "설정 > 위치 > 모두 보기"
    assert rec["depth"] == 2
    assert rec["anchor_state"] == "LEAF_LABEL_OBSERVED"
    assert rec["entry_method"] == "DEEPLINK_CANDIDATE"
    assert rec["text_input_required"] is False
    assert rec["recommended_probe"] == "NO_ANCHOR_DEEPEN_NEEDED"


def test_enrich_row_defers_mutation_detected_only_in_expected(idx):
    # #70 shape: PARTIAL anchor, no mutation cue in procedure, but expected
    # result confirms a state change -> must be deferred, not PROBE_PRIORITY.
    csv_row = {
        "excel_tc_id": "23.Settings#70.0",
        "excel_title": "Settings / 연결된 기기",
        "source_sheet": "23.Settings",
        "source_row_range": "70",
        "safety_class": "NAVIGATION_ONLY",
    }
    rec = sag.enrich_row(
        csv_row,
        "1. 설정 > 연결된 기기 > 기기 이름 TAP",
        idx,
        expected_text="이름 바꾸기 TAP하면 이름이 변경된다.",
    )
    assert rec["mutation_suspected"] is True
    assert rec["recommended_probe"] == "PROBE_DEFER_MUTATION"


def test_enrich_row_keys_match_locked_schema(idx):
    csv_row = {
        "excel_tc_id": "23.Settings#1.0",
        "excel_title": "Settings / 설정 검색",
        "source_sheet": "23.Settings",
        "source_row_range": "1",
        "safety_class": "READ_ONLY",
    }
    rec = sag.enrich_row(csv_row, "1. 설정 > 검색", idx)
    assert list(rec.keys()) == list(sag.FIELDNAMES)
    assert len(sag.FIELDNAMES) == 17


# --------------------------------------------------------------------------
# Cycle 5 — IO: join key / CSV filter / Excel continuation / reproducible main
# --------------------------------------------------------------------------

def test_join_key_normalizes_forms():
    assert sag._join_key("23.Settings#827.0") == "827.0"
    assert sag._join_key(827.0) == "827.0"
    assert sag._join_key("827") == "827.0"
    assert sag._join_key(149) == "149.0"


def test_load_export_rows_filters_settings_export(tmp_path):
    csvp = tmp_path / "join.csv"
    csvp.write_text(
        "excel_tc_id,source_sheet,recommended_next_action,excel_title,safety_class,source_row_range\n"
        "23.Settings#1.0,23.Settings,EXPORT_TO_APPIUM,Settings / 검색,READ_ONLY,1\n"
        "23.Settings#2.0,23.Settings,MANUAL_ONLY,Settings / 검색,INPUT_REQUIRED,2\n"
        "25.Call#9.0,25.Call,EXPORT_TO_APPIUM,Call / x,READ_ONLY,9\n",
        encoding="utf-8-sig",
    )
    rows = sag.load_export_rows(str(csvp))
    assert [r["excel_tc_id"] for r in rows] == ["23.Settings#1.0"]


def test_load_excel_texts_concats_continuation_rows_for_proc_and_expected(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "23.Settings"
    ws.append(["TC ID", "ITEM", "Functionality", "Pre", "Test procedure", "Expected"])
    ws.append(["81.0", "Settings", "앱", "", "1. 설정 > 앱 > 모두 보기", "목록 노출"])
    ws.append([None, None, None, None, "2. 설정 > 앱 > 모두 보기 > 정렬", "정렬됨"])  # continuation
    ws.append(["85.0", "Settings", "앱", "", "1. 설정 > 앱 > 기기 사용 시간", "결과"])
    xlsx = tmp_path / "src.xlsx"
    wb.save(str(xlsx))

    texts = sag.load_excel_texts(str(xlsx), sheet="23.Settings")
    assert "정렬" in texts["81.0"]["proc"]  # continuation appended to owning TC
    assert "정렬됨" in texts["81.0"]["expected"]  # expected continuation too
    assert texts["85.0"]["proc"].startswith("1. 설정 > 앱 > 기기 사용 시간")


def test_build_records_are_joined_and_sorted(idx):
    rows = [
        {"excel_tc_id": "23.Settings#85.0", "excel_title": "Settings / 앱",
         "source_sheet": "23.Settings", "source_row_range": "85", "safety_class": "NAVIGATION_ONLY"},
        {"excel_tc_id": "23.Settings#81.0", "excel_title": "Settings / 앱",
         "source_sheet": "23.Settings", "source_row_range": "81", "safety_class": "NAVIGATION_ONLY"},
    ]
    text_map = {
        "85.0": {"proc": "1. 설정 > 앱 > 기기 사용 시간", "expected": "결과"},
        "81.0": {"proc": "1. 설정 > 앱 > 모두 보기", "expected": "목록"},
    }
    recs = sag.build_records(rows, text_map, idx)
    assert [r["tc_id"] for r in recs] == ["23.Settings#81.0", "23.Settings#85.0"]  # sorted
    assert recs[0]["menu_path"] == "설정 > 앱 > 모두 보기"


def test_main_outputs_are_reproducible(tmp_path):
    import json as _json
    import openpyxl

    csvp = tmp_path / "join.csv"
    csvp.write_text(
        "excel_tc_id,source_sheet,recommended_next_action,excel_title,safety_class,source_row_range\n"
        "23.Settings#827.0,23.Settings,EXPORT_TO_APPIUM,Settings / 위치,NAVIGATION_ONLY,827\n",
        encoding="utf-8-sig",
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "23.Settings"
    ws.append(["TC ID", "ITEM", "Functionality", "Pre", "Test procedure", "Expected"])
    ws.append(["827.0", "Settings", "위치", "", "1. 설정 > 위치 > 모두 보기", "결과"])
    xlsx = tmp_path / "src.xlsx"
    wb.save(str(xlsx))
    blp = tmp_path / "baseline.json"
    blp.write_text(_json.dumps(_SYNTH_BASELINE, ensure_ascii=False), encoding="utf-8")

    out_csv = tmp_path / "enriched.csv"
    out_md = tmp_path / "summary.md"
    args = [
        "--csv", str(csvp), "--excel", str(xlsx), "--baseline", str(blp),
        "--out-csv", str(out_csv), "--out-md", str(out_md),
    ]
    assert sag.main(args) == 0
    first_csv = out_csv.read_bytes()
    first_md = out_md.read_bytes()
    assert sag.main(args) == 0
    assert out_csv.read_bytes() == first_csv  # byte-identical rerun
    assert out_md.read_bytes() == first_md
    header = first_csv.decode("utf-8-sig").splitlines()[0]
    assert header.split(",") == list(sag.FIELDNAMES)


def test_main_stdout_reports_new_anchor_states(tmp_path, capsys):
    import json as _json
    import openpyxl

    csvp = tmp_path / "join.csv"
    csvp.write_text(
        "excel_tc_id,source_sheet,recommended_next_action,excel_title,safety_class,source_row_range\n"
        "23.Settings#827.0,23.Settings,EXPORT_TO_APPIUM,Settings / 위치,NAVIGATION_ONLY,827\n",
        encoding="utf-8-sig",
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "23.Settings"
    ws.append(["TC ID", "ITEM", "Functionality", "Pre", "Test procedure", "Expected"])
    ws.append(["827.0", "Settings", "위치", "", "1. 설정 > 위치 > 모두 보기", "목록"])
    xlsx = tmp_path / "src.xlsx"
    wb.save(str(xlsx))
    blp = tmp_path / "baseline.json"
    blp.write_text(_json.dumps(_SYNTH_BASELINE, ensure_ascii=False), encoding="utf-8")

    rc = sag.main([
        "--csv", str(csvp), "--excel", str(xlsx), "--baseline", str(blp),
        "--out-csv", str(tmp_path / "e.csv"), "--out-md", str(tmp_path / "s.md"),
    ])
    assert rc == 0
    out = capsys.readouterr().out
    assert "RESOLVED=" not in out  # stale key removed
    assert "TARGET_REACHED=0" in out
    assert "LEAF_LABEL_OBSERVED=1" in out  # the one synthetic row is leaf-observed
    assert "PARTIAL=0" in out and "MISSING=0" in out and "UNKNOWN=0" in out


# --------------------------------------------------------------------------
# Cycle 6 — golden snapshot over the real audit corpus (skips if absent)
# --------------------------------------------------------------------------

def test_real_corpus_matches_golden_snapshot():
    import collections
    import json as _json

    csvp = _ROOT / sag._DEFAULT_CSV
    excelp = _ROOT / sag._DEFAULT_EXCEL
    blp = _ROOT / sag._DEFAULT_BASELINE
    if not (csvp.exists() and excelp.exists() and blp.exists()):
        pytest.skip("source audit corpus not present on this machine")

    index, run_id = sag.load_baseline_index(str(blp))
    rows = sag.load_export_rows(str(csvp))
    text_map = sag.load_excel_texts(str(excelp))
    matched = sum(1 for r in rows if sag._join_key(r["excel_tc_id"]) in text_map)
    recs = sag.build_records(rows, text_map, index)

    def dist(field):
        return dict(collections.Counter(str(r[field]) for r in recs))

    actual = {
        "baseline_run_id": run_id,
        "population": len(rows),
        "join_matched": matched,
        "anchor_state": dist("anchor_state"),
        "entry_method": dist("entry_method"),
        "recommended_probe": dist("recommended_probe"),
        "depth": {
            str(k): v
            for k, v in sorted(collections.Counter(r["depth"] for r in recs).items())
        },
        "traits": {
            "text_input_required": sum(1 for r in recs if r["text_input_required"]),
            "focus_nav_required": sum(1 for r in recs if r["focus_nav_required"]),
            "mutation_suspected": sum(1 for r in recs if r["mutation_suspected"]),
        },
    }
    golden = _json.loads(
        (_ROOT / "tests/fixtures/anchor/settings_anchor_gap_golden.json").read_text(
            encoding="utf-8"
        )
    )
    assert actual == golden
    assert matched == len(rows)  # join must stay total
