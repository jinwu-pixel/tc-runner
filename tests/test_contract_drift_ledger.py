"""Slice 0.5 ledger + Slice 1a shared-normalizer integration tests.

설계 source: docs/superpowers/specs/2026-07-14-canonical-execution-contract-design.md §6 + §10 Task 1.
Host-only: 단말/subprocess 없음 (adb probe는 subprocess.run을 fake로 patch).
"""
import copy
import hashlib
import importlib.util
import json
from pathlib import Path

import pytest

_REPO = Path(__file__).resolve().parent.parent
_PATH = _REPO / "scripts" / "contract_drift_ledger.py"
_spec = importlib.util.spec_from_file_location("contract_drift_ledger", _PATH)
L = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(L)


@pytest.fixture(scope="module")
def fixtures():
    return L.build_fixture_matrix()


@pytest.fixture(scope="module")
def consumer_rows(fixtures):
    return L.probe_consumers(fixtures)


@pytest.fixture(scope="module")
def producer_rows(fixtures):
    return L.probe_producers(fixtures)


def _rows(rows, **want):
    return [r for r in rows if all(r.get(k) == v for k, v in want.items())]


def _nj(row):
    return json.loads(row["normalized_json"])


def _structural_pair_rows():
    return [
        L._row(
            actor_kind="pair",
            producer=producer,
            consumer=consumer,
            variant=mode,
        )
        for producer in L.PRODUCERS
        for mode in L.PRODUCER_MODES
        for consumer in L.CONSUMERS
    ]


STEP_FAMILIES = [
    "target/text",
    "duration/seconds",
    "key/keycode",
    "x/y vs x1/y1",
    "target/id",
]
VARIANTS = ["canonical_only", "alias_only", "equal_duplicate", "conflicting_duplicate"]


def test_matrix_has_four_consumers_by_two_producers_and_two_modes(producer_rows):
    assert L.PRODUCERS == ["excel", "mmi"]
    assert L.PRODUCER_MODES == ["legacy", "canonical"]
    assert L.CONSUMERS == ["schema", "validate_tc", "tc_loader", "action_runner"]
    pairs = {
        (r["producer"], r["variant"], r["consumer"])
        for r in producer_rows
        if r["actor_kind"] == "pair"
    }
    assert pairs == {
        (producer, mode, consumer)
        for producer in L.PRODUCERS
        for mode in L.PRODUCER_MODES
        for consumer in L.CONSUMERS
    }
    assert len(pairs) == 16
    assert len([r for r in producer_rows if r["actor_kind"] == "pair"]) == 16


def test_current_pair_baseline_is_locked(producer_rows):
    actual = {
        (r["producer"], r["consumer"]): (r["verdict"], r["finding_code"])
        for r in producer_rows
        if r["actor_kind"] == "pair" and r["variant"] == "legacy"
    }
    assert actual == {
        ("excel", "schema"): ("reject", "-"),
        ("excel", "validate_tc"): ("reject", "-"),
        ("excel", "tc_loader"): ("accept", "-"),
        ("excel", "action_runner"): (
            "blocking", "EXCEL_SWIPE_ENDPOINT_MISSING",
        ),
        ("mmi", "schema"): ("reject", "-"),
        ("mmi", "validate_tc"): ("reject", "-"),
        ("mmi", "tc_loader"): ("accept", "-"),
        ("mmi", "action_runner"): (
            "accept", "UNRESOLVED_PLACEHOLDER_EXECUTED",
        ),
    }


def test_canonical_pair_baseline_is_nonblocking(producer_rows):
    actual = {
        (r["producer"], r["consumer"]): (r["verdict"], r["finding_code"])
        for r in producer_rows
        if r["actor_kind"] == "pair" and r["variant"] == "canonical"
    }
    assert actual == {
        ("excel", "schema"): ("accept", "-"),
        ("excel", "validate_tc"): ("accept", "-"),
        ("excel", "tc_loader"): ("accept", "-"),
        ("excel", "action_runner"): ("reject", "-"),
        ("mmi", "schema"): ("accept", "-"),
        ("mmi", "validate_tc"): ("accept", "-"),
        ("mmi", "tc_loader"): ("accept", "-"),
        ("mmi", "action_runner"): ("accept", "-"),
    }
    assert not any(
        row["verdict"] == "blocking"
        for row in producer_rows
        if row["variant"].startswith("canonical")
        or (row["actor_kind"] == "pair" and row["variant"] == "canonical")
    )


def test_seed_alias_pairs_are_all_enumerated(fixtures):
    have = {(f["family"], f["variant"]) for f in fixtures}
    for fam in STEP_FAMILIES:
        for var in VARIANTS:
            assert (fam, var) in have, f"missing fixture: {fam} / {var}"
    ids = {f["fixture_id"] for f in fixtures}
    for required in [
        "input_text_text_canonical",
        "input_text_target_rejected",
        "key_sequence_delay_seconds_observed",
        "screenshot_name",
        "verify_shell_timeout_5000_ms",
        "metadata_runnable_false",
        "metadata_runnable_reason_unresolved",
        "step_compile_status_unresolved",
        "shell_placeholder_brace",
        "excel_swipe_endpoint_missing",
        "shell_rc1_with_stdout",
    ]:
        assert required in ids, f"missing fixture_id: {required}"


def test_corrected_fixture_matrix_has_a_new_digest_version():
    assert L.FIXTURE_VERSION == 4


def test_ledger_uses_shared_contract_functions_directly():
    from src import execution_contract

    assert L.normalize_tc is execution_contract.normalize_tc
    assert L.derive_action_required is execution_contract.derive_action_required
    assert L.validate_canonical_tc is execution_contract.validate_canonical_tc
    assert L.TOOL_VERSION == "contract-drift-ledger-v4"


def test_validator_step_probe_exercises_public_wrapper(fixtures, monkeypatch):
    import validate_tc as validator_module

    fixture = next(
        f for f in fixtures if f["fixture_id"] == "wait_alias_only"
    )
    monkeypatch.setattr(
        validator_module,
        "validate_tc",
        lambda _tc, _schema: ["SENTINEL_WRAPPER_FAILURE"],
    )

    row = L._probe_validator_consumer(fixture, L._load_schema())
    normalized = _nj(row)

    assert row["verdict"] == "blocking"
    assert row["finding_code"] == "VALIDATOR_CORE_DIVERGENCE"
    assert normalized["wrapper_errors"] == ["SENTINEL_WRAPPER_FAILURE"]
    assert normalized["wrapper_core_match"] is False


def test_validator_toplevel_probe_exercises_public_wrapper(fixtures, monkeypatch):
    import validate_tc as validator_module

    fixture = next(f for f in fixtures if f["fixture_id"] == "tcname_alias_only")
    monkeypatch.setattr(
        validator_module,
        "validate_tc",
        lambda _tc, _schema: ["SENTINEL_WRAPPER_FAILURE"],
    )

    rows = L._probe_toplevel_fixture(fixture, L._load_schema())
    row = _rows(rows, consumer="validate_tc")[0]
    normalized = _nj(row)

    assert row["verdict"] == "blocking"
    assert row["finding_code"] == "VALIDATOR_CORE_DIVERGENCE"
    assert normalized["wrapper_errors"] == ["SENTINEL_WRAPPER_FAILURE"]
    assert normalized["wrapper_core_match"] is False


def test_producer_probe_uses_excel_fixture_input(fixtures, monkeypatch):
    from openpyxl import load_workbook
    from src import excel_converter

    custom = copy.deepcopy(fixtures)
    fixture = next(
        f for f in custom if f["fixture_id"] == "excel_swipe_endpoint_missing"
    )
    assert fixture["producer"] == "excel"
    fixture["producer_input"]["parameter1"] = 321
    fixture["producer_input"]["parameter2"] = 654

    converter_inputs = []
    original = excel_converter.convert_excel_to_yaml

    def recording_converter(excel_path, output_dir, **kwargs):
        workbook = load_workbook(excel_path, data_only=True)
        try:
            converter_inputs.extend(
                tuple(cell for cell in row)
                for row in workbook.active.iter_rows(min_row=2, values_only=True)
            )
        finally:
            workbook.close()
        return original(excel_path, output_dir, **kwargs)

    monkeypatch.setattr(excel_converter, "convert_excel_to_yaml", recording_converter)

    rows = L.probe_producers(custom)
    emitted = _rows(
        rows,
        actor_kind="producer",
        producer="excel",
        fixture_id="excel_swipe_endpoint_missing",
        action="swipe",
        variant="legacy_emission",
    )
    assert len(emitted) == 1
    assert _nj(emitted[0])["x1"] == 321
    assert _nj(emitted[0])["y1"] == 654
    assert ("EXCEL_TC1", 4, "swipe", 321, 654, None) in converter_inputs


def test_mmi_unresolved_shell_is_emitted_by_compiler(producer_rows):
    emitted = _rows(
        producer_rows,
        actor_kind="producer",
        producer="mmi",
        fixture_id="mmi_unresolved_shell_compiler",
        action="shell",
        variant="legacy_emission",
    )
    assert len(emitted) == 1
    step = _nj(emitted[0])
    assert step["compile_status"] == "UNRESOLVED_PARAMS"
    assert step["_unresolved_params"] == ["package"]

    doc_row = _rows(
        producer_rows,
        actor_kind="producer",
        producer="mmi",
        variant="legacy_document",
    )
    assert len(doc_row) == 1
    warnings = _nj(doc_row[0])["metadata_warnings"]
    assert any(w.startswith("unresolved_params:") for w in warnings)


def test_mmi_probe_calls_compile_classified_with_injected_map(fixtures, monkeypatch):
    from src.mmi_converter.compiler import TCRunnerCompiler

    calls = []
    original = TCRunnerCompiler.compile_classified

    def recording_compile(self, classified):
        result = original(self, classified)
        calls.append({
            "target": classified.intent.target,
            "execution_mode": classified.execution_mode,
            "map_type": type(self._shell_map).__name__,
            "steps": copy.deepcopy(result[0]),
        })
        return result

    monkeypatch.setattr(TCRunnerCompiler, "compile_classified", recording_compile)
    L.probe_producers(fixtures)

    probe_calls = [
        call for call in calls
        if call["target"] == "__ledger_unresolved_package__"
    ]
    assert len(probe_calls) == 1
    assert probe_calls[0]["execution_mode"] == "SHELL_AUTO"
    assert probe_calls[0]["map_type"] == "_ProbeShellActionMap"
    assert probe_calls[0]["steps"][0]["compile_status"] == "UNRESOLVED_PARAMS"


def test_producer_probe_invokes_actual_legacy_and_canonical_modes(
    fixtures, monkeypatch,
):
    from src import excel_converter
    from src.mmi_converter.compiler import TCRunnerCompiler
    from src.mmi_converter.exporter import YAMLExporter

    excel_modes = []
    compiler_modes = []
    exporter_modes = []
    original_excel = excel_converter.convert_excel_to_yaml
    original_compiler_init = TCRunnerCompiler.__init__
    original_exporter_init = YAMLExporter.__init__

    def recording_excel(*args, **kwargs):
        excel_modes.append(kwargs.get("contract_mode", "legacy"))
        return original_excel(*args, **kwargs)

    def recording_compiler_init(self, *args, **kwargs):
        compiler_modes.append(kwargs.get("contract_mode", "legacy"))
        original_compiler_init(self, *args, **kwargs)

    def recording_exporter_init(self, *args, **kwargs):
        exporter_modes.append(kwargs.get("contract_mode", "legacy"))
        original_exporter_init(self, *args, **kwargs)

    monkeypatch.setattr(excel_converter, "convert_excel_to_yaml", recording_excel)
    monkeypatch.setattr(TCRunnerCompiler, "__init__", recording_compiler_init)
    monkeypatch.setattr(YAMLExporter, "__init__", recording_exporter_init)

    L.probe_producers(fixtures)

    assert set(excel_modes) == {"legacy", "canonical"}
    assert set(compiler_modes) == {"legacy", "canonical"}
    assert set(exporter_modes) == {"legacy", "canonical"}


def test_producer_schema_pair_checks_metadata_and_step_rules(producer_rows):
    excel = _rows(
        producer_rows,
        actor_kind="pair",
        producer="excel",
        consumer="schema",
        variant="legacy",
    )
    assert len(excel) == 1
    excel_schema = _nj(excel[0])
    assert excel_schema["metadata_missing"] == [
        "execution_type", "manual_detail", "runnable", "tc_class",
    ]
    excel_steps = {
        item["action"]: item["missing"]
        for item in excel_schema["step_rule_violations"]
    }
    assert excel_steps["tap_text"] == ["target"]
    assert excel_steps["swipe"] == ["x", "x2", "y", "y2"]
    assert excel_steps["key"] == ["key"]
    assert excel_steps["wait"] == ["duration"]

    mmi = _rows(
        producer_rows,
        actor_kind="pair",
        producer="mmi",
        consumer="schema",
        variant="legacy",
    )
    assert len(mmi) == 1
    mmi_schema = _nj(mmi[0])
    assert mmi_schema["metadata_missing"] == [
        "execution_type", "manual_detail", "tc_class",
    ]
    assert any(
        item["action"] == "verify_text" and item["missing"] == ["target"]
        for item in mmi_schema["step_rule_violations"]
    )


def _canonical_wait_doc():
    return {
        "tc_name": "SCHEMA_PROBE",
        "metadata": {
            "runnable": True,
            "tc_class": "FULL_AUTO",
            "execution_type": "AUTO",
            "manual_detail": "NONE",
        },
        "steps": [{"action": "wait", "duration": 1000}],
    }


@pytest.mark.parametrize(
    ("case", "expected_path", "expected_keyword"),
    [
        ("metadata_runnable_string", "/metadata/runnable", "type"),
        ("wait_duration_missing", "/steps/0/duration", "required"),
        ("wait_duration_string", "/steps/0/duration", "type"),
        ("top_level_extra_name", "/name", "additionalProperties"),
    ],
)
def test_schema_validation_uses_current_document_contract(
    case, expected_path, expected_keyword,
):
    doc = _canonical_wait_doc()
    if case == "metadata_runnable_string":
        doc["metadata"]["runnable"] = "true"
    elif case == "wait_duration_missing":
        del doc["steps"][0]["duration"]
    elif case == "wait_duration_string":
        doc["steps"][0]["duration"] = "1000"
    elif case == "top_level_extra_name":
        doc["name"] = "LEGACY"

    violations = L.validate_schema_instance(doc, L._load_schema())
    observed = {(v["path"], v["keyword"]) for v in violations}
    assert (expected_path, expected_keyword) in observed


def test_schema_validation_preserves_metadata_and_step_extra_openness():
    doc = _canonical_wait_doc()
    doc["metadata"]["future_metadata"] = "allowed"
    doc["steps"][0]["future_step"] = "allowed"
    assert L.validate_schema_instance(doc, L._load_schema()) == []


def test_input_text_text_is_not_selector_alias(consumer_rows):
    # §6.2 필수 RED fixture: input_text.text 는 입력값 payload — selector alias 규칙 적용 금지
    rows = _rows(
        consumer_rows, fixture_id="input_text_text_canonical", consumer="action_runner"
    )
    assert len(rows) == 1
    nj = _nj(rows[0])
    assert nj["passed"] is True
    assert ["input_text", "hello"] in nj["calls"]
    assert rows[0]["finding_code"] == "INPUT_TEXT_VALUE_NOT_ALIAS"
    assert rows[0]["verdict"] != "blocking"
    # 대조: input_text 에서 'target' 은 입력값으로 수용되지 않는다 (selector alias 부재 입증)
    neg = _rows(
        consumer_rows, fixture_id="input_text_target_rejected", consumer="action_runner"
    )
    assert len(neg) == 1
    njn = _nj(neg[0])
    assert njn["passed"] is False
    assert "'text'" in njn["message"]
    assert neg[0]["verdict"] == "reject"


def test_key_sequence_delay_seconds_is_observed_not_normalized(consumer_rows):
    rows = _rows(
        consumer_rows,
        fixture_id="key_sequence_delay_seconds_observed",
        consumer="action_runner",
    )
    assert len(rows) == 1
    nj = _nj(rows[0])
    assert nj["sleeps"] == [7, 7]  # time.sleep(delay) 그대로 seconds — ms 변환 없음
    assert rows[0]["unit"] == "s"
    assert rows[0]["verdict"] == "observed"
    assert rows[0]["finding_code"] == "KEY_SEQUENCE_DELAY_SECONDS_OBSERVED"


def test_equal_duplicate_and_conflict_are_distinct(consumer_rows):
    eq = _rows(
        consumer_rows, fixture_id="wait_equal_duplicate", consumer="action_runner"
    )
    cf = _rows(
        consumer_rows, fixture_id="wait_conflicting_duplicate", consumer="action_runner"
    )
    assert len(eq) == 1 and len(cf) == 1
    # seconds(2) == duration(2000ms) → 일관 / seconds(5) != duration(2000ms) → alias 가 canonical 을 가림
    assert _nj(eq[0])["sleeps"] == [2]
    assert _nj(cf[0])["sleeps"] == [5]
    assert eq[0]["finding_code"] == "EQUAL_DUPLICATE_CONSISTENT"
    assert cf[0]["finding_code"] == "CONFLICT_ALIAS_SHADOWS_CANONICAL"
    assert eq[0]["finding_code"] != cf[0]["finding_code"]
    assert eq[0]["verdict"] != "blocking"
    assert cf[0]["verdict"] != "blocking"


@pytest.mark.parametrize(
    "fixture_id",
    [
        "tap_text_alias_only",
        "wait_alias_only",
        "key_alias_only",
        "swipe_alias_only",
        "tap_id_alias_only",
    ],
)
def test_validator_rows_record_shared_alias_normalization(consumer_rows, fixture_id):
    row = _rows(
        consumer_rows, fixture_id=fixture_id, consumer="validate_tc"
    )[0]
    normalized = _nj(row)

    assert row["verdict"] == "accept"
    assert row["finding_code"] == "ALIAS_NORMALIZED"
    assert normalized["contract_findings"][0]["code"] == "ALIAS_NORMALIZED"
    assert normalized["canonical"]["steps"][0] != next(
        f["step"] for f in L.build_fixture_matrix() if f["fixture_id"] == fixture_id
    )


@pytest.mark.parametrize(
    "fixture_id",
    [
        "tap_text_equal_duplicate",
        "wait_equal_duplicate",
        "key_equal_duplicate",
        "swipe_equal_duplicate",
        "tap_id_equal_duplicate",
        "tcname_equal_duplicate",
    ],
)
def test_validator_rows_record_nonblocking_alias_duplicates(consumer_rows, fixture_id):
    row = _rows(
        consumer_rows, fixture_id=fixture_id, consumer="validate_tc"
    )[0]

    assert row["verdict"] == "accept"
    assert row["finding_code"] == "ALIAS_DUPLICATE"
    assert _nj(row)["contract_findings"][0]["code"] == "ALIAS_DUPLICATE"


@pytest.mark.parametrize(
    "fixture_id",
    [
        "tap_text_conflicting_duplicate",
        "wait_conflicting_duplicate",
        "key_conflicting_duplicate",
        "swipe_conflicting_duplicate",
        "tap_id_conflicting_duplicate",
        "tcname_conflicting_duplicate",
    ],
)
def test_validator_rows_make_alias_conflicts_blocking(consumer_rows, fixture_id):
    row = _rows(
        consumer_rows, fixture_id=fixture_id, consumer="validate_tc"
    )[0]

    assert row["verdict"] == "blocking"
    assert row["finding_code"] == "ALIAS_CONFLICT"
    assert _nj(row)["contract_findings"][0]["severity"] == "ERROR"


def test_producer_documents_block_on_canonical_metadata_after_normalization(
    producer_rows,
):
    document_rows = _rows(
        producer_rows,
        actor_kind="producer",
        variant="legacy_document",
    )
    assert len(document_rows) == 2

    by_producer = {row["producer"]: row for row in document_rows}
    assert set(by_producer) == {"excel", "mmi"}
    for row in document_rows:
        assert row["verdict"] == "blocking"
        assert row["finding_code"] == "PRODUCER_DOC_NONCANONICAL"
        normalized = _nj(row)
        assert normalized["canonical_validation_errors"]
        assert normalized["canonical"]["tc_name"] == row["tc_name"]

    assert by_producer["excel"] and _nj(by_producer["excel"])[
        "missing_canonical_metadata"
    ] == ["execution_type", "manual_detail", "runnable", "tc_class"]
    assert _nj(by_producer["mmi"])["missing_canonical_metadata"] == [
        "execution_type",
        "manual_detail",
    ]


def test_canonical_producer_documents_are_clean(producer_rows):
    document_rows = _rows(
        producer_rows,
        actor_kind="producer",
        variant="canonical_document",
    )
    assert len(document_rows) == 2
    assert {row["producer"] for row in document_rows} == {"excel", "mmi"}
    for row in document_rows:
        normalized = _nj(row)
        assert row["verdict"] == "observed"
        assert row["finding_code"] == "-"
        assert normalized["contract_findings"] == []
        assert normalized["canonical_validation_errors"] == []
        assert normalized["missing_canonical_metadata"] == []


def test_mmi_exported_at_is_excluded_from_deterministic_comparison():
    assert hasattr(L, "_deterministic_document_view")
    first = {
        "tc_name": "MMI_TIME",
        "metadata": {"exported_at": "2026-07-21T12:00:00", "runnable": True},
        "steps": [{"action": "wait", "duration": 1}],
    }
    second = copy.deepcopy(first)
    second["metadata"]["exported_at"] = "2026-07-21T12:01:00"

    assert L._deterministic_document_view(first) == (
        L._deterministic_document_view(second)
    )
    assert first["metadata"]["exported_at"] == "2026-07-21T12:00:00"


def test_excel_swipe_missing_endpoint_is_blocking(producer_rows):
    emit = _rows(
        producer_rows,
        actor_kind="producer",
        producer="excel",
        action="swipe",
        variant="legacy_emission",
    )
    assert len(emit) == 1
    observed = json.loads(emit[0]["observed_fields"])
    assert "x1" in observed and "y1" in observed
    assert "x2" not in observed and "y2" not in observed
    assert emit[0]["verdict"] == "blocking"
    assert emit[0]["finding_code"] == "EXCEL_SWIPE_ENDPOINT_MISSING"
    pair = _rows(
        producer_rows,
        actor_kind="pair",
        producer="excel",
        consumer="action_runner",
        variant="legacy",
    )
    assert len(pair) == 1
    assert pair[0]["verdict"] == "blocking"
    assert pair[0]["finding_code"] == "EXCEL_SWIPE_ENDPOINT_MISSING"


def test_shell_nonzero_with_stdout_is_blocking(consumer_rows):
    adb_rows = _rows(consumer_rows, fixture_id="shell_rc1_with_stdout", actor="adb")
    assert len(adb_rows) == 1
    nj = _nj(adb_rows[0])
    assert nj["returncode"] == 1
    assert nj["stdout"] != ""
    assert nj["shell_returned_stdout_only"] is True
    assert adb_rows[0]["verdict"] == "blocking"
    assert adb_rows[0]["finding_code"] == "SHELL_RC_DISCARDED"
    runner_rows = _rows(
        consumer_rows,
        fixture_id="shell_rc1_with_stdout",
        actor="action_runner",
        variant="special",
    )
    assert len(runner_rows) == 1
    assert _nj(runner_rows[0])["passed"] is True  # rc=1 인데도 step PASS — 확정 결함
    assert runner_rows[0]["verdict"] == "blocking"
    assert runner_rows[0]["finding_code"] == "SHELL_RC_DISCARDED"


def test_canonical_runner_probes_record_task4_fixed_candidates(consumer_rows):
    shell_row = _rows(
        consumer_rows,
        fixture_id="shell_rc1_with_stdout",
        actor="action_runner",
        variant="canonical_runner",
    )
    assert len(shell_row) == 1
    assert _nj(shell_row[0])["passed"] is False
    assert shell_row[0]["verdict"] == "observed"
    assert shell_row[0]["finding_code"] == "SHELL_NONZERO_REJECTED"

    timeout_row = _rows(
        consumer_rows,
        fixture_id="verify_shell_timeout_5000_ms",
        actor="action_runner",
        variant="canonical_runner",
    )
    assert len(timeout_row) == 1
    timeout_observation = _nj(timeout_row[0])
    assert timeout_observation["passed"] is True
    assert ["shell_result", "getprop ro.build.type", 5.0] in (
        timeout_observation["calls"]
    )
    assert timeout_row[0]["finding_code"] == (
        "TIMEOUT_MS_CONVERTED_AT_ADB_BOUNDARY"
    )

    tap_row = _rows(
        consumer_rows,
        fixture_id="tap_id_canonical_only",
        actor="action_runner",
        variant="canonical_runner",
    )
    assert len(tap_row) == 1
    tap_observation = _nj(tap_row[0])
    assert tap_observation["passed"] is True
    assert tap_observation["lookups"] == [["id", "com.x:id/btn"]]
    assert tap_row[0]["finding_code"] == "CANONICAL_ACCEPTED_BY_RUNNER"


def test_current_blocking_baseline_is_locked(fixtures):
    rows = L.probe_consumers(fixtures) + L.probe_producers(fixtures)
    actual = {}
    for row in rows:
        if row["verdict"] == "blocking":
            actual[row["finding_code"]] = actual.get(row["finding_code"], 0) + 1
    assert actual == {
        "ALIAS_CONFLICT": 6,
        "EXCEL_SWIPE_ENDPOINT_MISSING": 2,
        "PRODUCER_DOC_NONCANONICAL": 2,
        "SHELL_RC_DISCARDED": 2,
    }


def test_corpus_counts_are_3_25_2_1():
    rows = L.scan_corpora()
    counts = {
        r["corpus"]: json.loads(r["normalized_json"])["file_count"]
        for r in rows
        if r["variant"] == "group_count"
    }
    primary = (
        counts["golden_tc_set"],
        counts["exported_tc1"],
        counts["thor2j_settings_smoke"],
        counts["tc_samples_legacy"],
    )
    assert primary == (3, 25, 2, 1)
    assert counts["thor2k_settings_smoke"] == 0  # informational only (§6.4)
    # 파일 row 의 경로는 정렬 상태여야 한다
    for group in ("golden_tc_set", "exported_tc1"):
        paths = [
            r["source_path"]
            for r in rows
            if r["corpus"] == group and r["variant"] == "file"
        ]
        assert paths == sorted(paths)


def test_current_corpus_classification_baseline_is_locked():
    rows = [
        row for row in L.scan_corpora()
        if row["actor_kind"] == "corpus" and row["variant"] == "file"
    ]
    actual = {}
    for row in rows:
        corpus = actual.setdefault(row["corpus"], {
            "files": 0,
            "verdicts": {},
            "findings": {},
            "metadata_files": 0,
            "screenshot_name_hits": 0,
            "canonical": {},
            "legacy": {},
        })
        corpus["files"] += 1
        corpus["verdicts"][row["verdict"]] = (
            corpus["verdicts"].get(row["verdict"], 0) + 1
        )
        corpus["findings"][row["finding_code"]] = (
            corpus["findings"].get(row["finding_code"], 0) + 1
        )
        fields = json.loads(row["observed_fields"])
        for dialect in ("canonical", "legacy"):
            for field, count in fields[dialect].items():
                corpus[dialect][field] = corpus[dialect].get(field, 0) + count
        normalized = json.loads(row["normalized_json"])
        corpus["metadata_files"] += int(normalized["has_metadata"])
        corpus["screenshot_name_hits"] += int(normalized["screenshot_name"])

    legacy_rows = [row for row in rows if row["corpus"] == "tc_samples_legacy"]
    assert len(legacy_rows) == 1
    assert json.loads(legacy_rows[0]["normalized_json"])["top_level"] == "name"

    assert actual == {
        "exported_tc1": {
            "files": 25,
            "verdicts": {"canonical": 25},
            "findings": {"-": 25},
            "metadata_files": 25,
            "screenshot_name_hits": 87,
            "canonical": {
                "duration": 283, "key": 61, "target": 127,
                "x": 55, "x2": 54, "y": 55, "y2": 54,
            },
            "legacy": {},
        },
        "golden_tc_set": {
            "files": 3,
            "verdicts": {"canonical": 3},
            "findings": {"-": 3},
            "metadata_files": 3,
            "screenshot_name_hits": 11,
            "canonical": {"duration": 13, "key": 5, "target": 4},
            "legacy": {},
        },
        "tc_samples_legacy": {
            "files": 1,
            "verdicts": {"legacy": 1},
            "findings": {"LEGACY_DIALECT_FILE": 1},
            "metadata_files": 0,
            "screenshot_name_hits": 1,
            "canonical": {},
            "legacy": {"keycode": 2, "seconds": 1},
        },
        "thor2j_settings_smoke": {
            "files": 2,
            "verdicts": {"canonical": 2},
            "findings": {"-": 2},
            "metadata_files": 2,
            "screenshot_name_hits": 3,
            "canonical": {"duration": 3, "target": 11},
            "legacy": {},
        },
    }


def test_self_check_enforces_only_exact_pair_structure():
    rows = _structural_pair_rows()
    rows.extend([
        L._row(
            actor_kind="corpus",
            corpus="thor2k_settings_smoke",
            variant="group_count",
            normalized_json=json.dumps({"file_count": 7, "primary": False}),
        ),
        L._row(
            actor_kind="consumer",
            verdict="observed",
            finding_code="SHELL_RC_PRESERVED",
        ),
    ])
    assert L._self_check(rows) == []

    duplicate = rows + [
        L._row(
            actor_kind="pair",
            producer="excel",
            consumer="schema",
            variant="legacy",
        )
    ]
    assert any("duplicate" in problem for problem in L._self_check(duplicate))

    missing = [
        row for row in rows
        if not (
            row["actor_kind"] == "pair"
            and row["producer"] == "mmi"
            and row["consumer"] == "action_runner"
            and row["variant"] == "canonical"
        )
    ]
    assert any("pair groups" in problem for problem in L._self_check(missing))


def test_snapshot_inputs_records_sha256_and_mtime_ns():
    snapshot = L.snapshot_inputs()
    expected_actor_paths = {
        "tc_step_schema.json",
        "validate_tc.py",
        "src/execution_contract.py",
        "src/tc_loader.py",
        "src/action_runner.py",
        "src/adb.py",
        "src/excel_converter.py",
        "src/mmi_converter/compiler.py",
        "src/mmi_converter/exporter.py",
    }
    declared_actor_paths = {
        path for paths in L.ACTOR_SOURCE_FILES.values() for path in paths
    }
    assert declared_actor_paths == expected_actor_paths
    expected_corpus_paths = {
        row["source_path"] for row in L.scan_corpora()
        if row["actor_kind"] == "corpus" and row["variant"] == "file"
    }
    assert set(snapshot) == expected_actor_paths | expected_corpus_paths
    for rel, state in snapshot.items():
        path = L.REPO_ROOT / rel
        assert state == {
            "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            "mtime_ns": path.stat().st_mtime_ns,
        }


@pytest.mark.parametrize("changed_field", ["sha256", "mtime_ns"])
def test_main_does_not_write_outputs_when_input_snapshot_changes(
    tmp_path, monkeypatch, changed_field,
):
    before = L.snapshot_inputs()
    after = copy.deepcopy(before)
    first_rel = sorted(after)[0]
    if changed_field == "sha256":
        after[first_rel][changed_field] = "0" * 64
    else:
        after[first_rel][changed_field] += 1
    snapshot_calls = 0

    def changing_snapshot():
        nonlocal snapshot_calls
        snapshot_calls += 1
        return before if snapshot_calls == 1 else after

    monkeypatch.setattr(L, "snapshot_inputs", changing_snapshot)
    monkeypatch.setattr(L, "_build_all_rows", _structural_pair_rows)

    out_dir = tmp_path / "out"
    assert L.main(["--out-dir", str(out_dir)]) == 3
    assert not out_dir.exists()


def test_main_does_not_write_outputs_when_pair_structure_is_invalid(
    tmp_path, monkeypatch,
):
    snapshot = L.snapshot_inputs()
    rows = _structural_pair_rows()[:-1]
    monkeypatch.setattr(L, "snapshot_inputs", lambda: snapshot)
    monkeypatch.setattr(L, "_build_all_rows", lambda: rows)

    out_dir = tmp_path / "out"
    assert L.main(["--out-dir", str(out_dir)]) == 3
    assert not out_dir.exists()


def test_main_does_not_write_outputs_when_determinism_check_fails(
    tmp_path, monkeypatch,
):
    snapshot = L.snapshot_inputs()
    first = _structural_pair_rows()
    second = copy.deepcopy(first)
    second[0]["fixture_id"] = "changed-on-second-build"
    builds = iter([first, second])
    monkeypatch.setattr(L, "snapshot_inputs", lambda: snapshot)
    monkeypatch.setattr(L, "_build_all_rows", lambda: next(builds))

    out_dir = tmp_path / "out"
    assert L.main([
        "--out-dir", str(out_dir), "--verify-determinism",
    ]) == 3
    assert not out_dir.exists()


def test_main_does_not_write_outputs_when_only_summary_is_nondeterministic(
    tmp_path, monkeypatch,
):
    snapshot = L.snapshot_inputs()
    rows = _structural_pair_rows()
    summaries = iter([b"summary-first", b"summary-second"])
    monkeypatch.setattr(L, "snapshot_inputs", lambda: snapshot)
    monkeypatch.setattr(L, "_build_all_rows", lambda: copy.deepcopy(rows))
    monkeypatch.setattr(L, "render_summary_bytes", lambda *args: next(summaries))

    out_dir = tmp_path / "out"
    assert L.main([
        "--out-dir", str(out_dir), "--verify-determinism",
    ]) == 3
    assert not out_dir.exists()


def test_outputs_are_byte_deterministic(tmp_path):
    out1 = tmp_path / "run1"
    out2 = tmp_path / "run2"
    assert L.main([
        "--out-dir", str(out1), "--verify-determinism",
    ]) == 0
    assert L.main(["--out-dir", str(out2)]) == 0
    d1 = [p for p in out1.iterdir() if p.is_dir()]
    d2 = [p for p in out2.iterdir() if p.is_dir()]
    assert len(d1) == 1 and len(d2) == 1
    assert d1[0].name == d2[0].name  # 동일 input digest
    csv1 = (d1[0] / "contract_drift_matrix.csv").read_bytes()
    csv2 = (d2[0] / "contract_drift_matrix.csv").read_bytes()
    assert csv1 == csv2
    assert (d1[0] / "SUMMARY.md").read_bytes() == (d2[0] / "SUMMARY.md").read_bytes()


def test_fail_on_blocking_returns_one_after_writing_evidence(tmp_path):
    out_dir = tmp_path / "out"
    assert L.main([
        "--out-dir", str(out_dir), "--fail-on-blocking",
    ]) == 1
    evidence_dirs = [path for path in out_dir.iterdir() if path.is_dir()]
    assert len(evidence_dirs) == 1
    assert (evidence_dirs[0] / "contract_drift_matrix.csv").is_file()
    assert (evidence_dirs[0] / "SUMMARY.md").is_file()


def test_summary_render_requires_a_captured_input_snapshot(fixtures):
    rows = L.probe_consumers(fixtures)
    rows.extend(L.probe_producers(fixtures))
    rows.extend(L.scan_corpora())
    csv_bytes = L.render_csv_bytes(rows)
    with pytest.raises(TypeError):
        L.render_summary_bytes(rows, "a" * 64, csv_bytes)


def test_summary_reports_expected_vs_observed_without_gating(fixtures):
    rows = L.probe_consumers(fixtures)
    rows.extend(L.probe_producers(fixtures))
    rows.extend(L.scan_corpora())
    csv_bytes = L.render_csv_bytes(rows)
    snapshot = L.snapshot_inputs()
    summary = L.render_summary_bytes(
        rows, "a" * 64, csv_bytes, snapshot,
    ).decode("utf-8")
    assert "acceptance matrix (16 pairs)" in summary
    for producer_mode in (
        "excel/legacy",
        "excel/canonical",
        "mmi/legacy",
        "mmi/canonical",
    ):
        assert f"| {producer_mode} |" in summary
    assert "primary counts observed:" in summary
    assert "baseline expected:" in summary
    assert "THOR2_K target count observed (informational only):" in summary
    assert "blocking codes observed:" in summary
    assert (
        '- canonical runner observations: '
        '`{"CANONICAL_ACCEPTED_BY_RUNNER":1,"SHELL_NONZERO_REJECTED":1,'
        '"TIMEOUT_MS_CONVERTED_AT_ADB_BOUNDARY":1}`'
    ) in summary
    assert (
        '- blocking baseline expected: '
        '`{"ALIAS_CONFLICT":6,"EXCEL_SWIPE_ENDPOINT_MISSING":2,'
        '"PRODUCER_DOC_NONCANONICAL":2,"SHELL_RC_DISCARDED":2}`'
    ) in summary
    assert (
        '- confirmed-defect baseline expected: '
        '`{"EXCEL_SWIPE_ENDPOINT_MISSING":2,"SHELL_RC_DISCARDED":2}`'
    ) in summary
    assert (
        f"contract_drift_matrix.csv sha256: "
        f"`{hashlib.sha256(csv_bytes).hexdigest()}`"
    ) in summary
    assert "self-check: PASS (structural invariants only" in summary
    assert "- [x] `id`→`target` (tap_id) + runner canonical 수용 정렬" in summary
    assert "- [x] `verify_shell.timeout` ms→s 변환을 ADB 경계에서 수행" in summary
    assert "- [x] shell returncode/stderr 구조화 반환" in summary

    changed = copy.deepcopy(rows)
    for row in changed:
        if row["actor_kind"] == "corpus" and row["variant"] == "group_count":
            observed = json.loads(row["normalized_json"])
            if row["corpus"] == "golden_tc_set":
                observed["file_count"] = 99
            elif row["corpus"] == "thor2k_settings_smoke":
                observed["file_count"] = 7
            row["normalized_json"] = json.dumps(observed)
        if row["verdict"] == "blocking":
            row["verdict"] = "observed"
    changed_csv = L.render_csv_bytes(changed)
    changed_summary = L.render_summary_bytes(
        changed, "a" * 64, changed_csv, snapshot,
    ).decode("utf-8")
    assert "primary counts observed: (99, 25, 2, 1)" in changed_summary
    assert "baseline expected: (3, 25, 2, 1)" in changed_summary
    assert "THOR2_K target count observed (informational only): 7" in changed_summary
    assert "blocking codes observed: `{}`" in changed_summary
    assert "confirmed defect" not in changed_summary
    assert "confirmed defect 1" not in changed_summary
    assert "confirmed defect 2" not in changed_summary
    assert "baseline defect 1 not observed (fixed candidate)" in changed_summary
    assert "baseline defect 2 not observed (fixed candidate)" in changed_summary


def test_source_hashes_unchanged_after_scan(tmp_path):
    inputs_before = L.snapshot_inputs()
    digest_before = L.compute_input_digest(inputs_before)
    assert L.main(["--out-dir", str(tmp_path / "out")]) == 0
    inputs_after = L.snapshot_inputs()
    digest_after = L.compute_input_digest(inputs_after)
    assert inputs_before == inputs_after
    assert digest_before == digest_after
