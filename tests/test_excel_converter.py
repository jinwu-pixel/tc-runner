import json
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from src.excel_converter import convert_excel_to_yaml
from src.execution_contract import validate_canonical_tc


CANONICAL_METADATA = {
    "runnable": True,
    "tc_class": "FULL_AUTO",
    "execution_type": "AUTO",
    "manual_detail": "NONE",
}


def create_test_excel(path: Path) -> None:
    """표준 템플릿 형식의 테스트 엑셀 파일을 생성한다."""
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["Wi-Fi 테스트", 1, "tap_text", "설정", "", ""])
    ws.append(["Wi-Fi 테스트", 2, "tap_text", "Wi-Fi", "", ""])
    ws.append(["Wi-Fi 테스트", 3, "wait", "2", "", ""])
    ws.append(["Wi-Fi 테스트", 4, "verify_text", "연결됨", "", ""])
    ws.append(["콜 테스트", 1, "tap_text", "전화", "", ""])
    ws.append(["콜 테스트", 2, "input_text", "01012345678", "", ""])
    ws.append(["콜 테스트", 3, "tap_id", "com.phone:id/call_btn", "", ""])
    ws.append(["콜 테스트", 4, "verify_shell", "dumpsys telephony.registry", "", "mCallState=2"])
    wb.save(path)


def test_convert_excel_to_yaml(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    create_test_excel(xlsx)
    output_dir = tmp_path / "output"

    files = convert_excel_to_yaml(xlsx, output_dir)

    assert len(files) == 2

    # Wi-Fi 테스트 파일 확인
    wifi_file = output_dir / "Wi-Fi 테스트.yaml"
    assert wifi_file.exists()
    with open(wifi_file, "r", encoding="utf-8") as f:
        tc = yaml.safe_load(f)
    assert tc["name"] == "Wi-Fi 테스트"
    assert len(tc["steps"]) == 4
    assert tc["steps"][0] == {"action": "tap_text", "text": "설정"}
    assert tc["steps"][2] == {"action": "wait", "seconds": 2}

    # 콜 테스트 파일 확인
    call_file = output_dir / "콜 테스트.yaml"
    assert call_file.exists()
    with open(call_file, "r", encoding="utf-8") as f:
        tc = yaml.safe_load(f)
    assert tc["name"] == "콜 테스트"
    assert len(tc["steps"]) == 4
    assert tc["steps"][3] == {
        "action": "verify_shell",
        "command": "dumpsys telephony.registry",
        "expected": "mCallState=2",
    }


def test_canonical_excel_emits_target_duration_key_and_tc_name(tmp_path):
    xlsx = tmp_path / "canonical.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_CANONICAL", 1, "tap_text", "Settings", "", ""])
    ws.append(["EXCEL_CANONICAL", 2, "tap_id", "com.app:id/ok", "", ""])
    ws.append(["EXCEL_CANONICAL", 3, "wait", "2", "", ""])
    ws.append(["EXCEL_CANONICAL", 4, "key", "KEYCODE_HOME", "", ""])
    ws.append(["EXCEL_CANONICAL", 5, "swipe", "10,20", "30,40", ""])
    wb.save(xlsx)

    files = convert_excel_to_yaml(
        xlsx,
        tmp_path / "canonical-output",
        contract_mode="canonical",
        metadata_by_tc={"EXCEL_CANONICAL": CANONICAL_METADATA},
    )

    assert len(files) == 1
    tc = yaml.safe_load(files[0].read_text(encoding="utf-8"))
    assert tc["tc_name"] == "EXCEL_CANONICAL"
    assert "name" not in tc
    assert tc["metadata"] == CANONICAL_METADATA
    assert tc["steps"] == [
        {"action": "tap_text", "target": "Settings"},
        {"action": "tap_id", "target": "com.app:id/ok"},
        {"action": "wait", "duration": 2000},
        {"action": "key", "key": "KEYCODE_HOME"},
        {"action": "swipe", "x": 10, "y": 20, "x2": 30, "y2": 40},
    ]
    schema = json.loads(
        (Path(__file__).parent.parent / "tc_step_schema.json").read_text(
            encoding="utf-8"
        )
    )
    assert validate_canonical_tc(tc, schema) == []


def test_canonical_excel_preserves_fractional_wait_seconds(tmp_path):
    xlsx = tmp_path / "fractional-wait.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_FRACTIONAL_WAIT", 1, "wait", 0.25, "", ""])
    wb.save(xlsx)

    files = convert_excel_to_yaml(
        xlsx,
        tmp_path / "fractional-wait-output",
        contract_mode="canonical",
        metadata_by_tc={"EXCEL_FRACTIONAL_WAIT": CANONICAL_METADATA},
    )

    tc = yaml.safe_load(files[0].read_text(encoding="utf-8"))
    assert tc["steps"] == [{"action": "wait", "duration": 250}]


def test_canonical_excel_rejects_boolean_wait(tmp_path):
    xlsx = tmp_path / "boolean-wait.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_BOOLEAN_WAIT", 1, "wait", True, "", ""])
    wb.save(xlsx)
    output_dir = tmp_path / "boolean-wait-output"

    with pytest.raises(ValueError, match="INVALID_UNIT") as error:
        convert_excel_to_yaml(
            xlsx,
            output_dir,
            contract_mode="canonical",
            metadata_by_tc={"EXCEL_BOOLEAN_WAIT": CANONICAL_METADATA},
        )

    assert [finding.code for finding in error.value.findings] == ["INVALID_UNIT"]
    assert not list(output_dir.glob("*.yaml"))


def test_canonical_excel_requires_explicit_metadata(tmp_path):
    xlsx = tmp_path / "missing-metadata.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_METADATA_REQUIRED", 1, "wait", "1", "", ""])
    wb.save(xlsx)
    output_dir = tmp_path / "missing-metadata-output"

    with pytest.raises(ValueError, match="METADATA_REQUIRED") as error:
        convert_excel_to_yaml(xlsx, output_dir, contract_mode="canonical")

    assert [finding.code for finding in error.value.findings] == [
        "METADATA_REQUIRED"
    ]
    assert not list(output_dir.glob("*.yaml"))


def test_canonical_excel_swipe_requires_two_coordinate_pairs(tmp_path):
    xlsx = tmp_path / "missing-endpoint.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_SWIPE_REQUIRED", 1, "swipe", 10, 20, "30,40"])
    wb.save(xlsx)
    output_dir = tmp_path / "missing-endpoint-output"

    with pytest.raises(ValueError, match="SWIPE_ENDPOINT_MISSING") as error:
        convert_excel_to_yaml(
            xlsx,
            output_dir,
            contract_mode="canonical",
            metadata_by_tc={"EXCEL_SWIPE_REQUIRED": CANONICAL_METADATA},
        )

    assert [finding.code for finding in error.value.findings] == [
        "SWIPE_ENDPOINT_MISSING"
    ]
    assert not list(output_dir.glob("*.yaml"))


def test_legacy_excel_output_is_unchanged(tmp_path):
    xlsx = tmp_path / "legacy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["TC Name", "Step", "Action", "Parameter1", "Parameter2", "Expected"])
    ws.append(["EXCEL_LEGACY", 1, "wait", "2", "", ""])
    ws.append(["EXCEL_LEGACY", 2, "key", "KEYCODE_HOME", "", ""])
    ws.append(["EXCEL_LEGACY", 3, "swipe", 10, 20, "30,40"])
    wb.save(xlsx)

    default_files = convert_excel_to_yaml(xlsx, tmp_path / "default")
    explicit_files = convert_excel_to_yaml(
        xlsx,
        tmp_path / "explicit",
        contract_mode="legacy",
    )

    assert [path.name for path in explicit_files] == [
        path.name for path in default_files
    ]
    assert [path.read_bytes() for path in explicit_files] == [
        path.read_bytes() for path in default_files
    ]
    tc = yaml.safe_load(default_files[0].read_text(encoding="utf-8"))
    assert tc == {
        "name": "EXCEL_LEGACY",
        "description": "",
        "steps": [
            {"action": "wait", "seconds": 2},
            {"action": "key", "keycode": "KEYCODE_HOME"},
            {"action": "swipe", "x1": 10, "y1": 20},
        ],
    }
