"""row_loader 단위 테스트.

실제 엑셀 파일에 의존하지 않고, openpyxl Workbook을 임시로 만들어 테스트한다.
"""
import pytest
from pathlib import Path
from openpyxl import Workbook

from src.mmi_converter.row_loader import load_mmi_rows, _find_header_row, _build_column_map


@pytest.fixture
def tmp_xlsx(tmp_path) -> Path:
    """정상 MMI 시트 구조의 임시 엑셀 파일."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ODIN 기본기능 TC(MMI 내용추가)(4번)"

    # 헤더
    ws.append([
        "  ", "`", "Functionality",
        "Pre-condition (사전 조건)", "Test procedure (재현 절차)",
        "Expected result (기대 결과)", "Priority",
    ])
    # 정상 데이터 행
    ws.append([
        7, "SOS button", "안심 기능",
        "ZEM 앱 설치", "설정 > 아이케어기능 > SOS",
        "SOS 메시지 전송", "S",
    ])
    ws.append([
        8, "Auto call receive", "안심 기능",
        "", "설정 > 아이케어기능 > 연속 전화 자동수신 > On",
        "연속 전화 자동수신 On", "S",
    ])
    # 빈 procedure + 빈 expected → skip 대상
    ws.append([
        9, "Empty row", "안심 기능",
        "", "", "", "S",
    ])
    # procedure만 있는 행 → 유효
    ws.append([
        10, "Proc only", "화면 잠금",
        "", "홈 > 설정 > 화면 잠금",
        "", "A",
    ])

    path = tmp_path / "test_mmi.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def normalized_header_xlsx(tmp_path) -> Path:
    """헤더에 공백/줄바꿈이 포함된 엑셀 파일."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ODIN 기본기능 TC(MMI 내용추가)(4번)"

    ws.append([
        "  ", "`", " Functionality ",
        " Pre-condition\n(사전 조건) ", " Test procedure\n(재현 절차) ",
        " Expected result\n(기대 결과) ", " Priority ",
    ])
    ws.append([
        1, "Wi-Fi", "네트워크",
        "", "설정 > 네트워크 > Wi-Fi",
        "Wi-Fi 메뉴 진입된다", "S",
    ])

    path = tmp_path / "test_normalized.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def carry_forward_xlsx(tmp_path) -> Path:
    """feature_name이 비어있어 carry-forward가 필요한 엑셀."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ODIN 기본기능 TC(MMI 내용추가)(4번)"

    ws.append([
        "  ", "`", "Functionality",
        "Pre-condition (사전 조건)", "Test procedure (재현 절차)",
        "Expected result (기대 결과)", "Priority",
    ])
    # 첫 행: feature_name 있음
    ws.append([
        1, "Block numbers", "안심 기능",
        "", "설정 > 안심 기능 > 수신 차단 > On",
        "수신 차단 On", "S",
    ])
    # 두 번째 행: feature_name 비어있음 → carry-forward
    ws.append([
        2, None, "안심 기능",
        "", "설정 > 안심 기능 > 수신 차단 > Off",
        "수신 차단 Off", "S",
    ])
    # 세 번째 행: feature_name 비어있음 + functionality도 비어있음
    ws.append([
        3, "", "",
        "", "설정 > 안심 기능 > 수신 차단 번호 추가",
        "번호 추가됨", "S",
    ])

    path = tmp_path / "test_carry.xlsx"
    wb.save(path)
    wb.close()
    return path


class TestLoadMMIRows:
    def test_loads_valid_rows(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        assert len(rows) == 3  # 빈 행 1건 제외

    def test_row_fields(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        first = rows[0]
        assert first.no == "7"
        assert first.feature_name == "SOS button"
        assert first.functionality == "안심 기능"
        assert "설정" in first.procedure
        assert first.priority == "S"

    def test_skips_empty_procedure_and_expected(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        nos = [r.no for r in rows]
        assert "9" not in nos  # 빈 행 skip

    def test_keeps_procedure_only_row(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        proc_only = [r for r in rows if r.no == "10"]
        assert len(proc_only) == 1
        assert proc_only[0].expected_result == ""

    def test_sheet_name_stored(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        assert all(r.sheet_name == "ODIN 기본기능 TC(MMI 내용추가)(4번)" for r in rows)

    def test_tc_name_generation(self, tmp_xlsx):
        rows = load_mmi_rows(tmp_xlsx)
        assert rows[0].tc_name == "7_SOS_button"


class TestHeaderNormalize:
    def test_normalized_headers(self, normalized_header_xlsx):
        rows = load_mmi_rows(normalized_header_xlsx)
        assert len(rows) == 1
        assert rows[0].functionality == "네트워크"
        assert "Wi-Fi" in rows[0].procedure


class TestCarryForward:
    def test_feature_name_carry_forward(self, carry_forward_xlsx):
        rows = load_mmi_rows(carry_forward_xlsx)
        assert len(rows) == 3
        assert rows[0].feature_name == "Block numbers"
        assert rows[1].feature_name == "Block numbers"  # carry-forward
        assert rows[2].feature_name == "Block numbers"  # carry-forward

    def test_functionality_carry_forward(self, carry_forward_xlsx):
        rows = load_mmi_rows(carry_forward_xlsx)
        assert rows[2].functionality == "안심 기능"  # carry-forward


class TestInvalidSheet:
    def test_missing_sheet_raises(self, tmp_xlsx):
        with pytest.raises(ValueError, match="시트"):
            load_mmi_rows(tmp_xlsx, sheet_name="존재하지 않는 시트")
