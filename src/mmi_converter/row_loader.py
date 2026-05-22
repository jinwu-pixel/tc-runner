from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import MMIRow


# 헤더 매칭용 정규식 (공백/줄바꿈 무시, 핵심 키워드 기반)
_HEADER_PATTERNS = {
    "functionality": re.compile(r"functionality|시험\s*목적|목적", re.IGNORECASE),
    "feature_name": re.compile(r"검증\s*항목", re.IGNORECASE),
    "precondition": re.compile(r"pre.?condition|사전\s*조건", re.IGNORECASE),
    "procedure": re.compile(r"test\s*procedure|재현\s*절차|시험\s*절차", re.IGNORECASE),
    "expected": re.compile(r"expected\s*result|기대\s*결과|예상\s*로그|판정\s*기준", re.IGNORECASE),
    "priority": re.compile(r"priority|우선순위", re.IGNORECASE),
    "no": re.compile(r"tc\s*id|번호", re.IGNORECASE),
}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _find_header_row(rows: list[tuple]) -> int:
    """핵심 헤더 키워드가 포함된 행 인덱스를 반환한다."""
    # procedure 또는 functionality 컬럼을 찾는다
    anchor_keys = ["procedure", "functionality"]
    for i, row in enumerate(rows[:10]):
        for cell in row:
            normalized = _normalize_header(cell)
            for key in anchor_keys:
                if _HEADER_PATTERNS[key].search(normalized):
                    return i
    raise ValueError("헤더 행을 찾을 수 없습니다 (시험 절차/Functionality 컬럼 누락)")


def _build_column_map(header_row: tuple) -> dict[str, int]:
    """헤더 행에서 키워드 기반으로 컬럼 인덱스를 매핑한다."""
    col_map: dict[str, int] = {}

    for i, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        for key, pattern in _HEADER_PATTERNS.items():
            if pattern.search(normalized) and key not in col_map:
                col_map[key] = i

    # functionality 컬럼 기준으로 앞 두 컬럼을 no, feature_name으로 추정
    # 단, 패턴으로 이미 매핑된 경우 positional 추정을 건너뛴다
    func_idx = col_map.get("functionality")
    if func_idx is not None and func_idx >= 2:
        if "no" not in col_map:
            col_map["no"] = func_idx - 2
        if "feature_name" not in col_map:
            col_map["feature_name"] = func_idx - 1

    return col_map


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_mmi_rows(
    xlsx_path: Path | str,
    sheet_name: str = "ODIN 기본기능 TC(MMI 내용추가)(4번)",
) -> list[MMIRow]:
    """MMI 시트를 읽어 MMIRow 리스트로 변환한다."""
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"시트 '{sheet_name}'이(가) 없습니다. 사용 가능: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header_idx = _find_header_row(rows)
    col_map = _build_column_map(rows[header_idx])

    required = ["functionality", "procedure", "expected"]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {missing}")

    result: list[MMIRow] = []
    prev_feature_name = ""
    prev_functionality = ""

    for row_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        # 충분한 컬럼이 없으면 skip
        max_col = max(col_map.values())
        if len(row) <= max_col:
            continue

        no = _safe_str(row[col_map["no"]]) if "no" in col_map else ""
        feature_name = _safe_str(row[col_map["feature_name"]]) if "feature_name" in col_map else ""
        functionality = _safe_str(row[col_map["functionality"]])
        precondition = _safe_str(row[col_map["precondition"]]) if "precondition" in col_map else ""
        procedure = _safe_str(row[col_map["procedure"]])
        expected_result = _safe_str(row[col_map["expected"]])
        priority = _safe_str(row[col_map["priority"]]) if "priority" in col_map else ""

        # carry-forward
        if feature_name:
            prev_feature_name = feature_name
        else:
            feature_name = prev_feature_name

        if functionality:
            prev_functionality = functionality
        else:
            functionality = prev_functionality

        # procedure 또는 expected_result 중 하나라도 있어야 유효 행
        if not procedure and not expected_result:
            continue

        result.append(
            MMIRow(
                row_index=row_idx,
                no=no,
                feature_name=feature_name,
                functionality=functionality,
                precondition=precondition,
                procedure=procedure,
                expected_result=expected_result,
                priority=priority,
                sheet_name=sheet_name,
            )
        )

    return result
