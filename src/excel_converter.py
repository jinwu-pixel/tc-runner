import copy
import json
from pathlib import Path
from collections import defaultdict
from typing import Any, Literal, Mapping

import yaml
from openpyxl import load_workbook

from src.execution_contract import (
    ContractFinding,
    normalize_step,
    validate_canonical_tc,
)


# action별 Parameter1/Parameter2 매핑 규칙
PARAM_MAP = {
    "tap_text":     {"param1": "text"},
    "tap_id":       {"param1": "id"},
    "tap_xy":       {"param1": "x", "param2": "y"},
    "swipe":        {"param1": "x1", "param2": "y1"},  # x2,y2는 추가 컬럼 필요 — 간소화
    "key":          {"param1": "keycode"},
    "shell":        {"param1": "command"},
    "wait":         {"param1": "seconds", "param1_type": int},
    "screenshot":   {"param1": "name"},
    "verify_text":  {"param1": "text"},
    "verify_shell": {"param1": "command", "param2": "expected"},
    "input_text":   {"param1": "text"},
}

_REQUIRED_METADATA = {
    "runnable",
    "tc_class",
    "execution_type",
    "manual_detail",
}


class ExcelConversionError(ValueError):
    def __init__(self, findings: tuple[ContractFinding, ...]):
        self.findings = findings
        detail = "; ".join(
            f"{finding.code} {finding.path}: {finding.detail}"
            for finding in findings
        )
        super().__init__(detail)


def convert_excel_to_yaml(
    xlsx_path: Path,
    output_dir: Path,
    *,
    contract_mode: Literal["legacy", "canonical"] = "legacy",
    metadata_by_tc: Mapping[str, Mapping[str, Any]] | None = None,
) -> list[Path]:
    """엑셀 T/C 파일을 YAML 파일들로 변환한다."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    sheet_title = ws.title

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 헤더 스킵
    wb.close()

    # TC Name별로 step 그룹핑
    tc_groups = defaultdict(list)
    for row_number, row in enumerate(rows, start=2):
        tc_name, step_num, action, param1, param2, expected = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )
        if not tc_name or not action:
            continue
        tc_groups[str(tc_name)].append({
            "action": str(action),
            "param1": param1,
            "param2": param2,
            "expected": expected,
            "row_number": row_number,
        })

    if contract_mode == "canonical":
        return _convert_canonical_groups(
            tc_groups,
            xlsx_path=xlsx_path,
            sheet_title=sheet_title,
            output_dir=output_dir,
            metadata_by_tc=metadata_by_tc,
        )
    if contract_mode != "legacy":
        raise ValueError(f"unsupported contract_mode: {contract_mode!r}")

    output_dir.mkdir(parents=True, exist_ok=True)
    created_files = []

    for tc_name, raw_steps in tc_groups.items():
        steps = []
        for raw in raw_steps:
            step = _build_step(raw)
            if step:
                steps.append(step)

        tc_data = {"name": tc_name, "description": "", "steps": steps}

        out_file = output_dir / f"{tc_name}.yaml"
        with open(out_file, "w", encoding="utf-8") as f:
            yaml.dump(tc_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
        created_files.append(out_file)

    return created_files


def _convert_canonical_groups(
    tc_groups: Mapping[str, list[dict]],
    *,
    xlsx_path: Path,
    sheet_title: str,
    output_dir: Path,
    metadata_by_tc: Mapping[str, Mapping[str, Any]] | None,
) -> list[Path]:
    findings: list[ContractFinding] = []
    documents: list[tuple[str, dict]] = []
    schema_path = Path(__file__).parent.parent / "tc_step_schema.json"
    schema = json.loads(schema_path.read_text(encoding="utf-8"))

    for tc_name, raw_steps in tc_groups.items():
        metadata = metadata_by_tc.get(tc_name) if metadata_by_tc else None
        missing = (
            sorted(_REQUIRED_METADATA - set(metadata))
            if isinstance(metadata, Mapping)
            else sorted(_REQUIRED_METADATA)
        )
        if missing:
            findings.append(
                _excel_finding(
                    "METADATA_REQUIRED",
                    f"{xlsx_path}:{sheet_title}:{tc_name}.metadata",
                    "metadata",
                    None,
                    f"explicit metadata fields required: {', '.join(missing)}",
                )
            )
            continue

        tc_findings_start = len(findings)
        steps: list[dict] = []
        for raw in raw_steps:
            source = f"{xlsx_path}:{sheet_title}:row {raw['row_number']}"
            if raw["action"] == "swipe":
                try:
                    x, y = _parse_coordinate_pair(raw.get("param1"))
                    x2, y2 = _parse_coordinate_pair(raw.get("param2"))
                except (TypeError, ValueError):
                    findings.append(
                        _excel_finding(
                            "SWIPE_ENDPOINT_MISSING",
                            source,
                            "x2/y2",
                            "Parameter1/Parameter2",
                            "canonical swipe requires two integer coordinate pairs",
                        )
                    )
                    continue
                steps.append(
                    {
                        "action": "swipe",
                        "x": x,
                        "y": y,
                        "x2": x2,
                        "y2": y2,
                    }
                )
                continue

            if raw["action"] == "wait":
                legacy_step = {"action": "wait"}
                if raw.get("param1") is not None:
                    # Canonical normalization owns the exact seconds -> ms
                    # conversion.  The legacy builder intentionally coerces
                    # through int(), which would truncate fractional cells and
                    # turn booleans into numeric durations.
                    legacy_step["seconds"] = raw["param1"]
            else:
                legacy_step = _build_step(raw)
            if legacy_step is None:
                continue
            normalized = normalize_step(legacy_step, path=source)
            findings.extend(
                finding
                for finding in normalized.findings
                if finding.severity == "ERROR"
            )
            steps.append(normalized.value)

        if len(findings) != tc_findings_start:
            continue

        tc_data = {
            "tc_name": tc_name,
            "description": "",
            "metadata": copy.deepcopy(dict(metadata)),
            "steps": steps,
        }
        validation_errors = validate_canonical_tc(tc_data, schema)
        if validation_errors:
            findings.append(
                _excel_finding(
                    "CANONICAL_VALIDATION_FAILED",
                    f"{xlsx_path}:{sheet_title}:{tc_name}",
                    None,
                    None,
                    "; ".join(validation_errors),
                )
            )
        documents.append((tc_name, tc_data))

    if findings:
        raise ExcelConversionError(tuple(findings))

    output_dir.mkdir(parents=True, exist_ok=True)
    created_files: list[Path] = []
    for tc_name, tc_data in documents:
        out_file = output_dir / f"{tc_name}.yaml"
        with open(out_file, "w", encoding="utf-8") as f:
            yaml.dump(
                tc_data,
                f,
                allow_unicode=True,
                default_flow_style=False,
                sort_keys=False,
            )
        created_files.append(out_file)
    return created_files


def _parse_coordinate_pair(value: Any) -> tuple[int, int]:
    if not isinstance(value, str):
        raise TypeError("coordinate pair must be text")
    parts = [part.strip() for part in value.split(",")]
    if len(parts) != 2 or not all(parts):
        raise ValueError("coordinate pair must have two values")
    return int(parts[0]), int(parts[1])


def _excel_finding(
    code: str,
    path: str,
    canonical_field: str | None,
    observed_field: str | None,
    detail: str,
) -> ContractFinding:
    return ContractFinding(
        code=code,
        path=path,
        severity="ERROR",
        canonical_field=canonical_field,
        observed_field=observed_field,
        detail=detail,
    )


def _build_step(raw: dict) -> dict | None:
    """엑셀 행 데이터를 YAML step dict로 변환한다."""
    action = raw["action"]
    mapping = PARAM_MAP.get(action)
    if not mapping:
        return {"action": action}

    step = {"action": action}

    if "param1" in mapping and raw["param1"] is not None:
        key = mapping["param1"]
        value = raw["param1"]
        if mapping.get("param1_type") == int:
            try:
                value = int(value)
            except (ValueError, TypeError):
                pass
        step[key] = value

    if "param2" in mapping and raw.get("param2") is not None and raw["param2"] != "":
        step[mapping["param2"]] = raw["param2"]

    # verify_shell의 expected는 Expected 컬럼에서 가져옴
    if action == "verify_shell" and raw.get("expected"):
        step["expected"] = str(raw["expected"])

    return step
