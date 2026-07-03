# -*- coding: utf-8 -*-
"""ALT Basic tc_id 충돌검사 · 유일성 감사 도구 (P-1 — synth prep 선행 게이트).

승격 배경 (scratch → scripts, 2026-07-03):
  tc_id 스킴 `ALTBASIC_<PREFIX>_<excel_row3>`는 비단사라 cross-batch 충돌이 가능하다.
  batch11 실충돌 4건의 실제 원인은 워크플로 phantom side-effect(batch10 dir 오기록)였고,
  Excel sheet 내 TC ID 중복(83건)은 **잠재(latent) 구조 위험**(실발현 0)이다
  (FAILURE_TAXONOMY C7 FM1). 두 위험을 합성 prep의 **선행 게이트**로 상시 검사한다.

순수 함수(단위 테스트 대상, IO 없음):
  - find_collisions(assigned, existing)          — cross-batch + internal 중복 탐지
  - resolve_collisions_with_suffix(assigned, existing) — 결정적 suffix 충돌 해소
  - audit_sheet_tcid_dups(records)               — (sheet, TC ID) 유일성 감사

IO wrapper(openpyxl/glob, lazy import):
  - existing_stems(base, exclude_subdir)
  - excel_tcid_records(xlsx)
  - main()  — 게이트: 유일성 감사 + cross-batch 회귀, 충돌 시 exit 1
"""
from __future__ import annotations

import glob
import os
import sys
from collections import Counter

BASE = os.path.join(os.path.dirname(__file__), "..", "THOR2 - ALT Basic TC Audit")
XLSX = os.path.join(os.path.dirname(__file__), "..", "doc", "[THOR 2] ALT Basic Test Case_FULL.xlsx")


# ─────────────────────────── 순수 함수 ───────────────────────────

def find_collisions(assigned, existing) -> dict:
    """assigned tc_id 리스트를 기존 집합·자기 자신과 대조.

    반환: {"cross_batch": [기존과 충돌한 tc_id], "internal_dup": [리스트 내 중복 tc_id]}
    (각 리스트는 최초 발견 순서, 중복 제거).
    """
    existing = set(existing)
    seen: set = set()
    cross: list = []
    internal: list = []
    for t in assigned:
        if t in existing and t not in cross:
            cross.append(t)
        if t in seen and t not in internal:
            internal.append(t)
        seen.add(t)
    return {"cross_batch": cross, "internal_dup": internal}


def resolve_collisions_with_suffix(assigned, existing) -> list:
    """충돌하는 tc_id에 결정적 suffix(`_2`, `_3`, ...)를 붙여 유일하게 만든다.

    규칙: assigned 순서대로, 이미 쓰인(existing ∪ 지금까지 해소된) 값과 충돌하면
    가장 작은 미사용 `_k`(k>=2)를 붙인다. 순서·입력이 같으면 결과도 같다(결정적).
    반환: assigned와 평행한 해소된 tc_id 리스트.
    """
    used = set(existing)
    out: list = []
    for t in assigned:
        if t not in used:
            resolved = t
        else:
            k = 2
            while f"{t}_{k}" in used:
                k += 1
            resolved = f"{t}_{k}"
        used.add(resolved)
        out.append(resolved)
    return out


def audit_sheet_tcid_dups(records) -> dict:
    """(sheet, tc_id) 레코드에서 **같은 sheet 내** 중복 TC ID를 감사.

    다른 sheet의 동일 TC ID는 중복이 아니다(per-sheet 유일성).
    반환: {sheet: {tc_id: count}}  (중복 없으면 {}).
    """
    per_sheet: dict = {}
    for sheet, tc_id in records:
        per_sheet.setdefault(sheet, []).append(tc_id)
    out: dict = {}
    for sheet, ids in per_sheet.items():
        dup = {k: c for k, c in Counter(ids).items() if c > 1}
        if dup:
            out[sheet] = dup
    return out


# ─────────────────────────── IO wrapper ───────────────────────────

def _rel(p, base=BASE):
    return os.path.relpath(p, base).replace("\\", "/")


def existing_stems(base=BASE, exclude_subdir=None) -> dict:
    """base 하위 모든 ALTBASIC_*.yaml tc_id stem → {stem: [상대경로]} (exclude_subdir 제외)."""
    stems: dict = {}
    for f in glob.glob(os.path.join(base, "**", "ALTBASIC_*.yaml"), recursive=True):
        norm = f.replace("\\", "/")
        if exclude_subdir and ("/" + exclude_subdir + "/") in norm:
            continue
        stem = os.path.basename(f).replace("_canonical.yaml", "").replace(".yaml", "")
        stems.setdefault(stem, []).append(_rel(f, base))
    return stems


def excel_tcid_records(xlsx=XLSX) -> list:
    """Excel corpus에서 (sheet, tc_id) 레코드 추출 (audit_sheet_tcid_dups 입력용)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    records: list = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        hr = None
        for i, row in enumerate(rows[:6]):
            j = " | ".join(str(c) if c is not None else "" for c in row).lower()
            if "functionality" in j and ("procedure" in j or "재현" in j) and ("expected" in j or "기대" in j):
                hr = i
                break
        if hr is None:
            continue
        for row in rows[hr + 1:]:
            if not row or row[0] is None:
                continue
            s = str(row[0]).strip()
            try:
                s = "%g" % float(s)
            except ValueError:
                pass
            if s:
                records.append((name, s))
    return records


def main() -> None:
    # 1) 첫 게이트: (sheet, TC ID) 유일성 감사
    records = excel_tcid_records()
    dups = audit_sheet_tcid_dups(records)
    total_dup = sum(len(v) for v in dups.values())
    print("=== (sheet, TC ID) 유일성 감사 (잠재 충돌 구조적 원인) ===")
    print("중복 TC ID 보유 sheet:", len(dups), "| 중복 TC ID 총:", total_dup)
    for sh, dd in sorted(dups.items()):
        print("  %-26s %s" % (sh, dict(list(dd.items())[:8])))

    # 2) cross-batch 회귀검증 (batch11 예시)
    b11 = [
        os.path.basename(f).replace("_canonical.yaml", "").replace(".yaml", "")
        for f in glob.glob(os.path.join(BASE, "stage1_s2_salvage_batch11", "*.yaml"))
    ]
    existing = set(existing_stems(exclude_subdir="stage1_s2_salvage_batch11"))
    res = find_collisions(b11, existing)
    print("\n=== batch11 cross-batch 회귀검증 ===")
    print("batch11 tc_id:", len(b11))
    print("cross_batch 충돌:", res["cross_batch"] or "none")
    print("internal 중복:", res["internal_dup"] or "none")
    if res["cross_batch"] or res["internal_dup"]:
        print("\n제안 해소(결정적 suffix):", resolve_collisions_with_suffix(b11, existing))
        sys.exit(1)


if __name__ == "__main__":
    main()
