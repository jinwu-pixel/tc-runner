#!/usr/bin/env python
# -*- coding: utf-8 -*-
# 06-17 복합 기능시나리오 워크북 SKT/LGU+ 결과·비고 기재 (실기 2026-06-17 runtime).
# 컬럼: 10=SKT결과 12=LGU+결과 13=시험자 14=비고. 행: 3=REG_A..10=LTE.
import openpyxl
from openpyxl.styles import PatternFill, Alignment

WB = r"ODIN2 - Engineer IMS/ODIN2_EngineerIMS_TC_2026-06-17.xlsx"
TESTER = "ALT_Chung"
GREEN = PatternFill("solid", fgColor="C6EFCE")   # PASS
YELLOW = PatternFill("solid", fgColor="FFF2CC")  # partial / NOTE
WRAP = Alignment(wrap_text=True, vertical="top")

# row -> (carrier_col, result_text, note)
REC = {
 4: (10, "runtime PASS", "SKT runtime 2026-06-17(USIM 교체 전). reboot 후 SIP User-Agent=ALT-UA-TEST/1.0·SUBSCRIBE Expires=3600 반영. SIP Timer T1=500 write 확정(재전송 미트리거=보조축). SubExp=3600이 SUBSCRIBE Invalid 유발=망 최소간격 NOTE. ★등록 override는 reboot-load 필요(airplane 토글로는 미반영)."),
 3: (12, "값반영 PASS / 403 NOTE", "LGU+ runtime 2026-06-17. reboot 후 REGISTER 반영: req-URI/realm=ims.mnc006.mcc450.3gppnetwork.org·Authorization username=PRID·Expires=1200. 단 403 FORBIDDEN+availableServices=[](override 도메인 비-routable). 값은 실리나 등록실패=SKT TC01의 404와 대응. reset+reboot로 복구."),
 5: (12, "부분 반영 (LGU+)", "LGU+ runtime 2026-06-17. write 전부 result=0. 호 offer SDP: AMR-WB offer됨(단 EVS 우선=LGU+ 코덱우선순위 미강제)·AMR/AMR-WB mode-set(4/8) fmtp 미반영(octet-align=1만). C1: HD=ON(마지막)이 Voice Codec=AMR-WB pref 덮어씀. fresh 재기재 후도 동일=LGU+ 거동(06-16 SKT는 반영)."),
 6: (12, "부분 반영 (LGU+)", "LGU+ runtime 2026-06-17. write 전부 result=0. 호 SDP: m=audio 포트 50004~50008(설정 50000~50010 범위)=Traffic Port 반영✓·refresher=uac✓. 단 Session-Expires=3600(설정 1810 미반영)=LGU+ 프로파일 고정. NV#73842/EFS#25038 reboot 비영속."),
 7: (12, "부분 반영 (LGU+)", "LGU+ runtime 2026-06-17. write 전부 result=0. 영상호 m=video offer: H265(119) 최우선=Video Codec=H.265 반영✓(기본 H263-first 대비 승격). 단 video Traffic Port(50020/30) 미반영=m=video 포트 동적(1254). RTP Timer=15 aliveness."),
 8: (12, "runtime PASS", "LGU+ runtime 2026-06-17(teardown). IMS Reset status=deactivateConfigs/enableAutoMode=true 'Done. IMS config restored'. reset+reboot 후 Domain/RegExp/VoiceCodec=(not configured) 환원✓·IMS 재등록 정상(VOICE/SMS/VIDEO)✓. BTS#25071 reset 거동 확인."),
 9: (10, "Way1+Way2 확인", "device NV(캐리어 무관, SKT 시점 실행). HSPA: diag_nv -w -id 3649 (03→05, 시험 후 03 복구)✓. Auto Answer: qcRilSetAutoAnwser toggle ON→OFF→복구 ON✓. 기능검증(WCDMA AS Release / 수신호 자동응답)=WCDMA·incoming-gated NOTE."),
 10: (10, "best-effort (no-op)", "device hook(캐리어 무관, SKT 시점). ROHC/CDRX write hook(qcRilSetRrcRohc/FgiDrx) result=true 발화하나 NV readback=(not set). ★BTS#18582 미구현(write no-op) 확정. RRC 반영 미검증."),
}

wb = openpyxl.load_workbook(WB)
ws = wb["복합 기능시나리오"]
for row, (col, result, note) in REC.items():
    c_res = ws.cell(row, col); c_res.value = result
    c_res.fill = GREEN if result.endswith("PASS") else YELLOW
    c_res.alignment = WRAP
    ws.cell(row, 13).value = TESTER
    cn = ws.cell(row, 14); cn.value = note; cn.alignment = WRAP
wb.save(WB)
print("SAVED:", WB)
# read-back verify
wb2 = openpyxl.load_workbook(WB)
ws2 = wb2["복합 기능시나리오"]
for row, (col, result, note) in REC.items():
    print(f"  row{row} col{col}={ws2.cell(row,col).value!r} 시험자={ws2.cell(row,13).value!r}")
