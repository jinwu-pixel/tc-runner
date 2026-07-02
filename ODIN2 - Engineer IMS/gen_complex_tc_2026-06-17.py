# -*- coding: utf-8 -*-
"""ODIN2 Engineer Mode — 복합 기능시나리오 TC 재설계 (2026-06-17).

배경: 06-16 '복합 TC (기능시나리오)' 시트는 한 케이스에 설정 6~9개로 과다.
  → 케이스당 설정 3~4개로 줄이고, IMS 뿐 아니라 General·LTE 카테고리 추가.
설계 원칙: '케이스 = 3~4개 설정 + 단일 공유 동적-로그 신호'.
  같은 로그 신호(REGISTER / 음성 SDP / 영상 SDP / LTE RRC)에 함께 실리는 설정만 한 케이스로 묶어,
  한 번의 로그 캡처로 그 케이스의 모든 설정을 동시 검증.
서술 톤: 06-16 '복합 TC'(필드격리) 시트와 동일 — 짧은 번호형 step + 짧은 번호형 expected + 1~2줄 비고.
  (검증 훅·제약은 비고/범례로 분리, step·expected는 쉽고 명확하게 유지.)
판정 = Way1(UI readback) ∧ Way2(write hook 로그 result=0) ∧ Way3(.qmdl 기능 신호 = ground truth).

출력 = 신규 ODIN2_EngineerIMS_TC_2026-06-17.xlsx (06-16 워크북·기존 스크립트 무수정).
  결과 컬럼(SKT/KT/LGU+) = 공란(계획·미실행). 현재 ODIN2 AT-M150 미연결.
결정적 생성(LLM 변동 없음). 카탈로그 값은 본 스크립트 dict 에 인코딩(단말 무접촉).
"""
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# 기본 = 06-17 워크북. argv[1] 주면 그 경로로(검증/임시용).
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent / "ODIN2_EngineerIMS_TC_2026-06-17.xlsx"
BUILD = "AT-M150Z0612U"
AUTHOR = "ALT_Chung"

COLS = ["TC ID", "ITEM", "Functionality", "Pre-condition (사전 조건)", "Test procedure (재현 절차)",
        "Expected result (기대 결과)", "Priority", "작성자", "시험 버전", "SKT 결과", "KT 결과", "LGU+ 결과",
        "시험자", "비고", "BTS 번호"]


def pre(tab, setline):
    head = {
        "IMS": "1. 단말 ODIN2 AT-M150 / 빌드 Z0612U\n2. SKT USIM, LTE 캠프 (IMS 등록 상태)\n3. Tele Engineer > [Enter Engineering Mode] > [IMS] 탭",
        "GENERAL": "1. 단말 ODIN2 AT-M150 / 빌드 Z0612U\n2. SKT USIM\n3. Tele Engineer > [Enter Engineering Mode] > [GENERAL] 탭",
        "LTE": "1. 단말 ODIN2 AT-M150 / 빌드 Z0612U\n2. SKT USIM, LTE 캠프\n3. Tele Engineer > [Enter Engineering Mode] > [LTE] 탭",
    }[tab]
    return head + "\n" + setline + "\n5. offline 로깅 + QXDM(.qmdl) ON · svc power stayon true"


# ───────────────────────── 8 케이스 ─────────────────────────
# 각 row = [TCID, ITEM(카테고리), Functionality, Pre, Procedure, Expected, Prio, BTS, 비고]
ROWS = []

# IMS-1 : REGISTER 식별자/주기 (3) → REGISTER
ROWS.append([
 "CMB_IMS_REG_A", "Engineer (IMS) 기능복합",
 "등록 식별자·주기 (Domain·PRID·Register Expires) → REGISTER 신호 반영",
 pre("IMS", "4. [설정 3] Domain · PRID · Register Expires"),
 "1. 'Domain' = ims.mnc0XX.mcc450.3gppnetwork.org (routable) > [Write]\n"
 "2. 'PRID' = <user>@<domain> > [Write]\n"
 "3. 'Register Expires' = 1200 > [Write]\n"
 "4. 단말 reboot (부팅 시 DataPopup '사용') → IMS 재등록\n"
 "5. 재등록 로그(.qmdl) 수집 — REGISTER 확인",
 "1. 3개 항목 Write status=OK / readback=입력값\n"
 "2. 로그 REGISTER에 반영: 도메인=Domain · privateURI=PRID(reboot 후) · Expires=1200\n"
 "3. (참고) 짧은 realm 'sktelecom2'는 404 — 값은 실리나 등록 실패. 200 OK는 routable 도메인 필요",
 "A", "25059 / 25066 / 25035",
 "Way2 EFS [INI_WRITE] result=0 / Way3 .qmdl REGISTER. 입력 전 focus→clear(EditText auto-clear 안 됨). 캡처는 200 OK/최종실패까지 유지."])

# IMS-2 : UA·구독·SIP 타이머 (3) → SIP 헤더/SUBSCRIBE
ROWS.append([
 "CMB_IMS_REG_B", "Engineer (IMS) 기능복합",
 "SIP 헤더·구독·타이머 (User Agent·Subscribe Expires·SIP Timer) → SIP 메시지 반영",
 pre("IMS", "4. [설정 3] User Agent · Subscribe Expires · SIP Timer"),
 "1. 'User Agent' = ALT-UA-TEST/1.0 > [Write]\n"
 "2. 'Subscribe Expires' = 3600 > [Write]\n"
 "3. 'SIP Timer' T1 = 500 > [Write]\n"
 "4. 단말 reboot → 재등록·재구독\n"
 "5. 로그(.qmdl) 수집 — REGISTER/SUBSCRIBE 확인",
 "1. 3개 항목 Write status=OK / readback=입력값\n"
 "2. 모든 SIP 메시지 User-Agent 헤더 = ALT-UA-TEST/1.0\n"
 "3. SUBSCRIBE Expires = 3600 · SIP Timer T1 = 재전송 간격에 반영",
 "A", "19420 / 25036 / 19429",
 "Way2 UA=NV [QCRIL_JAVA] / SubExp=EFS [INI_WRITE] result=0. SIP Timer는 보조축(write 로그가 권위), 메뉴 존재 재확인 선행."])

# IMS-3 : 음성 코덱 (4) → 음성호 SDP m=audio
ROWS.append([
 "CMB_IMS_VOICE", "Engineer (IMS) 기능복합",
 "음성 코덱 (Voice Codec·AMR·AMR-WB·HD Voice) → 음성호 SDP 반영",
 pre("IMS", "4. [설정 4] Voice Codec · AMR ModeSet · AMR-WB ModeSet · HD Voice"),
 "1. 'Voice Codec Priority' = AMR-WB preferred > [Write]\n"
 "2. 'AMR Codec ModeSet' = 4 > [Write]\n"
 "3. 'AMR-WB Codec ModeSet' = 8 > [Write]\n"
 "4. 'HD Voice Setting' = ON > [Write]\n"
 "5. 일반(음성)호 발신 → 수초 통화 → 종료\n"
 "6. 통화 로그(.qmdl) 수집 — INVITE/200 SDP 확인",
 "1. 4개 항목 Write status=OK / readback=입력값\n"
 "2. 음성호 SDP m=audio: 코덱 순서=설정 · AMR mode-set=0x04 · AMR-WB=0x08 · HD ON→AMR-WB 포함",
 "A", "25043 / 19443 / 19445 / 19410",
 "Voice Codec·HD Voice = 동일 AudioProfile1(C1 커플링). KT는 EVS 거부=NOTE. Way2 NV/EFS result=0."])

# IMS-4 : 세션/RTP (4) → 음성호 SIP/SDP
ROWS.append([
 "CMB_IMS_SESSION", "Engineer (IMS) 기능복합",
 "세션·RTP 전송 (Session Expires·Refresher·RTP Timer·Traffic Port) → 음성호 SIP/SDP 반영",
 pre("IMS", "4. [설정 4] Session Expires · Session Refresher · RTP Timer · Traffic Port"),
 "1. 'Session Expires' = 1810 > [Write]\n"
 "2. 'Session Refresher' = UAC(0) > [Write]\n"
 "3. 'RTP Timer' = 15 > [Write]\n"
 "4. 'Traffic Port' speech = 50000 / 50010 > [Write]\n"
 "5. 일반(음성)호 발신 → 통화 → 종료\n"
 "6. 통화 로그(.qmdl) 수집",
 "1. 4개 항목 Write status=OK / readback=입력값\n"
 "2. SIP Session-Expires = 1810 · refresher = uac\n"
 "3. SDP RTP 포트 = 50000~50010 · 통화 중 RTP alive 주기 = 15초",
 "A", "19425 / 25038 / 19581 / 19593",
 "Session Expires·RTP Timer = 동일 NV#73842 별도 필드. Traffic Port는 영상 케이스와 의도적 중복(음성 컨텍스트). Way2 result=0."])

# IMS-5 : 영상 코덱 (3) → 영상호 SDP m=video
ROWS.append([
 "CMB_IMS_VIDEO", "Engineer (IMS) 기능복합",
 "영상 코덱·전송 (Video Codec·Traffic Port·RTP Timer) → 영상호 SDP 반영",
 pre("IMS", "4. [설정 3] Video Codec · Traffic Port(video) · RTP Timer"),
 "1. 'Video Codec Priority' = H.265 > [Write]\n"
 "2. 'Traffic Port' video = 50020 / 50030 > [Write]\n"
 "3. 'RTP Timer' = 15 > [Write]\n"
 "4. 영상호 발신 → 통화 → 종료\n"
 "5. (선택) Video Codec = H.264 변경 후 영상호 재발신\n"
 "6. 통화 로그(.qmdl) 수집 — INVITE SDP 확인",
 "1. 3개 항목 Write status=OK / readback=입력값\n"
 "2. 영상호 SDP m=video: payload=H265 (H.264 변경 시 H264)\n"
 "3. video RTP 포트 = 50020~50030 · RTP alive = 15초",
 "A", "25049 / 19593 / 19581",
 "Video Codec = VideoProfile1(음성과 별도 키). Traffic·RTP는 음성 케이스와 의도적 중복(영상 컨텍스트). Way2 result=0."])

# IMS-6 : IMS Reset (action) → 환원/복원
ROWS.append([
 "CMB_IMS_RESET", "Engineer (IMS) 기능복합",
 "IMS Reset to Default → 전 항목 MBN default 환원 (reset 후 설정 비영속) · 복원 겸용",
 pre("IMS", "4. [선행] REG_A/VOICE/SESSION 설정값 적용 상태(또는 임의 값 다수 설정)"),
 "1. 'IMS Reset to Default' > [Reset] (다이얼로그 확인)\n"
 "2. (reboot 전) 주요 항목 [Read] — RTP / RegExp / Domain\n"
 "3. 단말 reboot (부팅 시 DataPopup '사용')\n"
 "4. 주요 항목 [Read]\n"
 "5. (대조) reset 없이 RTP=15·RegExp=36000 > [Write] → reboot → [Read]",
 "1. Reset status: deactivateConfigs=true / enableAutoMode=true\n"
 "2. (reboot 전) write는 정상 반영, readback 일치 — 즉시 무반영 아님\n"
 "3. reboot 후 전 항목 기본값 환원: RTP=10 · RegExp=(not configured) · Session=1800 · Domain=(not configured)\n"
 "4. (대조) reset 없는 write는 reboot 후 유지 → 환원은 reset 때문",
 "A", "25071",
 "BTS#25071 RESET-NONPERSIST (OBSERVED). 'reset 후 설정이 1회 reboot로 날아감' = 기대 갭 flag."])

# GENERAL-1 : HSPA·Auto Answer (2, best-effort)
ROWS.append([
 "CMB_GEN_01", "Engineer (GENERAL) 기능복합",
 "General 항목 (HSPA Setting·Auto Answer) → 동작/로그 반영 (best-effort)",
 pre("GENERAL", "4. [설정 2] HSPA Setting · Auto Answer"),
 "1. 'HSPA Setting' = Rel-9 > [Write]\n"
 "2. 'Auto Answer' = 활성 > [설정]\n"
 "3. (HSPA) WCDMA 캠프 전환 후 통화 시도\n"
 "4. (Auto Answer) 타 단말로 수신호 → 자동 응답 관찰\n"
 "5. 로그 수집 (logcat / WCDMA 시 .qmdl)",
 "1. HSPA Write status=OK / readback=Rel-9 (LTE에서도 write 확인 가능)\n"
 "2. (HSPA 기능) WCDMA RRC AS Release Indicator = 설정값 — WCDMA 캠프 필요\n"
 "3. (Auto Answer) 수신호 자동 응답 동작",
 "B", "16232 / —",
 "⚠ HSPA 기능검증=WCDMA 캠프 전제(LTE는 write까지) · 메뉴 'default 3' vs 기대 5 불일치 flag · Auto Answer=BTS 없음·inconclusive."])

# LTE-1 : ROHC·CDRX FGI (2, 18582 Active)
ROWS.append([
 "CMB_LTE_01", "Engineer (LTE) 기능복합",
 "LTE Protocol Feature (ROHC·CDRX FGI) → LTE RRC 반영 (best-effort)",
 pre("LTE", "4. [설정 2] ROHC · CDRX FGI (둘 다 BTS#18582)"),
 "1. 'ROHC' = Enable(1) > [Write]\n"
 "2. 'CDRX FGI' = FGI#4=0 · FGI#5=0 > [Write]\n"
 "3. 단말 reboot 또는 재캠프 (RRC 재설정 트리거)\n"
 "4. 로그(.qmdl) 수집 — LTE RRC 확인",
 "1. ROHC·CDRX Write status=OK / readback=설정값\n"
 "2. RRC Reconfiguration에 ROHC profile 활성\n"
 "3. UE Capability에 FGI#4 / #5 = 설정값",
 "B", "18582",
 "⚠ BTS#18582 = Active(메뉴 구현중) → write no-op 가능, 18582 구현 확인 선행. 현재 Read=(not set)."])

# ───────────────────────── 입력 가이드 (IMS + General + LTE) ─────────────────────────
GUIDE_HDR = ["항목 (BTS)", "인터페이스 / 필드", "입력 방식", "입력 범위 / 허용값", "예시 (SKT)", "기대 효과 (로그 신호)", "검증상태 / 주의"]
GUIDE = [
 ("User Agent (19420)", "NV#69689 / ims_user_agent (STRING)", "텍스트", "임의 문자열(SIP UA 헤더)", "ALT-UA-TEST/1.0", "전 SIP 메시지 User-Agent 헤더", "Way2 [QCRIL_JAVA]. BTS 19420 본문 미추출"),
 ("Subscribe Expires (25036)", "EFS [SIPConfig:StandardTimers] SipSubscribeValue", "텍스트(초)", "양의 정수 초", "3600", "SUBSCRIBE Expires(재구독 주기)", "Way2 [INI_WRITE]. 기본 (not configured)"),
 ("Session Expires (19425)", "NV#73842 / sessionExpires (UINT16)", "텍스트(초)", "0~65535", "1810", "SIP Session-Expires(RFC4028)", "관찰 1800→1810 정상. NV#73842(RTP와 공유)"),
 ("Register Expires (25035)", "EFS [SIPConfig:StandardTimers] SipRegValue", "텍스트(초)", "양의 정수 초", "1200 / 36000", "REGISTER Expires(재등록 주기)", "Way2 [INI_WRITE]. 기본 (not configured)"),
 ("Domain (25059)", "EFS [ParamConfig:RegistrationDefaultParams] domainName", "텍스트", "routable 홈도메인", "ims.mnc0XX.mcc450.3gppnetwork.org (carrier 확인)", "REGISTER realm/req-URI", "⚠ bare realm(sktelecom2)=404(TC1). 200 OK엔 routable 값"),
 ("PRID (25066)", "EFS [ParamConfig:RegistrationDefaultParams] privateURI", "텍스트", "<user>@<domain>", "4500612345678@ims.mnc006.mcc450.3gppnetwork.org", "Authorization/P-Preferred-Identity(등록 ID)", "reboot 후 반영. 오타 금지(TC1 PRID 오타→404 일조)"),
 ("SIP Timer (19429)", "NV / ims_sip_config (T1,T2,T4,TA~TK)", "텍스트(멀티)", "타이머별 ms", "T1=500", "SIP 재전송 트랜잭션 타이밍", "멀티필드 보조축. 메뉴 존재 inconclusive(재확인 선행)"),
 ("Voice Codec Priority (25043)", "EFS [QIPCALL:ImsMediaProfileConfig] AudioProfile1", "라디오", "Default/EVS/AMR-WB/AMR-NB/AMR-WB pref", "AMR-WB preferred", "음성 SDP m=audio 코덱 우선순위", "HD Voice와 동일 키. KT EVS 거부=NOTE"),
 ("AMR Codec ModeSet (19443)", "NV#73846 / amrModeSet (UINT32)", "텍스트(비트마스크)", "0x00~0xFF (0=미전송)", "4 (0x04, M2 5.90kbps)", "m=audio AMR fmtp mode-set", "Way2 [QCRIL_JAVA]. 관찰 0→4 정상"),
 ("AMR-WB Codec ModeSet (19445)", "NV#73846 / amrWbModeSet (UINT32)", "텍스트(비트마스크)", "0x000~0x1FF (0=미전송)", "8 (0x08, M3 14.25kbps)", "m=audio AMR-WB fmtp mode-set", "Way2 [QCRIL_JAVA]. AMR와 동일 NV(별도 subfield)"),
 ("RTP Timer (19581)", "NV#73842 / rtpLinkAlivenessTimer (UINT32)", "텍스트(초)", "정수 초(0=비활성)", "15 (SKT 기대 10)", "통화 중 RTP aliveness 주기", "Way2 [QCRIL_JAVA]. fix 0→10. 관찰 10→15 정상"),
 ("Session Refresher (25038)", "EFS [QIPCALL:ImsVoiceSessionTimerConfig] sessionRefresherType", "라디오", "Default/UAC(0)/UAS(1)", "UAC(0)", "SIP Session-Expires refresher 주체", "Way2 [INI_WRITE]. UAC=단말, UAS=망"),
 ("HD Voice Setting (19410)", "EFS [QIPCALL:ImsMediaProfileConfig] AudioProfile1", "라디오", "ON(AMR-WB+AMR)/OFF(AMR-NB)", "ON", "m=audio AMR-WB payload 포함 여부", "Voice Codec과 동일 키. ★EVS-only인데 ON표기 GAP후보"),
 ("Video Codec Priority (25049)", "EFS [QIPCALL:ImsMediaProfileConfig] VideoProfile1", "라디오", "Default/H.263/H.264/H.265", "H.265", "영상 SDP m=video payload", "Way2 [INI_WRITE]. VideoProfile1(음성과 별도 키)"),
 ("Traffic Port (19593)", "NV#73845 / IMSRTPDynamicConfig speech/video Start·Stop", "멀티필드", "0~65535 (Start/Stop 쌍)", "speech 50000/50010 · video 50020/50030", "SDP RTP 포트 범위", "Way2 [QCRIL_JAVA] 멀티필드"),
 ("IMS Reset to Default (25071)", "PDC (action)", "버튼", "—", "—", "전체 IMS config MBN default 리셋", "⚠ reset 후 manual write=1 reboot로 환원(RESET-NONPERSIST)"),
 ("HSPA Setting (16232)", "NV#3649 / accessStratumReleaseIndicator (Diag#3649)", "라디오/텍스트", "Rel-99/5/6/7/8/9", "Rel-9 (=5)", "WCDMA RRC AS Release Indicator", "⚠ 기능검증=WCDMA 캠프 필요. 메뉴 'default 3' vs BTS 5 불일치"),
 ("Auto Answer (—)", "QcRilHook (BTS 없음)", "—", "—", "—", "수신호 자동 응답 동작", "⚠ 로그경로 불명확. 행위검증(inconclusive)"),
 ("ROHC (18582)", "EFS .../rrc/rohc_supported (Hook)", "토글", "0(OFF)/1(ON)", "1 (Enable)", "RRC Reconfiguration PDCP rohc profiles", "⚠ 18582 Active(메뉴 구현중). 현재 Read=(not set)"),
 ("CDRX FGI (18582)", "EFS .../rrc/cap/fgi (Hook)", "필드", "FGI#4/#5 비트", "FGI#4=0·FGI#5=0", "UE Capability Information FGI 비트", "⚠ 18582 Active. ROHC와 동일 티켓(LTE Protocol Feature)"),
]

# ───────────────────────── 범례·검증·커버리지 ─────────────────────────
LEGEND = [
 ("ODIN2 Engineer Mode — 복합 기능시나리오 TC (2026-06-17)", ""),
 ("설계 원칙", "케이스 = 3~4개 설정 + 단일 공유 로그신호. 한 번의 로그 캡처로 케이스 전 설정 동시 검증."),
 ("단말/빌드", "ODIN2 AT-M150 / AT-M150Z0612U · carrier=SKT 1차(KT/LGU+ 매트릭스 확장)"),
 ("케이스 구성", "IMS 6(REG_A·REG_B·VOICE·SESSION·VIDEO·RESET) + General 1 + LTE 1 = 8 케이스. 케이스당 설정 ≤4."),
 ("판정 (3-way)", "Way1 = 단말 화면 readback=입력값·status OK · Way2 = write 로그 result=0 · Way3 = .qmdl 기능 신호(REGISTER/SDP/RRC)에 반영. step·expected는 쉬운 표현, 로그훅 상세는 본 표 참조."),
 ("Way2 로그훅 — NV", "[QCRIL_JAVA] writeNvField result:0 + readResp OK value=X. RTP·SessExp=73842 / AMR·AMRWB=73846 / Traffic=73845 / UA=69689 / SIPT=ims_sip_config / HSPA=Diag#3649"),
 ("Way2 로그훅 — EFS", "[INI_WRITE] /efsprofiles/overideconfig result=0 + [INI_READ]. RegExp·SubExp·Domain·PRID·Voice/Video Codec·Refresher·HD · (LTE ROHC/CDRX=Hook)"),
 ("Way3 신호", "REGISTER/SUBSCRIBE=SIP 0x156E(Call-ID 매칭) · 음성=INVITE/200 SDP m=audio · 영상=m=video · LTE=RRC Reconfig/UE Capability. AP logcat callProfile=비권위."),
 ("커버리지 — IMS 16/16", "REG_A: Domain·PRID·RegExp | REG_B: UA·SubExp·SIPT | VOICE: VoiceCodec·AMR·AMRWB·HD | SESSION: SessExp·Refresher·RTP·Traffic | VIDEO: VideoCodec·Traffic·RTP | RESET: IMS Reset"),
 ("커버리지 — General/LTE", "GEN_01: HSPA·Auto Answer | LTE_01: ROHC·CDRX FGI"),
 ("TC1 교훈 (등록)", "값 반영(REGISTER에 실림) ≠ 등록 성공(200 OK). bare realm(sktelecom2)=404. routable 도메인+PRID 정확형식. 캡처 윈도우=200 OK/최종실패까지."),
 ("앱 함정", "EditText auto-clear 안 됨(focus→clear→input) · 게이트 재진입=am force-stop 후 relaunch · btn bounds dump 추출 · svc power stayon true · 부팅 DataPopup '사용'."),
 ("General/LTE 제약", "HSPA 기능검증=WCDMA 캠프 전제(LTE는 write까지) · HSPA default 3 vs 5 flag · Auto Answer=BTS 없음 inconclusive · LTE ROHC/CDRX=BTS#18582 Active(메뉴 구현중)."),
 ("결과 표기", "SKT/KT/LGU+ 결과 컬럼 = 공란(계획·미실행). 현재 ODIN2 AT-M150 미연결. planned를 implemented로 표기 금지."),
 ("연관 문서", "VERIFY_PROTOCOL.md · BTS_TICKET_SPECS.md · APP_BTS_MATCH.md · SPEC_CATALOG.md · 06-16 워크북(필드격리 복합 TC·메뉴항목별 TC)"),
]

# ───────────────────────── 워크북 빌드 ─────────────────────────
wb = openpyxl.Workbook()
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")

# --- Sheet 1: 복합 기능시나리오 ---
ws = wb.active
ws.title = "복합 기능시나리오"
ws["A1"] = "ODIN2 Engineer Mode — 복합 기능시나리오 TC (IMS·General·LTE / 케이스당 설정 3~4 / 로그검증)"
ws["A1"].font = Font(bold=True, size=13); ws.merge_cells("A1:O1")
hdr_fill = PatternFill("solid", fgColor="548235"); hdr_font = Font(bold=True, color="FFFFFF", size=10)
for c, h in enumerate(COLS, 1):
    cell = ws.cell(2, c, h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
r = 3
for tcid, item, func, pre_, proc, exp, prio, bts, note in ROWS:
    full = [tcid, item, func, pre_, proc, exp, prio, AUTHOR, BUILD, "", "", "", AUTHOR, note, bts]
    for c, v in enumerate(full, 1):
        cell = ws.cell(r, c, v); cell.border = border; cell.alignment = wrap_top; cell.font = Font(size=9)
    ws.row_dimensions[r].height = 130
    r += 1
widths = [16, 22, 30, 36, 46, 46, 6, 9, 14, 8, 7, 8, 9, 40, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A3"; ws.sheet_view.showGridLines = False

# --- Sheet 2: 입력 가이드 ---
ws2 = wb.create_sheet("입력 가이드")
ws2["A1"] = "ODIN2 Engineer Mode — 입력값 가이드 (IMS·General·LTE / AT-M150 Z0612U / SKT)"
ws2["A1"].font = Font(bold=True, size=12); ws2.merge_cells("A1:G1")
for c, h in enumerate(GUIDE_HDR, 1):
    cell = ws2.cell(2, c, h); cell.fill = PatternFill("solid", fgColor="2E75B6"); cell.font = Font(bold=True, color="FFFFFF", size=10); cell.border = border; cell.alignment = center
r = 3
for g in GUIDE:
    for c, v in enumerate(g, 1):
        cell = ws2.cell(r, c, v); cell.border = border; cell.alignment = wrap_top; cell.font = Font(size=9)
    ws2.row_dimensions[r].height = 42
    r += 1
gwidths = [26, 44, 14, 26, 40, 34, 40]
for i, w in enumerate(gwidths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A3"; ws2.sheet_view.showGridLines = False

# --- Sheet 3: 범례·검증·커버리지 ---
ws3 = wb.create_sheet("범례·검증·커버리지")
ws3.column_dimensions["A"].width = 22; ws3.column_dimensions["B"].width = 120
r = 1
for k, v in LEGEND:
    a = ws3.cell(r, 1, k); a.font = Font(bold=True, size=10, color="385723"); a.alignment = wrap_top
    b = ws3.cell(r, 2, v); b.alignment = wrap_top; b.font = Font(size=9)
    ws3.row_dimensions[r].height = 30 if len(str(v)) < 90 else 46
    r += 1
ws3.sheet_view.showGridLines = False

wb.save(OUT)
print("OK saved:", OUT.name)
print("sheets:", wb.sheetnames)
print("케이스:", len(ROWS), "| 입력가이드 행:", len(GUIDE), "| 범례 행:", len(LEGEND))
