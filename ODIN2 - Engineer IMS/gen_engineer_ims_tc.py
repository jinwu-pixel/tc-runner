# -*- coding: utf-8 -*-
"""ODIN2 Engineer Mode IMS — 메뉴 항목당 TC 5개 생성기.
형식 = ALT Basic Test Case (doc/[THOR 2] ALT Basic Test Case_FULL.xlsx) 컬럼 준수:
  TC ID | ITEM | Functionality | Pre-condition(사전 조건) | Test procedure(재현 절차) |
  Expected result(기대 결과) | Priority | 작성자 | 시험 버전 | SKT 결과 | KT 결과 | LGU+ 결과 | 시험자 | 비고 | BTS 번호
절차/기대 = "1. … 2. …" 번호형. 결정적 생성(LLM 변동 없이 형식 통일)."""
# ⚠ 1회성 스캐폴드 — 재실행 금지. 출력 xlsx는 생성 후 SKT 결과가 수동 기입됨(authoritative).
#   재실행 시 (1) 수동 결과 전체 wipe, (2) 셀 텍스트 정정 되돌림. 콘텐츠 정정은 xlsx 직접 편집으로 한다.
#   2026-06-16: RESET-NONPERSIST 정정 반영(소스 문자열 동기) — NV/EFS 세분 노트는 xlsx만 보유.
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

OUT = Path(__file__).resolve().parent / "ODIN2_EngineerIMS_TC_2026-06-16.xlsx"
BUILD = "AT-M150Z0612U"
AUTHOR = "ALT_Chung"
ITEM = "Engineer Mode (IMS)"

PRE = ("1. 단말 ODIN2 AT-M150 / 빌드 Z0612U\n"
       "2. SKT USIM, LTE 캠프 (IMS 등록 상태)\n"
       "3. 앱서랍 > Tele Engineer > [Enter Engineering Mode] > [IMS] 탭")

COLS = ["TC ID","ITEM","Functionality","Pre-condition (사전 조건)","Test procedure (재현 절차)",
        "Expected result (기대 결과)","Priority","작성자","시험 버전","SKT 결과","KT 결과","LGU+ 결과",
        "시험자","비고","BTS 번호"]

# ---------- 항목 정의 ----------
# type: N(숫자/문자 입력형 Read/Write), R(라디오 선택형), S(문자열), M(멀티필드), A(액션)
ITEMS = [
 dict(abbr="RTP", func="RTP Timer", bts="19581", type="N",
      desc="RTP Link Aliveness Timer (NV#73842 rtpLinkAlivenessTimer, 초)",
      default="10 (SKT)", observed="10", w1="15", w2="20",
      shared="Session Expires (동일 NV#73842, 별도 필드)",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=10 = SKT 기대값 일치"),
           3:("PASS","Write 15 → Read=15"),4:("PASS","RTP 변경 시 Session=1800 보존(C2 격리)"),
           5:("PASS","reboot 후 MBN default 10 복원")}),
 dict(abbr="SESEXP", func="Session Expires", bts="19425", type="N",
      desc="IMS Session Expires (NV#73842 sessionExpires, 초)",
      default="1800 (SKT, BTS 일반 default 360 — USIM/MCFG별)", observed="1800", w1="720", w2="1810",
      shared="RTP Timer (동일 NV#73842, 별도 필드)",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=1800"),
           3:("PASS","Write 1810 → Read=1810"),4:("PASS","Session 변경 시 RTP=15 보존(C2 격리)"),
           5:("PASS","reboot 후 1800 복원")}),
 dict(abbr="REGEXP", func="Register Expires", bts="25035", type="N",
      desc="SIP Register Expires (EFS [SIPConfig:StandardTimers] SipRegValue, 초)",
      default="(not configured) / BTS reset default 60000", observed="(not configured)", w1="36000", w2="35000",
      shared="Subscribe Expires (동일 SIPConfig 섹션)",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=(not configured) = override 미설정")}),
 dict(abbr="SUBEXP", func="Subscribe Expires", bts="25036", type="N",
      desc="SIP Subscribe Expires (EFS [SIPConfig:StandardTimers] SipSubscribeValue, 초)",
      default="(not configured) / BTS reset default 600000", observed="(not configured)", w1="3600", w2="3500",
      shared="Register Expires (동일 SIPConfig 섹션)",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=(not configured)")}),
 dict(abbr="REFR", func="Session Refresher", bts="25038", type="R",
      desc="VoLTE Session Refresher Type (EFS [QIPCALL:ImsVoiceSessionTimerConfig] sessionRefresherType)",
      options="Default / UAC(0) — UE / UAS(1) — Network",
      default="Default (not configured)", observed="Default (not configured)",
      opt1="UAC(0)", opt1val="UAC (UE as refresher)", opt2="UAS(1)", opt2val="UAS (Network as refresher)",
      linked="Session Expires (세션 타이머 — 별도 키, 독립)",
      ver={1:("PASS","진입·라디오 노출 OK"),2:("PASS","현재 선택=Default")}),
 dict(abbr="VCODEC", func="Voice Codec Priority", bts="25043", type="R",
      desc="VoLTE Voice Codec Priority (EFS [QIPCALL:ImsMediaProfileConfig] AudioProfile1)",
      options="Default / EVS / AMR-WB / AMR-NB / AMR-WB preferred",
      default="Default (not configured)", observed="(not configured) / Default",
      opt1="EVS", opt1val="EVS_0_126;EVS_1_127", opt2="AMR-WB", opt2val="AMR_1_104;AMR_4_103",
      linked="HD Voice Setting (동일 AudioProfile1 키)",
      ver={1:("PASS","진입·라디오 노출 OK"),2:("PASS","현재 선택=Default, Read=(not configured)"),
           3:("PASS","EVS Write → Read=EVS_0_126;EVS_1_127"),
           4:("PASS","HD Voice=OFF 시 Voice 뷰=AMR_0_102;AMR_2_101/AMR-NB (C1 커플링)"),
           5:("PASS","IMS Reset 후 Default 복원. KT는 EVS 거부=carrier NOTE")}),
 dict(abbr="VIDCODEC", func="Video Codec Priority", bts="25049", type="R",
      desc="ViLTE Video Codec Priority (EFS [QIPCALL:ImsMediaProfileConfig] VideoProfile1)",
      options="Default(H.263,H.264,H.265) / H.263 / H.264 / H.265",
      default="Default (H263_0;H264_0;H265_0)", observed="H263_0;H264_0;H265_0 / Default",
      opt1="H.264", opt1val="H264_0", opt2="H.265", opt2val="H265_0",
      linked="Voice Codec Priority (별도 키 VideoProfile1 — 독립)",
      ver={1:("PASS","진입·라디오 노출 OK"),2:("PASS","현재=Default, Read=H263_0;H264_0;H265_0")}),
 dict(abbr="HDV", func="HD Voice Setting", bts="19410", type="R",
      desc="IMS HD Voice (AMR-WB) enable/disable (EFS [QIPCALL:ImsMediaProfileConfig] AudioProfile1)",
      options="ON (AMR-WB+AMR) / OFF (AMR-NB only)",
      default="ON", observed="ON",
      opt1="OFF", opt1val="AMR_0_102;AMR_2_101", opt2="ON", opt2val="AMR-WB+AMR(default)",
      linked="Voice Codec Priority (동일 AudioProfile1 키)",
      ver={1:("PASS","진입·라디오 노출 OK"),2:("PASS","현재 선택=ON"),
           3:("PASS","OFF Write → Read=OFF"),
           4:("PASS","Voice Codec=EVS 시 HD 뷰='ON (custom: EVS…)' — 동일키 커플링/표기 GAP 후보(C1)"),
           5:("PASS","IMS Reset 후 ON 복원")}),
 dict(abbr="AMR", func="AMR Codec ModeSet", bts="19443", type="N",
      desc="AMR Codec ModeSet (NV#73846 amrModeSet, bitmask)",
      default="0 (0x00, no mode-set sent)", observed="0 (0x00)", w1="4", w2="8",
      shared="AMR-WB Codec ModeSet (동일 NV#73846, 별도 subfield)",
      note_extra="값=bitmask 0x01~0x100 (앱이 0x04→M2(5.90kbps) 식 디코딩 표시)",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=0 (0x00)"),
           3:("PASS","Write 4 → Read='4 (0x04) M2(5.90kbps)'"),
           4:("PASS","AMR 변경 시 AMR-WB=0 보존(C3 격리)"),
           5:("PASS","explicit Write 0 복원")}),
 dict(abbr="AMRWB", func="AMR-WB Codec ModeSet", bts="19445", type="N",
      desc="AMR-WB Codec ModeSet (NV#73846 amrWbModeSet, bitmask)",
      default="0 (0x00, no mode-set sent)", observed="0 (0x00)", w1="8", w2="16",
      shared="AMR Codec ModeSet (동일 NV#73846, 별도 subfield)",
      note_extra="값=bitmask 0x01~0x100",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=0 (0x00)"),
           3:("PASS","Write 8 → Read='8 (0x08) M3(14.25kbps)'"),
           4:("PASS","AMR-WB 변경 시 AMR=4 보존(C3 격리)"),
           5:("PASS","explicit Write 0 복원")}),
 dict(abbr="DOMAIN", func="Domain", bts="25059", type="S",
      desc="IMS Domain (EFS [ParamConfig:RegistrationDefaultParams] domainName)",
      default="(not configured)", observed="(not configured)",
      w1="sktelecom2", w2="(carrier별: SKT=sktelecom2 / LGU=LGT2)", reboot=False,
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=(not configured), SKT 기대 sktelecom2")}),
 dict(abbr="PRID", func="PRID", bts="25066", type="S",
      desc="IMS Private User Identity (EFS [ParamConfig:RegistrationDefaultParams] privateURI)",
      default="(not configured) / 예 4500612345678@ims.mnc006.mcc450.3gppnetwork.org",
      observed="(not configured)", w1="test", w2="(형식 <user>@<domain>)", reboot=True,
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=(not configured)")}),
 dict(abbr="UA", func="User Agent", bts="19420", type="S",
      desc="IMS User Agent (NV#69689 ims_user_agent, STRING)",
      default="TTA-VoLTE/3.0 AT-M150S/Z0612U Device_Type/Android_Phone SKT",
      observed="TTA-VoLTE/3.0 AT-M150S/Z0612U Device_Type/Android_Phone SKT", w1="(테스트 문자열)", w2="—",
      reboot=False, note_extra="BTS 19420 본문 미추출 — 티켓 확인 필요",
      ver={1:("PASS","진입·Read OK"),2:("PASS","Read=빌드(Z0612U)·carrier(SKT) 반영된 UA 문자열")}),
 dict(abbr="TPORT", func="Traffic Port", bts="19593", type="M",
      desc="Voice/Video RTP Port (NV#73845 IMSRTPDynamicConfig: speech/videoStart·StopPort)",
      default="Ver 0 / speech 7010·7012 / video 7020·7022", observed="동일",
      ver={1:("PASS","진입·[Read All Fields] OK"),2:("PASS","Ver0/speech 7010·7012/video 7020·7022")}),
 dict(abbr="RESET", func="IMS Reset to Default", bts="25071", type="A",
      desc="IMS config를 MBN default로 전체 리셋 (PDC, [Reset IMS to Default] 버튼 + 확인 다이얼로그)",
      ver={1:("PASS","진입·버튼 노출"),2:("PASS","Reset 실행 → deactivateConfigs/enableAutoMode"),
           3:("PASS","Reset 후 EFS override Read=(not configured) / NV는 실제값 유지"),
           4:("PASS","★BUG-GAP(RESET-NONPERSIST): Reset 후 manual write는 런타임 커밋되나 1회 reboot로 환원(비영속, reset 특이)"),
           5:("PASS","Reset→reboot→MBN default 복원·Write 정상화")}),
 dict(abbr="SIPT", func="SIP Timer", bts="19429", type="N",
      desc="SIP Timer values T1/T2/T4/TA~TK (NV ims_sip_config)",
      default="—", observed="메뉴 존재 inconclusive(스캔 불안정)", w1="(타이머값)", w2="—",
      shared="Session Expires", note_extra="메뉴 항목 존재 수동 재확인 선행 필요",
      ver={}),
]

# ---------- Way2 (모뎀 RESP 로그) 판정 시그니처 ----------
# store: NV = QcRilHook readResp 값이 로그에 디코드되어 남음 / EFS = hook 트랜잭션은 남으나 필드 시그니처 실측 확정 필요
WAY2 = {
 "RTP":("NV","rtpLinkAlivenessTimer"), "SESEXP":("NV","sessionExpires"),
 "AMR":("NV","amrModeSet"), "AMRWB":("NV","amrWbModeSet"),
 "UA":("NV","ims_user_agent"), "TPORT":("NV","IMSRTPDynamicConfig"), "SIPT":("NV","ims_sip_config"),
 "REGEXP":("EFS","SipRegValue"), "SUBEXP":("EFS","SipSubscribeValue"), "REFR":("EFS","sessionRefresherType"),
 "VCODEC":("EFS","AudioProfile1"), "VIDCODEC":("EFS","VideoProfile1"), "HDV":("EFS","AudioProfile1"),
 "DOMAIN":("EFS","domainName"), "PRID":("EFS","privateURI"), "RESET":("ACTION","-"),
}
VERIFIED_LOG = {"RTP","SESEXP"}  # 2026-06-16 실측 로그 캡처 보유

def way2_read(it):
    store, field = WAY2.get(it["abbr"], ("EFS","-"))
    if store == "NV":
        return f"\n[판정 Way2] logcat QC_RIL_OEM_HOOK: readResp OK, field={field}, value=<표시값> (모뎀 반환 == UI 표시 일치)"
    if store == "EFS":
        return f"\n[판정 Way2] logcat QcRilHook 트랜잭션 확인 (EFS {field} — 필드 로그 시그니처 실측 확정)"
    return ""

def way2_write(it):
    store, field = WAY2.get(it["abbr"], ("EFS","-"))
    if store == "NV":
        return "\n[판정 Way2] logcat: writeNvField result=0 → 앱 자동 read-back readResp value=<입력값> (커밋 확인)"
    if store == "EFS":
        return "\n[판정 Way2] logcat: write 트랜잭션 OK → 자동 read-back 값 일치 (실측 확정)"
    return ""

# ---------- 타입별 5-TC 빌더 ----------
def tcs_for(it):
    t = it["type"]; f = it["func"]; out = []
    if t == "N":
        out = [
         ("진입 및 현재값 표시",
          f"1. [IMS] 탭에서 '{f}' 항목 탭하여 상세 진입\n2. [Read] 버튼 탭",
          f"1. '{f}' 상세 화면 진입 (타이틀에 ✅ + [{it['bts']}] 표기)\n2. Current Value에 현재값 표시, status=OK"),
         ("기본값/현재값 확인",
          f"1. '{f}' 진입 > [Read]\n2. 표시된 현재값과 기본값 비교",
          f"1. 현재값 = {it.get('default','—')}\n   (override 미설정 시 '(not configured)' = 기본 사용)"),
         ("값 변경(Write) 적용",
          f"1. '{f}' 진입\n2. New Value에 {it.get('w1','값')} 입력 > [Write]\n3. [Read]로 재조회",
          f"1. Write status = OK\n2. Read 값 = {it.get('w1','입력값')} (정상 반영)"),
         ("재변경 / 공유 NV 격리" if it.get("shared") else "재변경/경계값",
          (f"1. '{f}' = {it.get('w2','다른값')} 입력 > [Write]\n2. '{it['shared']}' 항목 진입 > [Read]"
           if it.get("shared") else f"1. '{f}' = {it.get('w2','다른값')} 입력 > [Write]\n2. [Read]"),
          (f"1. '{f}' = {it.get('w2','값')} 갱신\n2. '{it['shared'].split(' ')[0]}' 값 보존 (같은 NV 내 필드 격리)"
           if it.get("shared") else f"1. Read 값 = {it.get('w2','입력값')}로 갱신")),
         ("복원 / 복구 거동",
          f"1. New Value에 기본값 입력 > [Write] (또는 'IMS Reset to Default' 실행)\n2. [Read]\n3. (글로벌 Reset 사용 시) 단말 reboot 후 [Read]",
          f"1. 기본값 복원\n2. 주의: 글로벌 IMS Reset 후 manual write는 런타임 반영되나 1회 reboot로 환원(비영속, reset 특이). 영구 적용은 reset 없이 write. (NV는 post-reset Read=실제값, EFS override는 Read=(not configured))"),
        ]
    elif t == "R":
        out = [
         ("진입 및 옵션 표시",
          f"1. '{f}' 항목 탭하여 상세 진입\n2. [Read] 버튼 탭",
          f"1. '{f}' 상세 진입, 라디오 옵션 노출: {it['options']}\n2. Current Value + 현재 선택 라디오 표시, status=OK"),
         ("기본 선택 확인",
          f"1. '{f}' 진입 > [Read]\n2. 현재 선택 라디오 확인",
          f"1. 현재 선택 = {it.get('default','Default')} (Read={it.get('observed','—')})"),
         ("옵션 선택(Write) 적용",
          f"1. '{it['opt1']}' 라디오 선택\n2. [Write] > [Read]",
          f"1. 선택 = {it['opt1']}\n2. Current Value = {it['opt1val']}"),
         ("연동 항목 영향 확인" if "동일" in it.get("linked","") else "다른 옵션 적용",
          (f"1. '{f}' = {it['opt1']} Write 상태에서\n2. 연동 항목 '{it['linked'].split(' (')[0]}' 진입 > [Read]"
           if "동일" in it.get("linked","") else f"1. '{it['opt2']}' 라디오 선택 > [Write] > [Read]"),
          (f"1. 동일 키 공유로 연동 항목 뷰에 변경 반영 (커플링 일관성)\n   ※ HD Voice가 비-AMR-WB 프로파일에도 'ON(custom)' 표기 = GAP 후보"
           if "동일" in it.get("linked","") else f"1. 선택={it['opt2']}, Current Value={it['opt2val']}")),
         ("복원",
          f"1. 'Default' 라디오 선택 > [Write] (또는 [Reset] / 'IMS Reset to Default')\n2. [Read]",
          f"1. 기본 선택(Default) 복원"),
        ]
    elif t == "S":
        rb = "\n3. 단말 reboot (PRID 등 reboot 반영 항목)" if it.get("reboot") else ""
        rbe = "\n3. reboot 후 반영 (reboot 전 미반영은 참고)" if it.get("reboot") else ""
        out = [
         ("진입 및 현재값 표시",
          f"1. '{f}' 항목 탭하여 상세 진입\n2. [Read] 버튼 탭",
          f"1. '{f}' 상세 진입 (타이틀 [{it['bts']}])\n2. Current Value 표시, status=OK"),
         ("기본값/형식 확인",
          f"1. '{f}' 진입 > [Read]\n2. 값/형식 확인",
          f"1. 현재값 = {it.get('default','—')}"),
         ("값 변경(Write) 적용",
          f"1. '{f}' 진입\n2. New Value에 {it.get('w1','값')} 입력 > [Write]{rb}\n4. [Read]",
          f"1. Write status=OK{rbe}\n2. Read 값 = 입력값 반영"),
         ("형식/carrier 값 확인",
          f"1. carrier/형식별 값 입력 Write: {it.get('w2','—')}\n2. [Read]",
          f"1. 형식에 맞는 값 정상 반영"),
         ("복원",
          f"1. 기본값 입력 Write (또는 'IMS Reset to Default')\n2. [Read]",
          f"1. 기본값 복원"),
        ]
    elif t == "M":
        out = [
         ("진입 및 전 필드 조회",
          f"1. '{f}' 항목 탭하여 전용 화면 진입\n2. [Read All Fields] 버튼 탭",
          f"1. Traffic Port 화면 진입 (Version/speech/video 필드)\n2. 전 필드 Current Value 표시"),
         ("기본 포트값 확인",
          f"1. [Read All Fields]\n2. 각 필드값 확인",
          f"1. Ver=0, speechStart=7010, speechStop=7012, videoStart=7020, videoStop=7022"),
         ("speech 포트 변경",
          f"1. Speech Start/Stop 필드 New Value 입력 > 각 [Write]\n2. [Read]",
          f"1. speech 포트 변경값 반영"),
         ("video 포트 변경",
          f"1. Video Start/Stop 필드 New Value 입력 > 각 [Write]\n2. [Read]",
          f"1. video 포트 변경값 반영, speech 포트 보존"),
         ("복원",
          f"1. 각 필드 [Reset] 또는 기본값 Write\n2. [Read All Fields]",
          f"1. 기본 포트값(7010/7012/7020/7022) 복원"),
        ]
    elif t == "A":
        out = [
         ("진입 및 버튼 확인",
          f"1. '{f}' 항목 탭하여 진입",
          f"1. 'Reset IMS to Default' 버튼 + status('Click button to start') 노출"),
         ("Reset 실행",
          f"1. [Reset IMS to Default] 탭\n2. 확인 다이얼로그에서 [Reset] 탭",
          f"1. 확인 다이얼로그('Reset IMS config to MBN default?') 표시\n2. 실행 status: deactivateConfigs/enableAutoMode = true"),
         ("Reset 후 항목값 확인",
          f"1. Reset 실행 후 IMS 항목(예 RTP Timer) 진입 > [Read]",
          f"1. EFS override 항목 Read='(not configured)'(해제) / NV 항목(RTP 등)은 실제값 유지"),
         ("★Reset 후 Write 런타임반영·reboot 비영속(BUG-GAP)",
          f"1. Reset 직후(reboot 안함) RTP Timer 등 진입\n2. New Value 입력 > [Write] > [Read]\n3. 단말 reboot 후 [Read]",
          f"1. Write status='OK' = 런타임 반영됨(result=0 + 자동 readback 일치) — no-op 아님\n2. 단 reset 후 1회 reboot로 환원(비영속, reset 특이): NV=MBN default·EFS override=(not configured)\n   ※ 개발 보고 후보(기대 갭: reset 후 세팅이 1회 reboot로 소실)"),
         ("Reset → reboot → 복원",
          f"1. Reset 실행 후 단말 reboot\n2. (부팅 시 DataPopup '사용'으로 닫기)\n3. IMS 항목 [Read] / Write 재시도",
          f"1. RTP=10·Session=1800·AMR=0 등 MBN default 복원\n2. Write 정상 적용 재개"),
        ]
    # Way2 로그 판정 근거 부착 (Read=_1, Write=_3, BUG-GAP=A_4)
    def _augment(idx, extra):
        if extra and 0 <= idx < len(out):
            sub, proc, exp = out[idx]; out[idx] = (sub, proc, exp + extra)
    if t in ("N","R","S","M"):
        _augment(0, way2_read(it)); _augment(2, way2_write(it))
    elif t == "A":
        _augment(3, "\n[판정 Way2] logcat: writeNvField result=0 / [INI_WRITE] result=0 + 자동 readback 일치 = 런타임 커밋. post-reboot readback 환원으로 비영속 확정")

    # 결과 행 구성
    rows = []
    for i, (sub, proc, exp) in enumerate(out, start=1):
        v = it.get("ver", {}).get(i)
        skt = v[0] if v else ""
        bigo = (v[1] if v else "")
        tester = AUTHOR if v else ""
        if it.get("note_extra") and i == 1:
            bigo = (bigo + " / " if bigo else "") + it["note_extra"]
        # 판정 근거 비고 (Way2 부착 행)
        if (t in ("N","R","S","M") and i in (1, 3)) or (t == "A" and i == 4):
            tag = "판정=3-way(VERIFY_PROTOCOL.md)"
            if it["abbr"] in VERIFIED_LOG:
                tag += "; Way2 로그 실측" + ("=evidence/device/rtp_timer_*.log" if it["abbr"] == "RTP" else "(2026-06-16)")
            bigo = (bigo + " · " if bigo else "") + tag
        rows.append([
            f"IMS_{it['abbr']}_{i:02d}", ITEM, f, PRE, proc, exp,
            ("S" if i in (1,2,3) else "A"),  # Priority: 진입/기본/변경=S(필수), 그외=A
            AUTHOR, BUILD, skt, "", "", tester, bigo, it["bts"],
        ])
    return rows

# ---------- 빌드 ----------
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Engineer IMS TC"
ws.append(COLS)

allrows = []
for it in ITEMS:
    allrows.extend(tcs_for(it))
for r in allrows:
    ws.append(r)

# ---------- 스타일 ----------
hf = Font(bold=True, color='FFFFFF', size=10); hfill = PatternFill('solid', fgColor='2F5496')
pass_fill = PatternFill('solid', fgColor='C6EFCE'); gap_fill = PatternFill('solid', fgColor='FFF2CC')
item_fill = PatternFill('solid', fgColor='D9E2F3')
wrap = Alignment(wrap_text=True, vertical='top'); ctr = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin = Side(style='thin', color='BFBFBF'); border = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in range(1, len(COLS)+1):
    cell = ws.cell(1, c); cell.font = hf; cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border
ws.freeze_panes = 'A2'
widths = [13,15,18,30,40,40,7,9,14,9,9,9,9,30,8]
for w, col in zip(widths, ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']):
    ws.column_dimensions[col].width = w
prev_func = None; band = False
for ridx, r in enumerate(allrows, start=2):
    for c in range(1, len(COLS)+1):
        cell = ws.cell(ridx, c); cell.alignment = wrap; cell.border = border
    ws.cell(ridx,1).alignment = ctr; ws.cell(ridx,7).alignment = ctr
    for c in (10,11,12): ws.cell(ridx,c).alignment = ctr
    ws.cell(ridx,3).fill = item_fill
    skt = ws.cell(ridx,10).value
    if skt == "PASS": ws.cell(ridx,10).fill = pass_fill
    if "BUG-GAP" in (ws.cell(ridx,14).value or ""):
        for c in range(1,len(COLS)+1): ws.cell(ridx,c).fill = gap_fill

# 요약 시트
ws2 = wb.create_sheet("범례·요약")
summary = [
 ["ODIN2 Engineer Mode IMS — 메뉴 항목별 TC",""],
 ["형식","ALT Basic Test Case 컬럼 준수 (Pre-condition/Test procedure/Expected result/결과/BTS)"],
 ["단말/빌드", f"ODIN2 AT-M150 / {BUILD}"],
 ["구성","IMS 메뉴 16항목 × 5 TC = 80건 (항목당: 진입/기본값/변경적용/연동·경계/복원)"],
 ["검증 표기","SKT 결과 PASS(녹색)=2026-06-16 실기 manual evidence. 공란=계획(미실행). KT/LGU+=미실행"],
 ["판정 프로토콜","3-way: Way1 UI(tv_detail_value/tv_detail_status) == Way2 모뎀로그(QC_RIL_OEM_HOOK readResp value) == 의도값, ∧ Way3(reboot 영속·인접필드 격리). 상세=VERIFY_PROTOCOL.md"],
 ["Way2 로그 캡처","logcat -c 후 동작 → logcat -d | Select-String -CaseSensitive 'QC_RIL_OEM_HOOK:|TeleEngineer:'. readResp OK…value=N = 모뎀 반환값. NV항목 실측(RTP/Session=evidence/device/), EFS항목 시그니처 실측 확정"],
 ["BUG-GAP(노랑)","IMS Reset 후 manual write가 런타임 커밋되나 1회 reboot로 환원(비영속, reset 특이) — 개발 보고 후보(기대 갭). 판별=Way2 write result=0+readback 일치(커밋) ↔ post-reboot readback 환원"],
 ["carrier","SKT 1차. KT/LGU+ 컬럼 = 매트릭스 확장용. (BTS: KT는 EVS 거부=NOTE, Domain SKT=sktelecom2/LGU=LGT2)"],
 ["운영 주의","글로벌 IMS Reset 후 reboot 시 manual write가 MBN default로 환원(reset 특이·비영속). 정상 상태(reset 없음) write는 reboot 영속(Phase C/D control). silent no-op은 불성립"],
 ["복합 TC 시트","4그룹 7조합 × 3 TC = 21건. 동일NV격리(73842/73846)·EFS키커플링(AudioProfile1)·Reset영속·EFS섹션그룹. 켠 메뉴=Pre-condition 명시. PASS=오프라인로그/이전세션 검증, 공란=계획(미실행)"],
]
for row in summary: ws2.append(row)
ws2.column_dimensions['A'].width = 18; ws2.column_dimensions['B'].width = 90
ws2.cell(1,1).font = Font(bold=True, size=13)
for r in range(2, len(summary)+1):
    ws2.cell(r,1).font = Font(bold=True); ws2.cell(r,1).alignment = wrap; ws2.cell(r,2).alignment = wrap

# ---------- 복합(Combined) TC 시트 ----------
CPRE = ("1. 단말 ODIN2 AT-M150 / 빌드 Z0612U\n"
        "2. SKT USIM, LTE 캠프 (IMS 등록 상태)\n"
        "3. 앱서랍 > Tele Engineer > [Enter Engineering Mode] > [IMS] 탭")
COMBOS = [
 dict(cid="NV73842", func="RTP↔Session 필드격리 (NV#73842)", bts="19581 / 19425",
   setup="4. [복합 설정] 동일 NV#73842 두 필드 — RTP Timer(rtpLinkAlivenessTimer)·Session Expires(sessionExpires). 시작값 RTP=10·Session=1800",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'RTP Timer' 진입 > [Read]\n2. 'Session Expires' 진입 > [Read]",
      "1. RTP Read=10 / status=OK\n2. Session Read=1800 / status=OK (동일 NV#73842, 별도 필드)",
      "PASS","Way2 readResp 일치. offline_session_ledger Phase A/C"),
     ("상호작용·격리 판정",
      "1. 'Session Expires'=1810 > [Write]\n2. 'RTP Timer' [Read]\n3. (역) 'RTP'=15 > [Write] 후 'Session' [Read]",
      "1. Session readback=1810\n2. RTP=10 보존 (Session 변경이 RTP 미교란)\n3. RTP=15 시 Session=1810 보존 → 동일 NV 필드격리 OK",
      "PASS","C2 격리: Session=1810 중 RTP=10(0x0A) 보존. ledger 14:01:48"),
     ("reboot 영속·복원",
      "1. RTP=15·Session=1810 상태 reboot (부팅 시 DataPopup '사용')\n2. 두 항목 [Read]\n3. 기본값(RTP=10·Session=1800) [Write] 복원",
      "1. reboot 후 RTP=15·Session=1810 유지 (정상 상태 write 영속)\n2. 복원 확인",
      "PASS","Phase C/D 정상 write reboot survive. ledger Phase C"),
   ]),
 dict(cid="NV73846", func="AMR↔AMR-WB 필드격리 (NV#73846)", bts="19443 / 19445",
   setup="4. [복합 설정] 동일 NV#73846 두 subfield — AMR Codec ModeSet(amrModeSet)·AMR-WB Codec ModeSet(amrWbModeSet). 시작값 둘 다 0(0x00)",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'AMR Codec ModeSet' 진입 > [Read]\n2. 'AMR-WB Codec ModeSet' 진입 > [Read]",
      "1. AMR Read=0 (0x00)\n2. AMR-WB Read=0 (0x00) (동일 NV#73846, 별도 subfield)",
      "PASS","이전 세션 C3 baseline (오프라인로그 재검증 미실시)"),
     ("상호작용·격리 판정",
      "1. 'AMR Codec ModeSet'=4 > [Write]\n2. 'AMR-WB Codec ModeSet' [Read]\n3. (역) AMR-WB=8 > [Write] 후 AMR [Read]",
      "1. AMR=4 반영 ('4 (0x04) M2')\n2. AMR-WB=0 보존\n3. AMR-WB=8 시 AMR=4 보존 → subfield 격리 OK",
      "PASS","이전 세션 C3 격리 확인. NV#73846 오프라인로그 재검증은 후속"),
     ("reboot 영속·복원",
      "1. AMR=4·AMR-WB=8 상태 reboot\n2. [Read]\n3. AMR=0·AMR-WB=0 explicit [Write] 복원",
      "1. reboot 후 유지 예상 (정상 write 영속, Phase C/D NV control 근거)\n2. 복원 확인",
      "","영속은 Phase C/D NV control로 일반화 추정, 본 항목 미재현"),
   ]),
 dict(cid="AUDIOPROFILE1", func="Voice Codec↔HD Voice 커플링 (AudioProfile1)", bts="25043 / 19410",
   setup="4. [복합 설정] 동일 EFS [QIPCALL:ImsMediaProfileConfig] AudioProfile1 키 — Voice Codec Priority·HD Voice Setting. 시작 HD=ON",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'HD Voice Setting' 진입 > [Read] (라디오 현재 선택)\n2. 'Voice Codec Priority' 진입 > [Read]",
      "1. HD Voice=ON\n2. Voice Codec=Default/(not configured) (동일 AudioProfile1 키 공유)",
      "PASS","이전 세션 C1 baseline"),
     ("커플링 판정",
      "1. 'HD Voice Setting'=OFF 라디오 > [Write]\n2. 'Voice Codec Priority' [Read]\n3. (역) 'Voice Codec'=EVS > [Write] 후 'HD Voice' [Read]",
      "1. HD OFF 반영 (AudioProfile1=AMR_0_102;AMR_2_101 AMR-NB)\n2. Voice 뷰에 AMR-NB 반영 (동일 키 커플링 일관)\n3. ★GAP후보: Voice=EVS(AMR-WB 없음)인데 HD Voice 뷰='ON' 표기",
      "PASS","C1 커플링 OK. GAP후보=EVS-only에 HD ON 표기(개발 확인)"),
     ("복원",
      "1. HD='ON'·Voice='Default' 복원 (또는 'IMS Reset to Default'+reboot)\n2. [Read]",
      "1. HD=ON·Voice=Default 복원\n2. 주의: 글로벌 Reset 시 reboot 전 (not configured)",
      "","Reset 영속 거동은 Combo CMB_RESET 참조. 본 항목 reboot 미재현"),
   ]),
 dict(cid="RESET", func="IMS Reset 영속 상호작용 (Reset+NV/EFS)", bts="25071 / 19581 / 25035",
   setup="4. [복합 설정] IMS Reset to Default(PDC, 글로벌) + 대상 항목 RTP Timer(NV)·Register Expires(EFS)",
   tcs=[
     ("진입·세팅(pre-reset 커밋)",
      "1. 'RTP Timer'=15 > [Write] (NV)\n2. 'Register Expires'=36000 > [Write] (EFS)\n3. 각 [Read]",
      "1. RTP readback=15 (writeNvField result=0)\n2. RegExp readback=36000 ([INI_WRITE] result=0)\n   → 런타임 커밋 (no-op 아님)",
      "PASS","offline 로그 Phase B/D. Way2 + 단말 main 로그 일치"),
     ("Reset 후 거동 판정",
      "1. 'IMS Reset to Default' 실행 (다이얼로그 [Reset] 확인)\n2. Reset 직후(reboot 전) 'RTP'·'RegExp' [Read]\n3. Reset 직후 'RTP'=15 재 [Write] > [Read]",
      "1. status: deactivateConfigs/enableAutoMode=true\n2. RTP=10(NV 유지)·RegExp=(not configured)(EFS override clear)\n3. Reset 후 Write도 result=0+readback 일치 = 런타임 커밋(silent no-op 아님)",
      "PASS","Phase B: 이전 'no-op' 가설 불성립 확정"),
     ("reboot 영속=RESET 특이",
      "1. (Reset 후 write 상태) 단말 reboot\n2. 'RTP'·'RegExp' [Read]\n3. [대조] reset 없이 write→reboot 결과와 비교",
      "1. Reset+reboot 후 RTP=10·RegExp=(not configured) = MBN default 환원\n2. 대조(reset 없음): RTP=15·RegExp=36000 영속\n   → 비영속은 RESET 특이(auto-mode MBN reload). 버그 vs 의도=개발 판단",
      "PASS","Phase B/C/D control + .qmdl 교차검증"),
   ]),
 dict(cid="SIPTIMERS", func="Register↔Subscribe Expires 격리 (SIPConfig)", bts="25035 / 25036",
   setup="4. [복합 설정] 동일 EFS [SIPConfig:StandardTimers] 섹션 — Register Expires(SipRegValue)·Subscribe Expires(SipSubscribeValue). 시작 둘 다 (not configured)",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'Register Expires' [Read]\n2. 'Subscribe Expires' [Read]",
      "1. RegExp=(not configured) ([INI_READ] null)\n2. SubExp=(not configured) (동일 SIPConfig 섹션, 별도 key)",
      "","RegExp baseline은 Phase B 확인. SubExp 미Read(계획)"),
     ("격리 판정",
      "1. 'Register Expires'=36000 > [Write]\n2. 'Subscribe Expires' [Read]\n3. (역) SubExp=3600 > [Write] 후 RegExp [Read]",
      "1. RegExp=36000 반영\n2. SubExp=(not configured) 보존 (섹션 내 key 격리)\n3. SubExp=3600 시 RegExp=36000 보존",
      "","RegExp write/persist는 Phase D 확인. RegExp↔SubExp 격리는 미검증"),
     ("reboot 영속·복원",
      "1. RegExp=36000·SubExp=3600 상태 reboot\n2. [Read]\n3. 복원(IMS Reset+reboot)",
      "1. reset 없이 설정 시 reboot 유지 예상 (Phase D RegExp 영속 근거)\n2. 복원: (not configured)",
      "","EFS 정상 write 영속은 Phase D(RegExp) 확인. SubExp 미검증"),
   ]),
 dict(cid="REGPARAMS", func="Domain+PRID 등록 ID 조합 (RegistrationParams)", bts="25059 / 25066",
   setup="4. [복합 설정] 동일 EFS [ParamConfig:RegistrationDefaultParams] — Domain(domainName)·PRID(privateURI). SKT 기대 domain=sktelecom2. 시작 둘 다 (not configured)",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'Domain' [Read]\n2. 'PRID' [Read]",
      "1. Domain=(not configured) (SKT 기대 sktelecom2)\n2. PRID=(not configured)",
      "","이전 세션 둘 다 (not configured) 관찰. write 미실시"),
     ("등록 ID 조합 판정",
      "1. 'Domain'=sktelecom2 > [Write]\n2. 'PRID'=<user>@<domain> > [Write]\n3. 각 [Read]",
      "1. Domain=sktelecom2 반영 (REGISTER realm)\n2. PRID 반영 (주의: BTS#25066 — reboot 후 반영)",
      "","미검증. PRID reboot 반영 특성(BTS#25066)"),
     ("reboot 영속·복원",
      "1. Domain·PRID 설정 후 reboot\n2. [Read]\n3. 복원",
      "1. reboot 후 유지 (PRID는 reboot로 반영)\n2. 복원: IMS Reset+reboot",
      "","미검증"),
   ]),
 dict(cid="MEDIAPROFILE", func="Voice↔Video Codec 프로파일 격리 (ImsMediaProfileConfig)", bts="25043 / 25049",
   setup="4. [복합 설정] 동일 EFS [QIPCALL:ImsMediaProfileConfig] INI 파일, 별도 프로파일 — Voice Codec(AudioProfile1)·Video Codec(VideoProfile1). 시작 Video=Default(H263;H264;H265)",
   tcs=[
     ("진입·세팅(현재값)",
      "1. 'Voice Codec Priority' [Read]\n2. 'Video Codec Priority' [Read]",
      "1. Voice=Default/(not configured)\n2. Video=H263_0;H264_0;H265_0 (별도 VideoProfile1)",
      "","Video baseline 이전 세션 관찰. 격리 미검증"),
     ("프로파일 격리 판정",
      "1. 'Voice Codec Priority'=EVS > [Write]\n2. 'Video Codec Priority' [Read]\n3. (역) 'Video'=H.264 > [Write] 후 'Voice' [Read]",
      "1. Voice=EVS 반영 (AudioProfile1)\n2. Video=Default 보존 (별도 프로파일 키 격리)\n3. Video=H.264 시 Voice=EVS 보존",
      "","같은 INI 파일 내 프로파일 격리 — 미검증"),
     ("reboot 영속·복원",
      "1. Voice=EVS·Video=H.264 상태 reboot\n2. [Read]\n3. 복원(IMS Reset+reboot)",
      "1. reboot 후 유지 (정상 상태)\n2. 복원: Default/(not configured)",
      "","미검증"),
   ]),
]

ws3 = wb.create_sheet("복합 TC", 1)
ws3.append(COLS)
crows = []
for cb in COMBOS:
    pre = CPRE + "\n" + cb["setup"]
    for i, (sub, proc, exp, skt, bigo) in enumerate(cb["tcs"], start=1):
        tester = AUTHOR if skt == "PASS" else ""
        crows.append([f"CMB_{cb['cid']}_{i:02d}", "Engineer Mode (IMS) 복합", cb["func"], pre, proc, exp,
                      ("S" if i in (1, 2) else "A"), AUTHOR, BUILD, skt, "", "", tester, bigo, cb["bts"]])
for r in crows: ws3.append(r)
for c in range(1, len(COLS)+1):
    cell = ws3.cell(1, c); cell.font = hf; cell.fill = hfill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border
ws3.freeze_panes = 'A2'
for w, col in zip(widths, ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']):
    ws3.column_dimensions[col].width = w
ws3.column_dimensions['C'].width = 26; ws3.column_dimensions['D'].width = 42
for ridx, r in enumerate(crows, start=2):
    for c in range(1, len(COLS)+1):
        cell = ws3.cell(ridx, c); cell.alignment = wrap; cell.border = border
    ws3.cell(ridx,1).alignment = ctr; ws3.cell(ridx,7).alignment = ctr
    for c in (10,11,12): ws3.cell(ridx,c).alignment = ctr
    ws3.cell(ridx,3).fill = item_fill
    if ws3.cell(ridx,10).value == "PASS": ws3.cell(ridx,10).fill = pass_fill

wb.save(OUT)
print("SAVED:", OUT)
print("per-menu rows:", len(allrows), "| items:", len(ITEMS), "| combined rows:", len(crows), "| combos:", len(COMBOS))
