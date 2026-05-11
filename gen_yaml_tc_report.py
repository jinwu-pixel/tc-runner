"""범용 yaml TC → Excel 리포트 변환기.

ODIN2 - Music / ODIN2 - Settings 등 임의 트랙의 TC yaml 묶음을
사용자용 .xlsx 로 변환한다. minifile `MNF_TC_Report.xlsx` 의 시트 구조
(TC 요약 / 스텝 상세 / 수행 방식 분류) 를 답습한다.

사용:
    venv/Scripts/python.exe gen_yaml_tc_report.py --dir "ODIN2 - Music" \\
        --out "ODIN2 - Music/report/MUSIC_TC_Report.xlsx" \\
        --title "ODIN2 - Music TC Report"

검증결과 컬럼은 yaml metadata 에 존재하지 않으므로 기본 빈 칸이다.
--result-manifest <yaml> 옵션으로 {tc_name: 결과문자열} 매핑을 주입할 수 있다.
"""
import argparse
import io
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
SETUP_FILL = PatternFill("solid", fgColor="D9E2F3")
ACTION_FILL = PatternFill("solid", fgColor="FFF2CC")
ASSERT_FILL = PatternFill("solid", fgColor="E2EFDA")
TEARDOWN_FILL = PatternFill("solid", fgColor="EDEDED")
AUTO_FILL = PatternFill("solid", fgColor="B4C6E7")
MANUAL_LOCAL_FILL = PatternFill("solid", fgColor="FCE4D6")
EXTERNAL_FILL = PatternFill("solid", fgColor="F8CBAD")

STEP_ROLE_FILL = {
    "SETUP": SETUP_FILL,
    "ACTION": ACTION_FILL,
    "ASSERT": ASSERT_FILL,
    "TEARDOWN": TEARDOWN_FILL,
}
EXEC_TYPE_FILL = {
    "AUTO": AUTO_FILL,
    "MANUAL_LOCAL": MANUAL_LOCAL_FILL,
    "EXTERNAL_EVENT": EXTERNAL_FILL,
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def write_row(ws, row_num, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGN


def compose_step_command(step):
    """스텝의 command/description/target/duration 등을 하나의 셀에 응축."""
    parts = []
    if "command" in step:
        parts.append(str(step["command"]))
    if "target" in step:
        parts.append(f"target: {step['target']}")
    if "key" in step:
        parts.append(f"key: {step['key']}")
    if "name" in step and step.get("action") == "screenshot":
        parts.append(f"name: {step['name']}")
    if "description" in step:
        parts.append(step["description"])
    return "\n".join(parts)


def compose_duration(step):
    if "duration" in step:
        return f"{step['duration']}ms"
    if "timeout" in step:
        return f"timeout {step['timeout']}ms"
    return ""


def load_manifest(path):
    if not path:
        return {}
    with open(path, encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    if not isinstance(data, dict):
        raise SystemExit("ERROR: result manifest must be a mapping {tc_name: result}")
    return data


def main():
    parser = argparse.ArgumentParser(description="범용 yaml TC → Excel 리포트")
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--dir", dest="src_dir", help="yaml 디렉토리 (*.yaml 전체)")
    src.add_argument("--files", nargs="+", help="yaml 파일 명시 리스트")
    parser.add_argument("--out", required=True, help="출력 xlsx 경로")
    parser.add_argument("--title", default="TC Report", help="시트 제목 prefix (보고용)")
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="--dir 사용 시 하위 디렉토리 재귀 (functional/<카테고리>/ 분산 구조용)",
    )
    parser.add_argument(
        "--exclude",
        nargs="+",
        default=[],
        help="제외할 경로 부분 문자열 (예: catalog reports)",
    )
    parser.add_argument(
        "--result-manifest",
        dest="manifest",
        help="검증결과 매핑 yaml (선택). 형식: {TC_NAME: '결과 문자열'}",
    )
    args = parser.parse_args()

    if args.src_dir:
        base = Path(args.src_dir)
        candidates = base.rglob("*.yaml") if args.recursive else base.glob("*.yaml")
        tc_files = sorted(candidates)
    else:
        tc_files = [Path(f) for f in args.files]

    if args.exclude:
        tc_files = [
            p for p in tc_files if not any(ex in str(p) for ex in args.exclude)
        ]

    if not tc_files:
        raise SystemExit("ERROR: 대상 TC 파일이 없습니다")

    results = load_manifest(args.manifest)

    wb = openpyxl.Workbook()

    # ── Sheet 1: TC 요약 ──
    ws_sum = wb.active
    ws_sum.title = "TC 요약"
    sum_headers = [
        "TC ID",
        "설명",
        "tc_class",
        "execution_type",
        "manual_detail",
        "runnable",
        "step 수",
        "검증결과",
        "출처",
        "단말",
        "앱 (package)",
        "앱 버전",
        "사전조건",
    ]
    style_header(ws_sum, sum_headers)

    # ── Sheet 2: 스텝 상세 ──
    ws_steps = wb.create_sheet("스텝 상세")
    step_headers = [
        "TC ID",
        "Step#",
        "action",
        "step_role",
        "command / target / description",
        "execution_mode",
        "compile_status",
        "duration / timeout",
        "lint_allow",
    ]
    style_header(ws_steps, step_headers)

    # ── Sheet 3: 수행 방식 분류 ──
    ws_class = wb.create_sheet("수행 방식 분류")
    class_headers = [
        "TC ID",
        "tc_class",
        "execution_type",
        "manual_detail",
        "has_manual_steps",
        "has_shell_actions",
        "has_unresolved_params",
        "비고",
    ]
    style_header(ws_class, class_headers)

    sum_row = 2
    step_row = 2
    class_row = 2

    for f in tc_files:
        with open(f, encoding="utf-8") as fh:
            tc = yaml.safe_load(fh)
        if not tc:
            continue

        tc_id = tc.get("tc_name") or f.stem
        desc = (tc.get("description") or "").strip()
        meta = tc.get("metadata") or {}
        steps = tc.get("steps") or []
        precs = tc.get("preconditions") or []

        tc_class = meta.get("tc_class", "")
        exec_type = meta.get("execution_type", "")
        manual_detail = meta.get("manual_detail", "")
        runnable = "O" if meta.get("runnable") else "X"
        target_app = meta.get("target_app") or {}
        pkg = target_app.get("package", "")
        ver = target_app.get("version", "")
        device = meta.get("target_device", "")
        source = meta.get("source", "")
        result = results.get(tc_id, "")

        # Sheet 1
        write_row(
            ws_sum,
            sum_row,
            [
                tc_id,
                desc,
                tc_class,
                exec_type,
                manual_detail,
                runnable,
                len(steps),
                result,
                source,
                device,
                pkg,
                ver,
                "\n".join(f"- {p}" for p in precs),
            ],
        )
        # 결과 / execution_type 색상
        if result:
            r_cell = ws_sum.cell(row=sum_row, column=8)
            r_lower = result.lower()
            if "fail" in r_lower or "block" in r_lower:
                r_cell.fill = FAIL_FILL
            elif "pass" in r_lower:
                r_cell.fill = PASS_FILL
            r_cell.alignment = CENTER_ALIGN
        et_cell = ws_sum.cell(row=sum_row, column=4)
        if exec_type in EXEC_TYPE_FILL:
            et_cell.fill = EXEC_TYPE_FILL[exec_type]
        et_cell.alignment = CENTER_ALIGN

        ws_sum.cell(row=sum_row, column=6).alignment = CENTER_ALIGN
        ws_sum.cell(row=sum_row, column=7).alignment = CENTER_ALIGN
        sum_row += 1

        # Sheet 2
        for idx, step in enumerate(steps, 1):
            role = step.get("step_role", "")
            lint = step.get("lint_allow") or []
            write_row(
                ws_steps,
                step_row,
                [
                    tc_id,
                    idx,
                    step.get("action", ""),
                    role,
                    compose_step_command(step),
                    step.get("execution_mode", ""),
                    step.get("compile_status", ""),
                    compose_duration(step),
                    ", ".join(lint) if lint else "",
                ],
            )
            role_cell = ws_steps.cell(row=step_row, column=4)
            if role in STEP_ROLE_FILL:
                role_cell.fill = STEP_ROLE_FILL[role]
            role_cell.alignment = CENTER_ALIGN
            ws_steps.cell(row=step_row, column=2).alignment = CENTER_ALIGN
            step_row += 1

        # Sheet 3
        write_row(
            ws_class,
            class_row,
            [
                tc_id,
                tc_class,
                exec_type,
                manual_detail,
                "O" if meta.get("has_manual_steps") else "X",
                "O" if meta.get("has_shell_actions") else "X",
                "O" if meta.get("has_unresolved_params") else "X",
                "",
            ],
        )
        for col in (2, 3, 4, 5, 6, 7):
            ws_class.cell(row=class_row, column=col).alignment = CENTER_ALIGN
        et_cell2 = ws_class.cell(row=class_row, column=3)
        if exec_type in EXEC_TYPE_FILL:
            et_cell2.fill = EXEC_TYPE_FILL[exec_type]
        class_row += 1

    # 컬럼 폭 설정
    sum_widths = [36, 60, 14, 14, 14, 10, 9, 22, 38, 28, 26, 16, 38]
    for c, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    step_widths = [36, 7, 22, 11, 60, 14, 14, 16, 22]
    for c, w in enumerate(step_widths, 1):
        ws_steps.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    class_widths = [36, 14, 14, 14, 16, 16, 18, 30]
    for c, w in enumerate(class_widths, 1):
        ws_class.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # freeze panes
    ws_sum.freeze_panes = "A2"
    ws_steps.freeze_panes = "C2"
    ws_class.freeze_panes = "A2"

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(
        f"OK: {len(tc_files)} TC → {out_path} "
        f"(요약 {sum_row - 2} 행, 스텝 {step_row - 2} 행)"
    )


if __name__ == "__main__":
    main()
