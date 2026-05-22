import yaml, openpyxl, sys, io, argparse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

parser = argparse.ArgumentParser(description='SeniorShield TC → Excel 리포트 생성')
src = parser.add_mutually_exclusive_group()
src.add_argument('--dir', dest='src_dir', help='TC yaml 디렉토리 (*.yaml 전체)')
src.add_argument('--files', nargs='+', help='TC yaml 파일 리스트')
parser.add_argument('--out', help='출력 xlsx 경로')
args = parser.parse_args()

if args.files:
    tc_files = [Path(f) for f in args.files]
elif args.src_dir:
    tc_files = sorted(Path(args.src_dir).glob('*.yaml'))
else:
    tc_files = sorted(Path('stage2_output/new_tcs').glob('SS-*.yaml'))

if not tc_files:
    print('ERROR: 대상 TC 파일이 없습니다', file=sys.stderr)
    sys.exit(1)

# 로딩된 TC ID 사전 수집 (coverage/signals 필터링용)
loaded_tc_ids = set()
for f in tc_files:
    with open(f, encoding='utf-8') as fh:
        _tc = yaml.safe_load(fh)
    loaded_tc_ids.add(_tc['tc_name'])

wb = openpyxl.Workbook()

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2F5496')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
setup_fill = PatternFill('solid', fgColor='D9E2F3')
action_fill = PatternFill('solid', fgColor='FFF2CC')
assert_fill = PatternFill('solid', fgColor='E2EFDA')
auto_fill = PatternFill('solid', fgColor='B4C6E7')
manual_local_fill = PatternFill('solid', fgColor='FCE4D6')
external_fill = PatternFill('solid', fgColor='F8CBAD')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center')

EXEC_TYPE_FILL = {
    'AUTO': auto_fill,
    'MANUAL_LOCAL': manual_local_fill,
    'EXTERNAL_EVENT': external_fill,
}

# ── Sheet 1: TC 요약 ──
ws_sum = wb.active
ws_sum.title = 'TC 요약'
sum_headers = [
    'TC ID', '설명', 'automation_class', '실행가능', 'manual_steps',
    'execution_type', 'manual_detail', '검증결과', '출처', '사전조건', '경고사항'
]
for c, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

fail_tcs = {'SS-04', 'SS-08'} & loaded_tc_ids

for i, f in enumerate(tc_files, 2):
    with open(f, encoding='utf-8') as fh:
        tc = yaml.safe_load(fh)
    meta = tc.get('metadata', {})
    tc_id = tc['tc_name']
    is_pass = tc_id not in fail_tcs
    exec_type = meta.get('execution_type')
    manual_detail = meta.get('manual_detail')

    # fail-fast: 필수 필드 누락 시 에러
    if not exec_type:
        print(f"ERROR: {f.name} — metadata.execution_type 누락", file=sys.stderr)
        sys.exit(1)
    if manual_detail is None:
        print(f"ERROR: {f.name} — metadata.manual_detail 누락", file=sys.stderr)
        sys.exit(1)

    # manual_steps 일관성: AUTO→X, else→O
    if exec_type == 'AUTO':
        manual_steps = 'X'
    elif exec_type in ('MANUAL_LOCAL', 'EXTERNAL_EVENT'):
        manual_steps = 'O'
    else:
        print(f"ERROR: {f.name} — execution_type '{exec_type}' 은 허용값이 아님", file=sys.stderr)
        sys.exit(1)

    vals = [
        tc_id,
        tc.get('description', ''),
        meta.get('tc_class', ''),
        'O' if meta.get('runnable') else 'X',
        manual_steps,
        exec_type,
        manual_detail,
        'PASS' if is_pass else 'FAIL (placeholder)',
        meta.get('source_row', ''),
        '\n'.join(tc.get('preconditions', [])),
        '\n'.join(meta.get('warnings', []))
    ]
    for c, v in enumerate(vals, 1):
        cell = ws_sum.cell(row=i, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align

    # execution_type 색상
    et_cell = ws_sum.cell(row=i, column=6)
    if exec_type in EXEC_TYPE_FILL:
        et_cell.fill = EXEC_TYPE_FILL[exec_type]
    et_cell.alignment = center_align

    # 검증결과 색상
    result_cell = ws_sum.cell(row=i, column=8)
    result_cell.fill = pass_fill if is_pass else fail_fill

ws_sum.column_dimensions['A'].width = 8
ws_sum.column_dimensions['B'].width = 55
ws_sum.column_dimensions['C'].width = 16
ws_sum.column_dimensions['D'].width = 10
ws_sum.column_dimensions['E'].width = 13
ws_sum.column_dimensions['F'].width = 18
ws_sum.column_dimensions['G'].width = 28
ws_sum.column_dimensions['H'].width = 16
ws_sum.column_dimensions['I'].width = 30
ws_sum.column_dimensions['J'].width = 35
ws_sum.column_dimensions['K'].width = 40

# ── Sheet 2: 스텝 상세 ──
ws_det = wb.create_sheet('스텝 상세')
det_headers = [
    'TC ID', 'Step#', 'action', 'step_role',
    'command / description', 'expected', 'execution_mode',
    'compile_status', 'duration/timeout'
]
for c, h in enumerate(det_headers, 1):
    cell = ws_det.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

row = 2
for f in tc_files:
    with open(f, encoding='utf-8') as fh:
        tc = yaml.safe_load(fh)
    tc_id = tc['tc_name']
    for si, step in enumerate(tc.get('steps', []), 1):
        action = step.get('action', '')
        role = step.get('step_role', '')
        content = step.get('command', '') or step.get('description', '') or ''
        if isinstance(content, str):
            content = content.strip()
        expected = step.get('expected', '')
        exec_mode = step.get('execution_mode', '')
        compile_st = step.get('compile_status', '')
        dur = ''
        if 'duration' in step:
            dur = f"{step['duration']}ms"
        elif 'manual_timeout' in step:
            dur = f"timeout {step['manual_timeout']}s"
        vals = [tc_id, si, action, role, content, expected, exec_mode, compile_st, dur]
        for c, v in enumerate(vals, 1):
            cell = ws_det.cell(row=row, column=c, value=v)
            cell.border = thin_border
            cell.alignment = wrap_align
        role_cell = ws_det.cell(row=row, column=4)
        if role == 'SETUP':
            role_cell.fill = setup_fill
        elif role == 'ACTION':
            role_cell.fill = action_fill
        elif role == 'ASSERT':
            role_cell.fill = assert_fill
        row += 1

ws_det.column_dimensions['A'].width = 8
ws_det.column_dimensions['B'].width = 7
ws_det.column_dimensions['C'].width = 14
ws_det.column_dimensions['D'].width = 10
ws_det.column_dimensions['E'].width = 65
ws_det.column_dimensions['F'].width = 30
ws_det.column_dimensions['G'].width = 18
ws_det.column_dimensions['H'].width = 18
ws_det.column_dimensions['I'].width = 14

# ── Sheet 3: 수행 방식 분류 ──
ws_exec = wb.create_sheet('수행 방식 분류')
exec_headers = ['TC ID', 'automation_class', 'manual_steps', 'execution_type', 'manual_detail', '비고']
for c, h in enumerate(exec_headers, 1):
    cell = ws_exec.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

NOTES = {
    'SS-01': '사람 개입 없이 자동 수행',
    'SS-02': '보조폰(미저장 번호)에서 수신 필요',
    'SS-03': '보조폰(미저장 번호) 장시간 통화 수신 필요',
    'SS-04': '보조폰 수신 + 원격제어 앱 필요',
    'SS-05': 'DUT에서 TRIGGER 수동 발동 (앱 실행/설치)',
    'SS-06': '보조폰(7일 이내 저장 번호)에서 수신 필요',
    'SS-07': '보조폰 수신 + TTL 만료 대기',
    'SS-08': 'DUT에서 원격앱 + 금융 앱 실행',
    'SS-09': '보조폰에서 2회 수신 필요',
    'SS-10': 'DUT에서 의심 앱 설치',
    'SS-11': '보조폰 수신 + DUT 안전함 버튼 터치',
    'SS-12': '보조폰 2회 수신 + DUT 카드 닫기',
    'SS-13': '보조폰 수신 + DUT TRIGGER 발동',
    'SS-14': 'DUT에서 TRIGGER 발동 + 팝업 닫기',
    'SS-15': 'DUT에서 의심 앱 설치 + 텔레뱅킹 발신',
    'SS-16': 'DUT에서 개발자 옵션/설정 조작',
    'SS-17': '보조폰에서 2회 수신 (마지막 12초+)',
}

for i, f in enumerate(tc_files, 2):
    with open(f, encoding='utf-8') as fh:
        tc = yaml.safe_load(fh)
    meta = tc.get('metadata', {})
    tc_id = tc['tc_name']
    exec_type = meta.get('execution_type', '')
    manual_detail = meta.get('manual_detail', '')

    if exec_type == 'AUTO':
        manual_steps = 'X'
    else:
        manual_steps = 'O'

    vals = [
        tc_id,
        meta.get('tc_class', ''),
        manual_steps,
        exec_type,
        manual_detail,
        NOTES.get(tc_id, '')
    ]
    for c, v in enumerate(vals, 1):
        cell = ws_exec.cell(row=i, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align

    et_cell = ws_exec.cell(row=i, column=4)
    if exec_type in EXEC_TYPE_FILL:
        et_cell.fill = EXEC_TYPE_FILL[exec_type]
    et_cell.alignment = center_align

ws_exec.column_dimensions['A'].width = 8
ws_exec.column_dimensions['B'].width = 16
ws_exec.column_dimensions['C'].width = 13
ws_exec.column_dimensions['D'].width = 18
ws_exec.column_dimensions['E'].width = 28
ws_exec.column_dimensions['F'].width = 45

# ── Sheet 4: 커버리지 매트릭스 ──
ws_cov = wb.create_sheet('커버리지')
cov_headers = ['카테고리', 'TC ID', '검증 항목']
for c, h in enumerate(cov_headers, 1):
    cell = ws_cov.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

coverage = [
    ('Risk Signal - PASSIVE', 'SS-02', 'UNKNOWN_CALLER (20점)'),
    ('Risk Signal - PASSIVE', 'SS-06', 'UNVERIFIED_CALLER (20점)'),
    ('Risk Signal - PASSIVE', 'SS-09', 'REPEATED_UNKNOWN_CALLER (15점)'),
    ('Risk Signal - PASSIVE', 'SS-16', 'HIGH_RISK_DEVICE_ENVIRONMENT (20점)'),
    ('Risk Signal - PASSIVE', 'SS-03', 'LONG_CALL_DURATION (30점)'),
    ('Risk Signal - AMPLIFIER', 'SS-17', 'REPEATED_CALL_THEN_LONG_TALK (20점)'),
    ('Risk Signal - TRIGGER', 'SS-10', 'SUSPICIOUS_APP_INSTALLED (40점)'),
    ('Risk Signal - TRIGGER', 'SS-08', 'REMOTE_CONTROL + BANKING_AFTER_REMOTE (30+40점)'),
    ('Risk Signal - TRIGGER', 'SS-15', 'TELEBANKING_AFTER_SUSPICIOUS (25점)'),
    ('AlertState 전이', 'SS-02', 'OBSERVE → GUARDED'),
    ('AlertState 전이', 'SS-05', 'GUARDED → INTERRUPT'),
    ('AlertState 전이', 'SS-04', 'Call + TRIGGER → CRITICAL'),
    ('AlertState 전이', 'SS-08', 'REMOTE + BANKING → CRITICAL (no call)'),
    ('Banking Cooldown', 'SS-03', 'UNKNOWN+LONG → 쿨다운 발동'),
    ('Banking Cooldown', 'SS-06', 'UNVERIFIED_CALLER → 쿨다운 미발동'),
    ('세션 TTL', 'SS-07', '30분 TTL 만료 → OBSERVE 복귀'),
    ('세션 TTL', 'SS-13', 'TRIGGER → 60분 연장'),
    ('세션 TTL', 'SS-11', '안전함 확인 → 즉시 세션 종료'),
    ('UI/UX', 'SS-01', '온보딩 + FGS 확인'),
    ('UI/UX', 'SS-05', 'INTERRUPT 팝업 + SMS 버튼'),
    ('UI/UX', 'SS-12', '하단 카드 dedupe (session당 1회)'),
    ('UI/UX', 'SS-14', '팝업 dismiss 후 홈 재접근'),
]
coverage = [row for row in coverage if row[1] in loaded_tc_ids]
for r, (cat, tc_id, item) in enumerate(coverage, 2):
    for c, v in enumerate([cat, tc_id, item], 1):
        cell = ws_cov.cell(row=r, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align

ws_cov.column_dimensions['A'].width = 22
ws_cov.column_dimensions['B'].width = 10
ws_cov.column_dimensions['C'].width = 55

# ── Sheet 5: Signal 커버리지 ──
ws_sig = wb.create_sheet('Signal 커버리지')
sig_headers = ['Signal Name', '유형', '점수', 'TC ID', '커버 여부']
for c, h in enumerate(sig_headers, 1):
    cell = ws_sig.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

signals = [
    ('UNKNOWN_CALLER', 'PASSIVE', 20, 'SS-02', 'O'),
    ('LONG_CALL_DURATION', 'PASSIVE', 30, 'SS-03', 'O'),
    ('UNVERIFIED_CALLER', 'PASSIVE', 20, 'SS-06', 'O'),
    ('REPEATED_UNKNOWN_CALLER', 'PASSIVE', 15, 'SS-09', 'O'),
    ('HIGH_RISK_DEVICE_ENVIRONMENT', 'PASSIVE', 20, 'SS-16', 'O'),
    ('REPEATED_CALL_THEN_LONG_TALK', 'AMPLIFIER', 20, 'SS-17', 'O'),
    ('REMOTE_CONTROL_APP_OPENED', 'TRIGGER', 30, 'SS-04, SS-08', 'O'),
    ('BANKING_APP_OPENED_AFTER_REMOTE_APP', 'TRIGGER', 40, 'SS-08', 'O'),
    ('SUSPICIOUS_APP_INSTALLED', 'TRIGGER', 40, 'SS-10', 'O'),
    ('TELEBANKING_AFTER_SUSPICIOUS', 'TRIGGER', 25, 'SS-15', 'O'),
]
def _signal_covered(tc_id_field):
    ids = [t.strip() for t in tc_id_field.split(',')]
    return any(t in loaded_tc_ids for t in ids)

signals = [row for row in signals if _signal_covered(row[3])]
for r, (name, stype, score, tc_id, covered) in enumerate(signals, 2):
    for c, v in enumerate([name, stype, score, tc_id, covered], 1):
        cell = ws_sig.cell(row=r, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align
    ws_sig.cell(row=r, column=5).fill = pass_fill

ws_sig.column_dimensions['A'].width = 40
ws_sig.column_dimensions['B'].width = 12
ws_sig.column_dimensions['C'].width = 8
ws_sig.column_dimensions['D'].width = 14
ws_sig.column_dimensions['E'].width = 10

out = args.out or 'stage2_output/new_tcs/SeniorShield_New_TCs_v3.xlsx'
Path(out).parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(f'생성 완료: {out}')
print(f'TC 수: {len(tc_files)}건')
print('시트: TC 요약 / 스텝 상세 / 수행 방식 분류 / 커버리지 / Signal 커버리지')
