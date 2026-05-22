import openpyxl, sys, io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.Workbook()

# ── 스타일 ──
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2F5496')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
gallery_fill = PatternFill('solid', fgColor='D9E2F3')
memosy_fill = PatternFill('solid', fgColor='FCE4D6')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

def write_row(ws, row_num, values, app_fill=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align
    # 앱 구분 색상
    if app_fill:
        ws.cell(row=row_num, column=2).fill = app_fill
    # 결과 색상
    result_col = len(values)
    result_cell = ws.cell(row=row_num, column=result_col)
    if result_cell.value == 'PASS':
        result_cell.fill = pass_fill
    elif result_cell.value == 'FAIL':
        result_cell.fill = fail_fill
    result_cell.alignment = center_align


# ══════════════════════════════════════════
# Sheet 1: TC 요약
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = 'TC 요약'
write_header(ws1, ['TC ID', '앱', '기능', '테스트 설명', '사전조건', '결과'])

tcs = [
    # Gallery
    ('G-01', 'Gallery', '사진 목록', '메인 화면에서 날짜별 그룹핑된 사진 목록 표시 확인', '갤러리 앱 설치, 미디어 파일 존재', 'PASS'),
    ('G-02', 'Gallery', '사진 상세보기', '사진 탭 → 이미지 항목 전체화면 표시 확인', 'G-01 통과', 'PASS'),
    ('G-03', 'Gallery', '스와이프 넘기기', '상세보기에서 좌→우 스와이프로 다음 사진 이동', 'G-02 통과', 'PASS'),
    ('G-04', 'Gallery', '즐겨찾기', '상세보기에서 즐겨찾기 아이콘 탭 → 토글 동작, 메인 목록에 즐겨찾기 섹션 생성', 'G-02 통과', 'PASS'),
    ('G-05', 'Gallery', '공유', '상세보기에서 공유 버튼 탭 → 시스템 공유 시트(Quick Share, 메시지, 블루투스 등) 호출', 'G-02 통과', 'PASS'),
    ('G-06', 'Gallery', '편집', '상세보기에서 편집 버튼 탭 → 미디어 편집 화면(회전/자르기/지우기/보정) 진입', 'G-02 통과', 'PASS'),
    ('G-07', 'Gallery', '상세 정보', '더보기 메뉴 → 상세 정보 → 날짜/파일명/경로/해상도 표시', 'G-02 통과', 'PASS'),
    ('G-08', 'Gallery', '배경화면 설정', '더보기 메뉴 → 배경화면으로 설정 → 홈/잠금/홈+잠금 선택 다이얼로그 표시', 'G-02 통과', 'PASS'),
    ('G-09', 'Gallery', '동영상 재생', '동영상 항목 탭 → 인앱 플레이어에서 재생 확인', '동영상 파일 존재', 'PASS'),
    ('G-10', 'Gallery', '롱프레스 다중 선택', '사진 항목 롱프레스 → 체크 표시 + 하단 공유/삭제 액션바 표시', 'G-01 통과', 'PASS'),
    ('G-11', 'Gallery', '앨범 탭', '하단 앨범 탭 전환 → Movies/Screenshots 등 폴더 목록 표시', '미디어 파일 존재', 'PASS'),
    ('G-12', 'Gallery', '앨범 내부 진입', '앨범(Screenshots) 탭 → 항목 수/용량 표시, 내부 사진 목록', 'G-11 통과', 'PASS'),
    ('G-13', 'Gallery', '앨범 정렬', '앨범 탭에서 최신순 드롭다운 → 5가지 정렬 옵션(최신순/이름순 ㄱ-ㅎ/ㅎ-ㄱ/수정날짜순/크기순)', 'G-11 통과', 'PASS'),
    ('G-14', 'Gallery', '휴지통 탭', '하단 휴지통 탭 전환 → 빈 상태 메시지("미디어가 없습니다") 표시', '앱 실행', 'PASS'),
    # Memosy
    ('M-01', 'Memosy', '메모 작성', '+ 버튼 → 새 메모 화면 → 제목/내용 텍스트 입력 → 저장', '앱 실행', 'PASS'),
    ('M-02', 'Memosy', '배경 색상 변경', '하단 팔레트 아이콘 → 색상 원형 선택 → 메모 배경 즉시 반영', 'M-01 통과', 'PASS'),
    ('M-03', 'Memosy', '핀 고정', '상단 핀 아이콘 탭 → 토글 활성화, 목록에서 핀 아이콘 표시', 'M-01 통과', 'PASS'),
    ('M-04', 'Memosy', '공유', '하단 공유 아이콘 → 텍스트 공유 시트([제목]\\n내용 형식) + 복사 버튼 + 앱 목록', 'M-01 통과', 'PASS'),
    ('M-05', 'Memosy', '음성 녹음', '상단 마이크 아이콘 → 녹음 시작(타이머 표시) → 정지 → 재생 버튼(00:13) 저장', 'RECORD_AUDIO 권한 부여', 'PASS'),
    ('M-06', 'Memosy', '메모 저장 + 목록', '뒤로가기 → 메인 목록에 카드 표시(제목/내용/시간/핀/마이크 아이콘/배경색)', '메모 작성 완료', 'PASS'),
    ('M-07', 'Memosy', '검색', '돋보기 아이콘 → "Test" 입력 → 실시간 필터링으로 해당 메모 표시', '메모 1건 이상 존재', 'PASS'),
    ('M-08', 'Memosy', '보기 전환', '상단 보기 전환 아이콘 탭 → 그리드 ↔ 리스트 레이아웃 전환', '메모 1건 이상 존재', 'PASS'),
    ('M-09', 'Memosy', '롱프레스 다중 선택', '메모 카드 롱프레스 → 선택 모드(선택 해제/공유하기/전체 보관/삭제)', '메모 1건 이상 존재', 'PASS'),
    ('M-10', 'Memosy', '보관함 이동', '롱프레스 → 전체 보관 → 메인 목록에서 제거, 보관함(1)에 표시', 'M-09 통과', 'PASS'),
    ('M-11', 'Memosy', '삭제 (휴지통 이동)', '보관함에서 롱프레스 → 삭제 → 휴지통(1)에 메모 표시', 'M-10 통과', 'PASS'),
    ('M-12', 'Memosy', '휴지통 복원', '휴지통에서 롱프레스 → 전체 복원 → 원래 위치(보관함)로 복귀', 'M-11 통과', 'PASS'),
    ('M-13', 'Memosy', '메모 내 삭제 버튼', '메모 열기 → 하단 휴지통 아이콘 탭 → 즉시 삭제 → 목록에서 제거', '메모 존재', 'PASS'),
    ('M-14', 'Memosy', '휴지통 비우기', '휴지통에서 롱프레스 → 선택 해제/전체 복원/휴지통 비우기 옵션 표시', '휴지통에 메모 존재', 'PASS'),
]

for i, (tc_id, app, func, desc, precond, result) in enumerate(tcs, 2):
    app_f = gallery_fill if app == 'Gallery' else memosy_fill
    write_row(ws1, i, [tc_id, app, func, desc, precond, result], app_fill=app_f)

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 65
ws1.column_dimensions['E'].width = 30
ws1.column_dimensions['F'].width = 10


# ══════════════════════════════════════════
# Sheet 2: 스텝 상세
# ══════════════════════════════════════════
ws2 = wb.create_sheet('스텝 상세')
write_header(ws2, ['TC ID', 'Step#', '동작', '상세 절차', '기대 결과', '실제 결과', '판정'])

steps = [
    # G-01
    ('G-01', 1, 'adb 명령', 'adb shell am start -n com.example.mygalleryapp/.MainActivity', '갤러리 앱 실행', '앱 실행됨, 사진 탭 표시', 'PASS'),
    ('G-01', 2, 'UI 확인', '메인 화면에서 날짜별 헤더(2026년 04월 03일 등) 확인', '날짜별 그룹핑 표시', '2026-04-03, 03-12, 03-11 그룹 표시', 'PASS'),
    # G-02
    ('G-02', 1, '탭', '사진 항목(이미지) 탭', '상세보기 화면 전환', '전체화면 사진 표시, 상단바(뒤로/공유/편집/즐겨찾기/더보기)', 'PASS'),
    # G-03
    ('G-03', 1, '스와이프', '상세보기에서 좌→우 스와이프 (400→80, y=400)', '다음 사진으로 이동', '스크린샷(소프트웨어 업데이트) 이미지로 전환됨', 'PASS'),
    # G-04
    ('G-04', 1, '탭', '상세보기 상단 즐겨찾기(별) 아이콘 탭', '아이콘 토글(활성화)', '별 아이콘 채워짐', 'PASS'),
    ('G-04', 2, 'UI 확인', '메인 목록으로 복귀', '"즐겨찾기" 섹션 생성', '즐겨찾기 섹션 상단에 표시됨', 'PASS'),
    # G-05
    ('G-05', 1, '탭', '상세보기 공유 버튼(btnShare) 탭', '시스템 공유 시트 호출', '"이미지 공유" 시트 + Quick Share/Files/메시지/블루투스 표시', 'PASS'),
    # G-06
    ('G-06', 1, '탭', '상세보기 편집 버튼(btnEdit) 탭', '편집 화면 진입', '"미디어 편집" 화면 + 회전/자르기/지우기/보정 탭 표시', 'PASS'),
    # G-07
    ('G-07', 1, '탭', '더보기(btnMore) → 상세 정보 탭', '파일 상세 정보 표시', '날짜: 2026-04-03, 파일명, 경로, 해상도: 480x800', 'PASS'),
    # G-08
    ('G-08', 1, '탭', '더보기 → 배경화면으로 설정 탭', '배경화면 선택 다이얼로그', '홈 화면 / 잠금 화면 / 홈 및 잠금 화면 3가지 옵션 표시', 'PASS'),
    # G-09
    ('G-09', 1, '탭', '동영상 항목(00:46 표시) 탭', '비디오 재생', 'DetailActivity에서 인앱 비디오 플레이어 재생됨', 'PASS'),
    # G-10
    ('G-10', 1, '롱프레스', '사진 항목 롱프레스 (1.5초)', '다중 선택 모드 진입', '항목에 체크 표시, 하단 공유/삭제 액션바 표시', 'PASS'),
    # G-11
    ('G-11', 1, '탭', '하단 앨범 탭 전환', '앨범 목록 표시', 'Movies(1개), Screenshots(3개) 폴더 카드 표시', 'PASS'),
    # G-12
    ('G-12', 1, '탭', 'Screenshots 앨범 탭', '앨범 내부 진입', '"Screenshots - 항목 3개 · 0.4MB" + 사진 목록 표시', 'PASS'),
    # G-13
    ('G-13', 1, '탭', '앨범 탭 → 최신순 드롭다운 탭', '정렬 옵션 표시', '최신순/이름순(ㄱ-ㅎ)/이름순(ㅎ-ㄱ)/수정 날짜순/크기순', 'PASS'),
    # G-14
    ('G-14', 1, '탭', '하단 휴지통 탭 전환', '휴지통 화면 표시', '"미디어가 없습니다" 빈 상태 메시지', 'PASS'),
    # M-01
    ('M-01', 1, 'adb 명령', 'adb shell am start -n com.example.memosy/.MainActivity', '앱 실행', '"메모 (0)" 메인 화면 표시', 'PASS'),
    ('M-01', 2, '탭', '+ FAB 버튼 탭', '새 메모 화면 진입', '"새 메모" 화면 + 제목/내용 입력 필드', 'PASS'),
    ('M-01', 3, '텍스트 입력', '제목: "Test_Memo_Title", 내용: "Hello_this_is_test_content"', '텍스트 입력됨', '제목/내용 정상 입력, 수정일 자동 표시', 'PASS'),
    # M-02
    ('M-02', 1, '탭', '하단 팔레트 아이콘 탭', '색상 선택 팔레트 표시', '기본/빨강/분홍/보라/하늘 등 색상 원형 표시', 'PASS'),
    ('M-02', 2, '탭', '빨강(분홍) 색상 원형 탭', '메모 배경색 변경', '배경이 분홍색으로 즉시 변경됨', 'PASS'),
    # M-03
    ('M-03', 1, '탭', '상단 핀 아이콘 탭', '핀 고정 토글', '핀 아이콘 활성화(파란색), 목록 카드에 핀 표시', 'PASS'),
    # M-04
    ('M-04', 1, '탭', '하단 공유 아이콘 탭', '공유 시트 호출', '"텍스트 공유" + [Test_Memo_Title]\\nHello... 미리보기 + 복사 + 앱 목록', 'PASS'),
    # M-05
    ('M-05', 1, '탭', '상단 마이크 아이콘 탭', '음성 녹음 시작', '마이크 아이콘 녹색 활성화, 빨간 정지 버튼, 녹음 타이머(00:02)', 'PASS'),
    ('M-05', 2, '탭', '정지 버튼 탭 (13초 후)', '녹음 저장', '재생 버튼 표시(00:00 / 00:13)', 'PASS'),
    # M-06
    ('M-06', 1, '탭', '뒤로(←) 버튼 탭', '메인 목록 복귀', '"메모 (1)" + 카드(제목/내용/시간/핀📌/마이크🎤/분홍배경) 표시', 'PASS'),
    # M-07
    ('M-07', 1, '탭', '돋보기 아이콘 탭', '검색 화면 진입', '"메모 검색" 입력 필드 + 키보드 표시', 'PASS'),
    ('M-07', 2, '텍스트 입력', '"Test" 입력', '실시간 필터링', 'Test_Memo_Title 메모 카드 표시됨', 'PASS'),
    # M-08
    ('M-08', 1, '탭', '상단 보기 전환 아이콘 탭', '레이아웃 전환', '그리드 → 리스트 뷰, 전체 폭 카드로 변경, 아이콘 변경', 'PASS'),
    # M-09
    ('M-09', 1, '롱프레스', '메모 카드 롱프레스 (1.5초)', '다중 선택 모드', '체크 표시 + 하단 바(선택 해제/공유하기/전체 보관/삭제)', 'PASS'),
    # M-10
    ('M-10', 1, '탭', '하단 "전체 보관" 버튼 탭', '보관함으로 이동', '"메모 (0)" + 메인에서 제거됨', 'PASS'),
    ('M-10', 2, 'UI 확인', '햄버거 메뉴 → 보관함 진입', '보관함에 메모 존재', '"보관함 (1)" + 메모 카드 표시', 'PASS'),
    # M-11
    ('M-11', 1, '롱프레스', '보관함에서 메모 롱프레스', '선택 모드', '선택 해제/전체 보관 해제/삭제 버튼 표시', 'PASS'),
    ('M-11', 2, '탭', '"삭제" 버튼 탭', '휴지통으로 이동', '"보관함 (0)" + 제거됨', 'PASS'),
    ('M-11', 3, 'UI 확인', '햄버거 메뉴 → 휴지통 진입', '휴지통에 메모 존재', '"휴지통 (1)" + 메모 카드 표시', 'PASS'),
    # M-12
    ('M-12', 1, '롱프레스', '휴지통에서 메모 롱프레스', '선택 모드', '선택 해제/전체 복원/휴지통 비우기 버튼 표시', 'PASS'),
    ('M-12', 2, '탭', '"전체 복원" 버튼 탭', '원래 위치로 복원', '"보관함 (1)"에 메모 복귀 확인', 'PASS'),
    # M-13
    ('M-13', 1, '탭', '메모 열기 → 하단 휴지통 아이콘 탭', '즉시 삭제', '"보관함 (0)" + 메모 제거됨', 'PASS'),
    # M-14
    ('M-14', 1, 'UI 확인', '휴지통 롱프레스 → 옵션 확인', '3가지 옵션 표시', '선택 해제/전체 복원/휴지통 비우기 표시됨', 'PASS'),
]

for i, (tc_id, step_no, action, detail, expected, actual, result) in enumerate(steps, 2):
    app_f = gallery_fill if tc_id.startswith('G') else memosy_fill
    write_row(ws2, i, [tc_id, step_no, action, detail, expected, actual, result], app_fill=None)
    # TC ID 색상
    ws2.cell(row=i, column=1).fill = app_f

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 7
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 60
ws2.column_dimensions['E'].width = 35
ws2.column_dimensions['F'].width = 45
ws2.column_dimensions['G'].width = 8


# ══════════════════════════════════════════
# Sheet 3: 테스트 환경
# ══════════════════════════════════════════
ws3 = wb.create_sheet('테스트 환경')
write_header(ws3, ['항목', '값'])

env = [
    ('단말 모델', 'AT-M140'),
    ('Android 버전', '14'),
    ('테스트 일시', '2026-04-06 17:53~19:00'),
    ('테스트 방식', 'ADB + UIAutomator (자동화) + 재검증'),
    ('Gallery 패키지', 'com.example.mygalleryapp'),
    ('Gallery 버전', '1.0.26040614'),
    ('Gallery 권한', 'READ_MEDIA_IMAGES, READ_MEDIA_VIDEO, ACCESS_MEDIA_LOCATION, SET_WALLPAPER'),
    ('Memosy 패키지', 'com.example.memosy'),
    ('Memosy 버전', '1.0.26040610'),
    ('Memosy 권한', 'RECORD_AUDIO'),
    ('총 TC 수', '28건 (Gallery 14 + Memosy 14)'),
    ('총 스텝 수', f'{len(steps)}건'),
    ('PASS', '28건'),
    ('FAIL', '0건'),
    ('통과율', '100%'),
]

for i, (k, v) in enumerate(env, 2):
    ws3.cell(row=i, column=1, value=k).border = thin_border
    ws3.cell(row=i, column=1).alignment = wrap_align
    ws3.cell(row=i, column=1).font = Font(bold=True)
    ws3.cell(row=i, column=2, value=v).border = thin_border
    ws3.cell(row=i, column=2).alignment = wrap_align

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 55


# ══════════════════════════════════════════
# Sheet 4: 기능 커버리지
# ══════════════════════════════════════════
ws4 = wb.create_sheet('기능 커버리지')
write_header(ws4, ['앱', '카테고리', '기능', 'TC ID', '커버 여부'])

coverage = [
    ('Gallery', '미디어 탐색', '사진 목록 (날짜별)', 'G-01', 'O'),
    ('Gallery', '미디어 탐색', '사진 상세보기', 'G-02', 'O'),
    ('Gallery', '미디어 탐색', '스와이프 넘기기', 'G-03', 'O'),
    ('Gallery', '미디어 탐색', '동영상 재생', 'G-09', 'O'),
    ('Gallery', '미디어 관리', '즐겨찾기', 'G-04', 'O'),
    ('Gallery', '미디어 관리', '롱프레스 다중 선택', 'G-10', 'O'),
    ('Gallery', '미디어 관리', '휴지통', 'G-14', 'O'),
    ('Gallery', '공유/편집', '공유', 'G-05', 'O'),
    ('Gallery', '공유/편집', '편집 (회전/자르기/지우기/보정)', 'G-06', 'O'),
    ('Gallery', '공유/편집', '배경화면 설정', 'G-08', 'O'),
    ('Gallery', '정보/정렬', '상세 정보', 'G-07', 'O'),
    ('Gallery', '정보/정렬', '앨범 목록', 'G-11', 'O'),
    ('Gallery', '정보/정렬', '앨범 내부', 'G-12', 'O'),
    ('Gallery', '정보/정렬', '앨범 정렬 (5가지)', 'G-13', 'O'),
    ('Memosy', '메모 CRUD', '메모 작성 (제목+내용)', 'M-01', 'O'),
    ('Memosy', '메모 CRUD', '메모 저장 + 목록 표시', 'M-06', 'O'),
    ('Memosy', '메모 CRUD', '메모 내 삭제', 'M-13', 'O'),
    ('Memosy', '미디어/서식', '배경 색상 변경', 'M-02', 'O'),
    ('Memosy', '미디어/서식', '핀 고정', 'M-03', 'O'),
    ('Memosy', '미디어/서식', '음성 녹음 + 재생', 'M-05', 'O'),
    ('Memosy', '공유/검색', '공유', 'M-04', 'O'),
    ('Memosy', '공유/검색', '검색', 'M-07', 'O'),
    ('Memosy', 'UI/UX', '보기 전환 (그리드↔리스트)', 'M-08', 'O'),
    ('Memosy', 'UI/UX', '롱프레스 다중 선택', 'M-09', 'O'),
    ('Memosy', '데이터 관리', '보관함 이동', 'M-10', 'O'),
    ('Memosy', '데이터 관리', '삭제 (휴지통 이동)', 'M-11', 'O'),
    ('Memosy', '데이터 관리', '휴지통 복원', 'M-12', 'O'),
    ('Memosy', '데이터 관리', '휴지통 비우기', 'M-14', 'O'),
]

for i, (app, cat, func, tc_id, covered) in enumerate(coverage, 2):
    app_f = gallery_fill if app == 'Gallery' else memosy_fill
    vals = [app, cat, func, tc_id, covered]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=i, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align
    ws4.cell(row=i, column=1).fill = app_f
    ws4.cell(row=i, column=5).fill = pass_fill
    ws4.cell(row=i, column=5).alignment = center_align

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 35
ws4.column_dimensions['D'].width = 8
ws4.column_dimensions['E'].width = 10


# ══════════════════════════════════════════
# Sheet 5: 하드키 테스트
# ══════════════════════════════════════════
ws5 = wb.create_sheet('하드키 테스트')
write_header(ws5, ['TC ID', '앱', '키', '화면', '테스트 절차', '기대 결과', '실제 결과', '판정'])

warn_fill = PatternFill('solid', fgColor='FFF2CC')

hk_tcs = [
    # ── Gallery 하드키 ──
    ('HK-G01', 'Gallery', 'KEY_DOWN (↓)', '메인(사진 목록)',
     '사진 탭에서 D-pad ↓ 반복 입력',
     '사진 항목 간 포커스 아래로 이동 + 스크롤',
     '보라색 포커스 테두리 이동, 스크롤 동작', 'PASS'),
    ('HK-G02', 'Gallery', 'KEY_UP (↑)', '메인(사진 목록)',
     '포커스가 아래에 있는 상태에서 D-pad ↑ 입력',
     '포커스 위로 이동',
     '포커스 상단으로 이동됨', 'PASS'),
    ('HK-G03', 'Gallery', 'KEY_RIGHT (→)', '메인(사진 목록)',
     '사진 목록에서 D-pad → 입력',
     '같은 행의 오른쪽 항목으로 포커스 이동',
     '오른쪽 항목으로 포커스 이동', 'PASS'),
    ('HK-G04', 'Gallery', 'KEY_LEFT (←)', '메인(사진 목록)',
     '포커스가 오른쪽에 있는 상태에서 D-pad ← 입력',
     '왼쪽 항목으로 포커스 이동',
     '왼쪽 항목으로 이동', 'PASS'),
    ('HK-G05', 'Gallery', 'KEY_ENTER (OK)', '메인(사진 포커스)',
     '사진 항목에 포커스 → ENTER 키 입력',
     '사진 상세보기 화면 진입',
     '상세보기(DetailActivity) 진입됨', 'PASS'),
    ('HK-G06', 'Gallery', 'KEY_LEFT (←)', '상세보기',
     '사진 상세보기에서 D-pad ← 입력',
     '이전 화면(목록)으로 복귀',
     '메인 사진 목록으로 복귀됨', 'PASS'),
    ('HK-G07', 'Gallery', 'KEY_RIGHT (→)', '상세보기',
     '사진 상세보기에서 D-pad → 입력',
     '다음 사진으로 이동 또는 반응 없음',
     '다음 사진(SW 업데이트 스크린샷)으로 정상 이동', 'PASS'),
    ('HK-G08', 'Gallery', 'KEY_BACK', '상세보기',
     '상세보기에서 BACK 키 입력',
     '이전 화면(목록)으로 복귀',
     '메인 사진 목록으로 정상 복귀 (detail→list 네비게이션)', 'PASS'),
    ('HK-G09', 'Gallery', 'KEY_BACKSPACE', '상세보기',
     '상세보기에서 BACKSPACE(취소) 키 입력',
     '이전 화면으로 복귀',
     '메인 사진 목록으로 복귀 (BACK과 동일 동작)', 'PASS'),
    ('HK-G10', 'Gallery', 'VOLUME UP/DOWN', '상세보기',
     '상세보기에서 볼륨 상/하 키 입력',
     '시스템 볼륨 조절',
     '볼륨 바 표시, 볼륨 조절됨', 'PASS'),
    ('HK-G11', 'Gallery', 'KEY_LEFT/RIGHT', '하단 탭 바',
     '하단 네비게이션(사진/앨범/휴지통)에서 ← → 키 입력',
     '탭 간 포커스 이동',
     '사진→앨범→휴지통 포커스 이동됨', 'PASS'),
    ('HK-G12', 'Gallery', 'KEY_ENTER', '하단 탭 바',
     '앨범 탭 포커스 상태에서 ENTER 입력',
     '앨범 탭 활성화',
     'D-pad로 하단 탭 포커스 도달 불가, ENTER로 탭 전환 미지원', 'FAIL'),
    # ── Memosy 하드키 ──
    ('HK-M01', 'Memosy', 'KEY_DOWN (↓)', '메인(메모 목록)',
     '메모 목록에서 D-pad ↓ 입력',
     '메모 카드에 포커스 이동',
     '☰ → 메모 카드로 포커스 정상 이동 (파란 테두리 표시)', 'PASS'),
    ('HK-M02', 'Memosy', 'KEY_RIGHT (→)', '메인(메모 목록)',
     '메모 목록에서 D-pad → 반복 입력',
     '상단 버튼(검색/보기전환/+) 간 포커스 이동',
     '검색→목록보기→추가 버튼 순차 이동, 앱 내 유지', 'PASS'),
    ('HK-M03', 'Memosy', 'KEY_ENTER (OK)', '메인(메뉴 포커스)',
     'DOWN으로 햄버거 버튼 포커스 → ENTER 입력',
     '햄버거 메뉴 열림',
     '보관함/휴지통 메뉴 열림', 'PASS'),
    ('HK-M04', 'Memosy', 'KEY_BACK', '메인(메모 목록)',
     '메인 화면에서 BACK 키 입력',
     '앱 종료 또는 확인 다이얼로그',
     '앱 즉시 종료 (홈 화면)', 'PASS'),
    ('HK-M05', 'Memosy', 'VOLUME UP/DOWN', '메인(메모 목록)',
     '메모 목록에서 볼륨 상/하 키 입력',
     '시스템 볼륨 조절',
     '볼륨 바 표시, 볼륨 조절됨', 'PASS'),
    ('HK-M06', 'Memosy', 'KEY_1~9 (숫자)', '메인(메모 목록)',
     '메모 목록에서 숫자 키 입력',
     '반응 없음 또는 무시',
     '앱 내 무반응 (정상). 물리 키패드에서는 시스템 단축키로 SMS 실행 가능', 'PASS'),
]

for i, (tc_id, app, key, screen, proc, expected, actual, result) in enumerate(hk_tcs, 2):
    app_f = gallery_fill if app == 'Gallery' else memosy_fill
    vals = [tc_id, app, key, screen, proc, expected, actual, result]
    for c, v in enumerate(vals, 1):
        cell = ws5.cell(row=i, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align
    ws5.cell(row=i, column=2).fill = app_f
    result_cell = ws5.cell(row=i, column=8)
    result_cell.alignment = center_align
    if result == 'PASS':
        result_cell.fill = pass_fill
    elif result == 'FAIL':
        result_cell.fill = fail_fill

ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 10
ws5.column_dimensions['C'].width = 22
ws5.column_dimensions['D'].width = 18
ws5.column_dimensions['E'].width = 40
ws5.column_dimensions['F'].width = 30
ws5.column_dimensions['G'].width = 40
ws5.column_dimensions['H'].width = 8


# ══════════════════════════════════════════
# Sheet 6: 하드키 매핑
# ══════════════════════════════════════════
ws6 = wb.create_sheet('하드키 매핑')
write_header(ws6, ['물리 키', 'Linux Key Code', 'Android KeyEvent', '입력 장치', '비고'])

keymap = [
    ('방향키 ↑', 'KEY_UP', 'KEYCODE_DPAD_UP', 'mtk-kpd', ''),
    ('방향키 ↓', 'KEY_DOWN', 'KEYCODE_DPAD_DOWN', 'mtk-kpd', ''),
    ('방향키 ←', 'KEY_LEFT', 'KEYCODE_DPAD_LEFT', 'mtk-kpd', ''),
    ('방향키 →', 'KEY_RIGHT', 'KEYCODE_DPAD_RIGHT', 'mtk-kpd', ''),
    ('가운데(OK)', 'KEY_ENTER', 'KEYCODE_ENTER', 'mtk-kpd', '선택/확인'),
    ('숫자 1', 'KEY_1', 'KEYCODE_1', 'mtk-kpd', ''),
    ('숫자 2', 'KEY_2', 'KEYCODE_2', 'mtk-kpd', ''),
    ('숫자 3', 'KEY_3', 'KEYCODE_3', 'mtk-kpd', ''),
    ('숫자 4', 'KEY_4', 'KEYCODE_4', 'mtk-kpd', ''),
    ('숫자 5', 'KEY_5', 'KEYCODE_5', 'mtk-kpd', ''),
    ('숫자 6', 'KEY_6', 'KEYCODE_6', 'mtk-kpd', ''),
    ('숫자 7', 'KEY_7', 'KEYCODE_7', 'mtk-kpd', ''),
    ('숫자 8', 'KEY_8', 'KEYCODE_8', 'mtk-kpd', ''),
    ('숫자 9', 'KEY_9', 'KEYCODE_9', 'mtk-kpd', ''),
    ('숫자 0', 'KEY_0', 'KEYCODE_0', 'mtk-kpd', ''),
    ('* (별)', 'KEY_NUMERIC_STAR', 'KEYCODE_STAR', 'mtk-kpd', ''),
    ('# (샵)', 'KEY_NUMERIC_POUND', 'KEYCODE_POUND', 'mtk-kpd', ''),
    ('통화', 'KEY_PHONE', 'KEYCODE_CALL', 'mtk-kpd', '전화 발신/수신'),
    ('취소/지우기', 'KEY_BACKSPACE', 'KEYCODE_DEL', 'mtk-kpd', ''),
    ('뒤로', 'KEY_BACK', 'KEYCODE_BACK', 'mtk-kpd', ''),
    ('홈', 'KEY_HOMEPAGE', 'KEYCODE_HOME', 'gpio-keys', ''),
    ('최근앱', 'KEY_APPSELECT', 'KEYCODE_APP_SWITCH', 'gpio-keys', ''),
    ('볼륨 ↑', 'KEY_VOLUMEUP', 'KEYCODE_VOLUME_UP', 'mtk-pmic-keys', ''),
    ('볼륨 ↓', 'KEY_VOLUMEDOWN', 'KEYCODE_VOLUME_DOWN', 'mtk-pmic-keys', ''),
]

for i, (phys, linux, android, device, note) in enumerate(keymap, 2):
    vals = [phys, linux, android, device, note]
    for c, v in enumerate(vals, 1):
        cell = ws6.cell(row=i, column=c, value=v)
        cell.border = thin_border
        cell.alignment = wrap_align

ws6.column_dimensions['A'].width = 14
ws6.column_dimensions['B'].width = 22
ws6.column_dimensions['C'].width = 26
ws6.column_dimensions['D'].width = 16
ws6.column_dimensions['E'].width = 20


# ── 테스트 환경 업데이트 (Sheet 3) ──
env_row = ws3.max_row + 2
ws3.cell(row=env_row, column=1, value='[하드키 추가 정보]').font = Font(bold=True, size=11)
extra_env = [
    ('단말 유형', '스타일폴더2 (플립폰)'),
    ('입력 장치', 'gpio-keys, mtk-kpd, mtk-pmic-keys'),
    ('하드키 TC 수', f'{len(hk_tcs)}건 (Gallery 12 + Memosy 6)'),
    ('하드키 PASS', f'{sum(1 for t in hk_tcs if t[7]=="PASS")}건'),
    ('하드키 FAIL', f'{sum(1 for t in hk_tcs if t[7]=="FAIL")}건'),
]
for j, (k, v) in enumerate(extra_env, env_row + 1):
    ws3.cell(row=j, column=1, value=k).border = thin_border
    ws3.cell(row=j, column=1).font = Font(bold=True)
    ws3.cell(row=j, column=2, value=v).border = thin_border


# ── 저장 ──
out = 'reports/App_Feature_TC_Report.xlsx'
wb.save(out)

hk_pass = sum(1 for t in hk_tcs if t[7] == 'PASS')
hk_fail = sum(1 for t in hk_tcs if t[7] == 'FAIL')
total_tc = len(tcs) + len(hk_tcs)
print(f'생성 완료: {out}')
print(f'TC 수: {total_tc}건 (터치 28 + 하드키 {len(hk_tcs)})')
print(f'  터치 TC: 28건 PASS')
print(f'  하드키 TC: {hk_pass}건 PASS, {hk_fail}건 FAIL')
print(f'스텝 수: {len(steps)}건')
print('시트: TC 요약 / 스텝 상세 / 테스트 환경 / 기능 커버리지 / 하드키 테스트 / 하드키 매핑')
